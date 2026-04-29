#!/usr/bin/env python3
"""
LAMMPS data viewer/editor for atom-centric inspection.

Features
--------
- Read a LAMMPS data file (Atoms/Bonds/Angles parsing focused on common atom styles)
- 3D visualization with PySide6 + PyVistaQt
- Default visual style: ball-and-stick-like spheres + bonds
- Mouse bindings:
  * Right-drag: rotate model
  * Middle-drag: pan model
  * Wheel: zoom
  * Left-click: pick atom, show id/type, and highlight selection
  * Alt + double left-click: select all atoms of the same type
- Selection tools:
  * Hide selected atoms
  * Show only selected atoms
  * Show all atoms
  * Select fragment(s) containing selected atom(s)
  * Change selected atom type
  * Copy selected atom ids
- Save modified atom types back to a new data file

Notes
-----
This is a practical desktop viewer/editor, not a full Materials Studio clone.
It focuses on accurate atom picking, selection, visibility control, fragment query,
and light editing (atom type reassignment + save-as).
"""

from __future__ import annotations

import argparse
import fnmatch
import json
import math
import os
import re
import shutil
import shlex
import subprocess
import sys
import time
from collections import defaultdict, deque
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import numpy as np

try:
    import yaml
except ImportError:  # pragma: no cover - optional dependency path
    yaml = None

os.environ.setdefault("QT_API", "pyside6")

from PySide6 import QtCore, QtGui, QtWidgets

pv = None
QtInteractor = None
vtkCellPicker = None


def ensure_3d_imports() -> None:
    global pv, QtInteractor, vtkCellPicker
    if pv is not None and QtInteractor is not None and vtkCellPicker is not None:
        return
    import pyvista as _pv
    from pyvistaqt import QtInteractor as _QtInteractor
    from vtkmodules.vtkRenderingCore import vtkCellPicker as _vtkCellPicker

    pv = _pv
    QtInteractor = _QtInteractor
    vtkCellPicker = _vtkCellPicker


SECTION_HEADERS = {
    "Masses",
    "Atoms",
    "Velocities",
    "Bonds",
    "Angles",
    "Dihedrals",
    "Impropers",
    "Pair Coeffs",
    "Bond Coeffs",
    "Angle Coeffs",
    "Dihedral Coeffs",
    "Improper Coeffs",
    "BondBond Coeffs",
    "BondAngle Coeffs",
    "MiddleBondTorsion Coeffs",
    "EndBondTorsion Coeffs",
    "AngleTorsion Coeffs",
    "AngleAngleTorsion Coeffs",
    "BondBond13 Coeffs",
    "AngleAngle Coeffs",
    "Ellipsoids",
    "Lines",
    "Triangles",
    "Bodies",
    "CS-Info",
}

TOPOLOGY_KIND_TO_SECTION = {
    "bond": "Bonds",
    "angle": "Angles",
    "dihedral": "Dihedrals",
    "improper": "Impropers",
}
SECTION_TO_TOPOLOGY_KIND = {value: key for key, value in TOPOLOGY_KIND_TO_SECTION.items()}
TOPOLOGY_TYPE_ATTR = {
    "bond": "bond_type",
    "angle": "angle_type",
    "dihedral": "dihedral_type",
    "improper": "improper_type",
}
SECTION_COUNT_LABELS = {
    "Atoms": "atoms",
    "Bonds": "bonds",
    "Angles": "angles",
    "Dihedrals": "dihedrals",
    "Impropers": "impropers",
}
COEFF_SECTION_TYPE_KIND = {
    "Pair Coeffs": "atom",
    "Bond Coeffs": "bond",
    "Angle Coeffs": "angle",
    "Dihedral Coeffs": "dihedral",
    "Improper Coeffs": "improper",
    "BondBond Coeffs": "angle",
    "BondAngle Coeffs": "angle",
    "MiddleBondTorsion Coeffs": "dihedral",
    "EndBondTorsion Coeffs": "dihedral",
    "AngleTorsion Coeffs": "dihedral",
    "AngleAngleTorsion Coeffs": "dihedral",
    "BondBond13 Coeffs": "dihedral",
    "AngleAngle Coeffs": "improper",
}
PRIMARY_COEFF_SECTION_TYPE_KIND = {
    "Bond Coeffs": "bond",
    "Angle Coeffs": "angle",
    "Dihedral Coeffs": "dihedral",
    "Improper Coeffs": "improper",
}


def parse_integer_selector(text: str) -> Set[int]:
    """Parse selectors like '1,3~5' into a set of integers."""
    cleaned = re.sub(r"\s+", "", text.replace("，", ","))
    if not cleaned:
        raise ValueError("empty selector")
    values: Set[int] = set()
    for part in cleaned.split(","):
        if not part:
            raise ValueError("empty selector part")
        if "~" in part:
            pieces = part.split("~")
            if len(pieces) != 2 or not pieces[0] or not pieces[1]:
                raise ValueError("invalid range selector")
            start = int(pieces[0])
            end = int(pieces[1])
            low, high = sorted((start, end))
            values.update(range(low, high + 1))
        else:
            values.add(int(part))
    return values


def spherical_offset(radius: float, theta_deg: float, phi_deg: float) -> Tuple[float, float, float]:
    theta = math.radians(float(theta_deg))
    phi = math.radians(float(phi_deg))
    r = float(radius)
    return (
        r * math.sin(theta) * math.cos(phi),
        r * math.sin(theta) * math.sin(phi),
        r * math.cos(theta),
    )


FORCEFIELD_FILE_SUFFIXES = {".yaml", ".yml", ".json", ".xlsx"}
MSI2LMP_INPUT_SUFFIXES = {".car", ".cor", ".mdf"}


def _safe_forcefield_filename(text: str) -> str:
    cleaned = re.sub(r"[^\w.\-]+", "_", text.strip(), flags=re.UNICODE).strip("._")
    return cleaned or "forcefield"


def _parse_forcefield_number(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return None
    text = text.replace("−", "-")
    match = re.fullmatch(r"([-+]?\d+(?:\.\d+)?)\s*x\s*10\s*([-+]?\d+)", text, flags=re.IGNORECASE)
    if match:
        return float(match.group(1)) * (10 ** int(match.group(2)))
    try:
        return float(text)
    except ValueError:
        return value


@dataclass
class ForceFieldDefinition:
    name: str
    version: str = ""
    units: str = "real"
    source_path: str = ""
    atom_types: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    bond_types: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    angle_types: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    dihedral_types: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    improper_types: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def display_name(self) -> str:
        return f"{self.name} ({self.version})" if self.version else self.name

    def to_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "name": self.name,
            "version": self.version,
            "units": self.units,
            "metadata": dict(self.metadata),
            "atom_types": self.atom_types,
            "bond_types": self.bond_types,
            "angle_types": self.angle_types,
            "dihedral_types": self.dihedral_types,
            "improper_types": self.improper_types,
        }
        return data

    @classmethod
    def from_dict(cls, data: Dict[str, Any], source_path: Path) -> "ForceFieldDefinition":
        name = str(data.get("name") or data.get("forcefield") or source_path.stem)
        return cls(
            name=name,
            version=str(data.get("version") or ""),
            units=str(data.get("units") or "real"),
            source_path=str(source_path),
            atom_types=dict(data.get("atom_types") or {}),
            bond_types=dict(data.get("bond_types") or {}),
            angle_types=dict(data.get("angle_types") or {}),
            dihedral_types=dict(data.get("dihedral_types") or {}),
            improper_types=dict(data.get("improper_types") or {}),
            metadata=dict(data.get("metadata") or {}),
        )


def _forcefield_entry_element(symbol: str, entry: Dict[str, Any]) -> str:
    explicit = entry.get("element")
    if explicit:
        return str(explicit).strip()
    species = str(entry.get("species") or "").lower()
    symbol_l = str(symbol).lower()
    species_map = [
        ("hydrogen", "H"),
        ("oxygen", "O"),
        ("silicon", "Si"),
        ("aluminum", "Al"),
        ("aluminium", "Al"),
        ("magnesium", "Mg"),
        ("calcium", "Ca"),
        ("iron", "Fe"),
        ("sodium", "Na"),
        ("potassium", "K"),
        ("chlor", "Cl"),
        ("sulfur", "S"),
        ("nitrogen", "N"),
        ("uranium", "U"),
        ("barium", "Ba"),
        ("strontium", "Sr"),
        ("cesium", "Cs"),
    ]
    for marker, element in species_map:
        if marker in species:
            return element
    symbol_map = {
        "h": "H",
        "ho": "H",
        "hn": "H",
        "o": "O",
        "ob": "O",
        "on": "O",
        "oh": "O",
        "os": "O",
        "ou": "O",
        "st": "Si",
        "si": "Si",
        "at": "Al",
        "ao": "Al",
        "al": "Al",
        "mgo": "Mg",
        "mgh": "Mg",
        "mg": "Mg",
        "cao": "Ca",
        "cah": "Ca",
        "ca": "Ca",
        "feo": "Fe",
        "fe": "Fe",
        "na": "Na",
        "k": "K",
        "cl": "Cl",
        "s": "S",
        "n": "N",
    }
    normalized = re.sub(r"[^a-z]", "", symbol_l)
    for prefix, element in sorted(symbol_map.items(), key=lambda item: len(item[0]), reverse=True):
        if normalized.startswith(prefix):
            return element
    return ""


def _format_forcefield_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _forcefield_params_text(entry: Dict[str, Any]) -> str:
    params = entry.get("params")
    if not isinstance(params, dict):
        pair = entry.get("pair")
        if isinstance(pair, dict):
            params = pair.get("params")
    if not isinstance(params, dict):
        return ""
    return ", ".join(f"{key}={_format_forcefield_value(value)}" for key, value in params.items())


def _atom_pair_coeff_tokens(entry: Dict[str, Any]) -> Optional[List[str]]:
    pair = entry.get("pair") if isinstance(entry.get("pair"), dict) else {}
    params = pair.get("params", {}) if isinstance(pair, dict) else {}
    epsilon = params.get("epsilon", params.get("eps", params.get("Do", params.get("D0"))))
    sigma = params.get("sigma", params.get("σ"))
    if epsilon is None or sigma is None:
        return None
    return [_format_forcefield_value(epsilon), _format_forcefield_value(sigma)]


def _topology_coeff_tokens(entry: Dict[str, Any], kind: str) -> Optional[List[str]]:
    params = entry.get("params", {})
    if not isinstance(params, dict):
        return None
    if kind == "bond":
        k_value = params.get("k", params.get("K", params.get("K1")))
        r0 = params.get("r0", params.get("R0"))
        if k_value is not None and r0 is not None:
            return [_format_forcefield_value(k_value), _format_forcefield_value(r0)]
    if kind == "angle":
        k_value = params.get("k", params.get("K", params.get("K2")))
        theta0 = params.get("theta0", params.get("theta", params.get("θ0")))
        if k_value is not None and theta0 is not None:
            return [_format_forcefield_value(k_value), _format_forcefield_value(theta0)]
    if params:
        return [_format_forcefield_value(value) for _, value in sorted(params.items())]
    return None


def load_forcefield_definition(path: Path) -> ForceFieldDefinition:
    suffix = path.suffix.lower()
    if suffix in {".yaml", ".yml"}:
        if yaml is None:
            raise RuntimeError("当前 Python 环境缺少 PyYAML，无法读取 YAML 力场文件。")
        with path.open("r", encoding="utf-8") as handle:
            data = yaml.safe_load(handle) or {}
        if not isinstance(data, dict):
            raise ValueError("YAML 力场文件顶层必须是字典结构。")
        return ForceFieldDefinition.from_dict(data, path)
    if suffix == ".json":
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        if not isinstance(data, dict):
            raise ValueError("JSON 力场文件顶层必须是字典结构。")
        return ForceFieldDefinition.from_dict(data, path)
    if suffix == ".xlsx":
        return load_forcefield_from_excel(path)
    raise ValueError(f"不支持的力场文件格式: {path.suffix}")


def load_forcefield_from_excel(path: Path) -> ForceFieldDefinition:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency path
        raise RuntimeError("当前 Python 环境缺少 openpyxl，无法读取 Excel 力场文件。") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    definition = ForceFieldDefinition(
        name=path.stem,
        version="",
        units="real",
        source_path=str(path),
        metadata={
            "source_format": "xlsx",
            "source_file": str(path),
            "sheets": list(workbook.sheetnames),
        },
    )

    if "L-J" in workbook.sheetnames:
        _parse_clayff_lj_sheet(workbook["L-J"], definition)
    if "Bond+Angle" in workbook.sheetnames:
        _parse_clayff_bond_angle_sheet(workbook["Bond+Angle"], definition)
    for sheet_name in workbook.sheetnames:
        normalized = sheet_name.strip().lower()
        if normalized in {"atom_types", "bond_types", "angle_types", "dihedral_types", "improper_types"}:
            _parse_general_forcefield_sheet(workbook[sheet_name], normalized, definition)

    if not any(
        [
            definition.atom_types,
            definition.bond_types,
            definition.angle_types,
            definition.dihedral_types,
            definition.improper_types,
        ]
    ):
        raise ValueError("Excel 文件中未识别到可用的力场参数表。")
    return definition


def _parse_clayff_lj_sheet(worksheet, definition: ForceFieldDefinition) -> None:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return
    for row in rows[1:]:
        species = row[0] if len(row) > 0 else None
        symbol = row[1] if len(row) > 1 else None
        if not symbol:
            continue
        symbol_text = str(symbol).strip()
        if not symbol_text:
            continue
        charge = _parse_forcefield_number(row[2] if len(row) > 2 else None)
        do_value = _parse_forcefield_number(row[3] if len(row) > 3 else None)
        ro_value = _parse_forcefield_number(row[4] if len(row) > 4 else None)
        sigma = _parse_forcefield_number(row[5] if len(row) > 5 else None)
        entry: Dict[str, Any] = {
            "species": str(species).strip() if species is not None else "",
            "element": _forcefield_entry_element(symbol_text, {"species": species or ""}),
            "charge": charge,
            "pair": {
                "style": "lj/cut/coul/long",
                "params": {
                    "Do": do_value,
                    "Ro": ro_value,
                    "sigma": sigma,
                },
            },
        }
        definition.atom_types[symbol_text] = entry


def _parse_clayff_bond_angle_sheet(worksheet, definition: ForceFieldDefinition) -> None:
    mode = None
    for row in worksheet.iter_rows(values_only=True):
        values = [value for value in row if value is not None and str(value).strip()]
        if not values:
            continue
        first = str(values[0]).strip().lower()
        if "bond stretch" in first:
            mode = "bond"
            continue
        if "angle bend" in first:
            mode = "angle"
            continue
        if first.startswith("species") or "注意" in first or "参数来源" in first:
            continue
        if mode == "bond" and len(values) >= 4:
            atom1, atom2 = str(values[0]).strip(), str(values[1]).strip()
            if not atom1 or not atom2:
                continue
            symbol = f"{atom1}-{atom2}"
            definition.bond_types[symbol] = {
                "atoms": [atom1, atom2],
                "style": "harmonic",
                "params": {
                    "k": _parse_forcefield_number(values[2]),
                    "r0": _parse_forcefield_number(values[3]),
                },
            }
        elif mode == "angle" and len(values) >= 5:
            atom1, atom2, atom3 = str(values[0]).strip(), str(values[1]).strip(), str(values[2]).strip()
            if not atom1 or not atom2 or not atom3:
                continue
            symbol = f"{atom1}-{atom2}-{atom3}"
            definition.angle_types[symbol] = {
                "atoms": [atom1, atom2, atom3],
                "style": "harmonic",
                "params": {
                    "k": _parse_forcefield_number(values[3]),
                    "theta0": _parse_forcefield_number(values[4]),
                },
            }


def _parse_general_forcefield_sheet(worksheet, sheet_kind: str, definition: ForceFieldDefinition) -> None:
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    target_map = {
        "atom_types": definition.atom_types,
        "bond_types": definition.bond_types,
        "angle_types": definition.angle_types,
        "dihedral_types": definition.dihedral_types,
        "improper_types": definition.improper_types,
    }
    target = target_map[sheet_kind]
    for row in rows[1:]:
        record = {
            header: row[index]
            for index, header in enumerate(headers)
            if header and index < len(row) and row[index] is not None
        }
        symbol = record.get("symbol") or record.get("ff-type") or record.get("name")
        if not symbol:
            continue
        symbol_text = str(symbol).strip()
        entry = target.setdefault(symbol_text, {})
        for key, value in record.items():
            if key in {"symbol", "ff-type"}:
                continue
            if key == "atoms" and isinstance(value, str):
                entry[key] = [part.strip() for part in re.split(r"[,;\s]+", value) if part.strip()]
            elif key == "param_name":
                param_name = str(value).strip()
                entry.setdefault("params", {})[param_name] = _parse_forcefield_number(record.get("param_value"))
            elif key != "param_value":
                entry[key] = _parse_forcefield_number(value)

MATERIAL_STUDIO_ELEMENT_COLORS: Dict[str, Tuple[int, int, int]] = {
    "H": (255, 255, 255),
    "He": (217, 255, 255),
    "Li": (204, 128, 255),
    "Be": (194, 255, 0),
    "B": (255, 181, 181),
    "C": (144, 144, 144),
    "N": (48, 80, 248),
    "O": (255, 13, 13),
    "F": (144, 224, 80),
    "Ne": (179, 227, 245),
    "Na": (171, 92, 242),
    "Mg": (138, 255, 0),
    "Al": (191, 166, 166),
    "Si": (240, 200, 160),
    "P": (255, 128, 0),
    "S": (255, 255, 48),
    "Cl": (31, 240, 31),
    "Ar": (128, 209, 227),
    "K": (143, 64, 212),
    "Ca": (61, 255, 0),
    "Ti": (191, 194, 199),
    "Fe": (224, 102, 51),
    "Cu": (200, 128, 51),
    "Zn": (125, 128, 176),
}

ELEMENT_MASS_TABLE: List[Tuple[str, str, float]] = [
    ("H", "Hydrogen", 1.008),
    ("He", "Helium", 4.0026),
    ("Li", "Lithium", 6.94),
    ("Be", "Beryllium", 9.0122),
    ("B", "Boron", 10.81),
    ("C", "Carbon", 12.011),
    ("N", "Nitrogen", 14.007),
    ("O", "Oxygen", 15.999),
    ("F", "Fluorine", 18.998),
    ("Ne", "Neon", 20.180),
    ("Na", "Sodium", 22.990),
    ("Mg", "Magnesium", 24.305),
    ("Al", "Aluminium", 26.982),
    ("Si", "Silicon", 28.085),
    ("P", "Phosphorus", 30.974),
    ("S", "Sulfur", 32.06),
    ("Cl", "Chlorine", 35.45),
    ("Ar", "Argon", 39.948),
    ("K", "Potassium", 39.098),
    ("Ca", "Calcium", 40.078),
    ("Sc", "Scandium", 44.956),
    ("Ti", "Titanium", 47.867),
    ("V", "Vanadium", 50.942),
    ("Cr", "Chromium", 51.996),
    ("Mn", "Manganese", 54.938),
    ("Fe", "Iron", 55.845),
    ("Co", "Cobalt", 58.933),
    ("Ni", "Nickel", 58.693),
    ("Cu", "Copper", 63.546),
    ("Zn", "Zinc", 65.38),
    ("Ga", "Gallium", 69.723),
    ("Ge", "Germanium", 72.630),
    ("As", "Arsenic", 74.922),
    ("Se", "Selenium", 78.971),
    ("Br", "Bromine", 79.904),
    ("Kr", "Krypton", 83.798),
    ("Rb", "Rubidium", 85.468),
    ("Sr", "Strontium", 87.62),
    ("Y", "Yttrium", 88.906),
    ("Zr", "Zirconium", 91.224),
    ("Nb", "Niobium", 92.906),
    ("Mo", "Molybdenum", 95.95),
    ("Ru", "Ruthenium", 101.07),
    ("Rh", "Rhodium", 102.905),
    ("Pd", "Palladium", 106.42),
    ("Ag", "Silver", 107.868),
    ("Cd", "Cadmium", 112.414),
    ("In", "Indium", 114.818),
    ("Sn", "Tin", 118.710),
    ("Sb", "Antimony", 121.760),
    ("Te", "Tellurium", 127.60),
    ("I", "Iodine", 126.904),
    ("Xe", "Xenon", 131.293),
    ("Cs", "Caesium", 132.905),
    ("Ba", "Barium", 137.327),
    ("La", "Lanthanum", 138.905),
    ("Ce", "Cerium", 140.116),
    ("Pr", "Praseodymium", 140.908),
    ("Nd", "Neodymium", 144.242),
    ("Sm", "Samarium", 150.36),
    ("Eu", "Europium", 151.964),
    ("Gd", "Gadolinium", 157.25),
    ("Tb", "Terbium", 158.925),
    ("Dy", "Dysprosium", 162.500),
    ("Ho", "Holmium", 164.930),
    ("Er", "Erbium", 167.259),
    ("Tm", "Thulium", 168.934),
    ("Yb", "Ytterbium", 173.045),
    ("Lu", "Lutetium", 174.967),
    ("Hf", "Hafnium", 178.49),
    ("Ta", "Tantalum", 180.948),
    ("W", "Tungsten", 183.84),
    ("Re", "Rhenium", 186.207),
    ("Os", "Osmium", 190.23),
    ("Ir", "Iridium", 192.217),
    ("Pt", "Platinum", 195.084),
    ("Au", "Gold", 196.967),
    ("Hg", "Mercury", 200.592),
    ("Tl", "Thallium", 204.38),
    ("Pb", "Lead", 207.2),
    ("Bi", "Bismuth", 208.980),
    ("Th", "Thorium", 232.038),
    ("U", "Uranium", 238.029),
]


@dataclass
class AtomRecord:
    uid: int
    atom_id: int
    atom_type: int
    x: float
    y: float
    z: float
    mol: Optional[int] = None
    charge: Optional[float] = None
    image: Tuple[int, int, int] = (0, 0, 0)
    raw_tokens: List[str] = field(default_factory=list)
    raw_comment: str = ""
    line_index: Optional[int] = None


@dataclass
class BondRecord:
    bond_id: int
    bond_type: int
    atom1: int
    atom2: int
    raw_comment: str = ""
    line_index: Optional[int] = None

    @property
    def record_id(self) -> int:
        return self.bond_id

    @property
    def record_type(self) -> int:
        return self.bond_type

    @property
    def atom_ids(self) -> Tuple[int, int]:
        return self.atom1, self.atom2


@dataclass
class AngleRecord:
    angle_id: int
    angle_type: int
    atom1: int
    atom2: int
    atom3: int
    raw_comment: str = ""
    line_index: Optional[int] = None

    @property
    def record_id(self) -> int:
        return self.angle_id

    @property
    def record_type(self) -> int:
        return self.angle_type

    @property
    def atom_ids(self) -> Tuple[int, int, int]:
        return self.atom1, self.atom2, self.atom3


@dataclass
class DihedralRecord:
    dihedral_id: int
    dihedral_type: int
    atom1: int
    atom2: int
    atom3: int
    atom4: int
    raw_comment: str = ""
    line_index: Optional[int] = None

    @property
    def record_id(self) -> int:
        return self.dihedral_id

    @property
    def record_type(self) -> int:
        return self.dihedral_type

    @property
    def atom_ids(self) -> Tuple[int, int, int, int]:
        return self.atom1, self.atom2, self.atom3, self.atom4


@dataclass
class ImproperRecord:
    improper_id: int
    improper_type: int
    atom1: int
    atom2: int
    atom3: int
    atom4: int
    raw_comment: str = ""
    line_index: Optional[int] = None

    @property
    def record_id(self) -> int:
        return self.improper_id

    @property
    def record_type(self) -> int:
        return self.improper_type

    @property
    def atom_ids(self) -> Tuple[int, int, int, int]:
        return self.atom1, self.atom2, self.atom3, self.atom4


@dataclass
class SectionBlock:
    name: str
    header_index: int
    trailing: str = ""
    lines: List[Tuple[int, str]] = field(default_factory=list)


@dataclass
class ParsedDataFile:
    file_path: Path
    atom_style: str
    atoms: Dict[int, AtomRecord]
    bonds: List[BondRecord]
    angles: List[AngleRecord]
    dihedrals: List[DihedralRecord]
    impropers: List[ImproperRecord]
    masses: Dict[int, float]
    mass_labels: Dict[int, str]
    coeff_labels: Dict[str, Dict[int, str]]
    coeff_values: Dict[str, Dict[int, List[str]]]
    cs_info: List[Tuple[int, int]]
    sections: Dict[str, List[Tuple[int, str]]]
    section_order: List[SectionBlock]
    original_lines: List[str]
    box: Dict[str, Tuple[float, float]]


class LammpsDataParser:
    def __init__(self, path: Path):
        self.path = Path(path)

    @staticmethod
    def _split_comment(line: str) -> Tuple[str, str]:
        if "#" in line:
            idx = line.index("#")
            return line[:idx].rstrip(), line[idx:].rstrip("\n")
        return line.rstrip("\n"), ""

    @staticmethod
    def _canonical_header(line: str) -> Optional[Tuple[str, str]]:
        stripped = line.strip()
        if not stripped:
            return None
        for header in sorted(SECTION_HEADERS, key=len, reverse=True):
            if stripped == header:
                return header, ""
            if stripped.startswith(header + " "):
                trailing = stripped[len(header):].strip()
                return header, trailing
        return None

    @staticmethod
    def _parse_box(lines: Sequence[str]) -> Dict[str, Tuple[float, float]]:
        box: Dict[str, Tuple[float, float]] = {}
        patterns = {
            "x": re.compile(r"^\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)\s+xlo\s+xhi"),
            "y": re.compile(r"^\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)\s+ylo\s+yhi"),
            "z": re.compile(r"^\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)\s+zlo\s+zhi"),
        }
        for line in lines[:80]:
            for axis, pat in patterns.items():
                m = pat.match(line)
                if m:
                    box[axis] = (float(m.group(1)), float(m.group(2)))
        return box

    def parse(self) -> ParsedDataFile:
        lines = self.path.read_text(encoding="utf-8", errors="replace").splitlines(True)
        box = self._parse_box(lines)
        sections: Dict[str, List[Tuple[int, str]]] = defaultdict(list)
        section_order: List[SectionBlock] = []
        section_trailing: Dict[str, str] = {}
        current_block: Optional[SectionBlock] = None

        for idx, raw in enumerate(lines):
            header = self._canonical_header(raw)
            if header:
                current_section, trailing = header
                section_trailing[current_section] = trailing
                current_block = SectionBlock(name=current_section, header_index=idx, trailing=trailing)
                section_order.append(current_block)
                continue
            if current_block is not None:
                sections[current_block.name].append((idx, raw))
                current_block.lines.append((idx, raw))

        atom_style = self._infer_atom_style(section_trailing.get("Atoms", ""), sections.get("Atoms", []))
        atoms = self._parse_atoms(atom_style, sections.get("Atoms", []), lines)
        bonds = self._parse_bonds(sections.get("Bonds", []))
        angles = self._parse_angles(sections.get("Angles", []))
        dihedrals = self._parse_dihedrals(sections.get("Dihedrals", []))
        impropers = self._parse_impropers(sections.get("Impropers", []))
        masses, mass_labels = self._parse_masses(sections.get("Masses", []))
        coeff_labels = self._parse_coeff_labels(sections)
        coeff_values = self._parse_coeff_values(sections)
        cs_info = self._parse_cs_info(sections.get("CS-Info", []))

        return ParsedDataFile(
            file_path=self.path,
            atom_style=atom_style,
            atoms=atoms,
            bonds=bonds,
            angles=angles,
            dihedrals=dihedrals,
            impropers=impropers,
            masses=masses,
            mass_labels=mass_labels,
            coeff_labels=coeff_labels,
            coeff_values=coeff_values,
            cs_info=cs_info,
            sections=sections,
            section_order=section_order,
            original_lines=lines,
            box=box,
        )

    def _infer_atom_style(self, trailing: str, atom_lines: List[Tuple[int, str]]) -> str:
        trailing_lower = trailing.lower()
        if trailing_lower.startswith("#"):
            trailing_lower = trailing_lower[1:].strip()
        for style in ("full", "charge", "molecular", "atomic"):
            if style in trailing_lower:
                return style

        for _, raw in atom_lines:
            main, _ = self._split_comment(raw)
            if not main.strip():
                continue
            n = len(main.split())
            if n >= 7:
                return "full"
            if n == 6:
                return "charge"
            if n == 5:
                return "atomic"
        return "full"

    def _parse_atoms(
        self, atom_style: str, atom_lines: List[Tuple[int, str]], original_lines: List[str]
    ) -> Dict[int, AtomRecord]:
        atoms: Dict[int, AtomRecord] = {}
        for idx, raw in atom_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            atom = self._make_atom(atom_style, tokens)
            atom.raw_tokens = tokens[:]
            atom.raw_comment = comment
            atom.line_index = idx
            atoms[atom.atom_id] = atom
        if not atoms:
            raise ValueError("未能从 Atoms section 解析到原子。")
        return atoms

    def _make_atom(self, atom_style: str, tokens: List[str]) -> AtomRecord:
        if atom_style == "full":
            if len(tokens) < 7:
                raise ValueError(f"full 风格原子行字段不足: {' '.join(tokens)}")
            atom_id = int(tokens[0])
            mol = int(tokens[1])
            atom_type = int(tokens[2])
            charge = float(tokens[3])
            x, y, z = map(float, tokens[4:7])
            image = tuple(map(int, tokens[7:10])) if len(tokens) >= 10 else (0, 0, 0)
            return AtomRecord(atom_id, atom_id, atom_type, x, y, z, mol=mol, charge=charge, image=image)

        if atom_style == "charge":
            if len(tokens) < 6:
                raise ValueError(f"charge 风格原子行字段不足: {' '.join(tokens)}")
            atom_id = int(tokens[0])
            atom_type = int(tokens[1])
            charge = float(tokens[2])
            x, y, z = map(float, tokens[3:6])
            image = tuple(map(int, tokens[6:9])) if len(tokens) >= 9 else (0, 0, 0)
            return AtomRecord(atom_id, atom_id, atom_type, x, y, z, charge=charge, image=image)

        if atom_style == "molecular":
            if len(tokens) < 6:
                raise ValueError(f"molecular 风格原子行字段不足: {' '.join(tokens)}")
            atom_id = int(tokens[0])
            mol = int(tokens[1])
            atom_type = int(tokens[2])
            x, y, z = map(float, tokens[3:6])
            image = tuple(map(int, tokens[6:9])) if len(tokens) >= 9 else (0, 0, 0)
            return AtomRecord(atom_id, atom_id, atom_type, x, y, z, mol=mol, image=image)

        if atom_style == "atomic":
            if len(tokens) < 5:
                raise ValueError(f"atomic 风格原子行字段不足: {' '.join(tokens)}")
            atom_id = int(tokens[0])
            atom_type = int(tokens[1])
            x, y, z = map(float, tokens[2:5])
            image = tuple(map(int, tokens[5:8])) if len(tokens) >= 8 else (0, 0, 0)
            return AtomRecord(atom_id, atom_id, atom_type, x, y, z, image=image)

        raise ValueError(f"暂不支持的 atom_style: {atom_style}")

    def _parse_bonds(self, bond_lines: List[Tuple[int, str]]) -> List[BondRecord]:
        bonds: List[BondRecord] = []
        for idx, raw in bond_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 4:
                continue
            bonds.append(BondRecord(int(tokens[0]), int(tokens[1]), int(tokens[2]), int(tokens[3]), raw_comment=comment, line_index=idx))
        return bonds

    def _parse_angles(self, angle_lines: List[Tuple[int, str]]) -> List[AngleRecord]:
        angles: List[AngleRecord] = []
        for idx, raw in angle_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 5:
                continue
            angles.append(
                AngleRecord(
                    int(tokens[0]),
                    int(tokens[1]),
                    int(tokens[2]),
                    int(tokens[3]),
                    int(tokens[4]),
                    raw_comment=comment,
                    line_index=idx,
                )
            )
        return angles

    def _parse_dihedrals(self, dihedral_lines: List[Tuple[int, str]]) -> List[DihedralRecord]:
        dihedrals: List[DihedralRecord] = []
        for idx, raw in dihedral_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 6:
                continue
            dihedrals.append(
                DihedralRecord(
                    int(tokens[0]),
                    int(tokens[1]),
                    int(tokens[2]),
                    int(tokens[3]),
                    int(tokens[4]),
                    int(tokens[5]),
                    raw_comment=comment,
                    line_index=idx,
                )
            )
        return dihedrals

    def _parse_impropers(self, improper_lines: List[Tuple[int, str]]) -> List[ImproperRecord]:
        impropers: List[ImproperRecord] = []
        for idx, raw in improper_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 6:
                continue
            impropers.append(
                ImproperRecord(
                    int(tokens[0]),
                    int(tokens[1]),
                    int(tokens[2]),
                    int(tokens[3]),
                    int(tokens[4]),
                    int(tokens[5]),
                    raw_comment=comment,
                    line_index=idx,
                )
            )
        return impropers

    def _parse_masses(self, mass_lines: List[Tuple[int, str]]) -> Tuple[Dict[int, float], Dict[int, str]]:
        masses: Dict[int, float] = {}
        mass_labels: Dict[int, str] = {}
        for _, raw in mass_lines:
            main, comment = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 2:
                continue
            try:
                atom_type = int(tokens[0])
                masses[atom_type] = float(tokens[1])
                mass_labels[atom_type] = comment[1:].strip() if comment.startswith("#") else comment.strip()
            except ValueError:
                continue
        return masses, mass_labels

    def _parse_coeff_labels(self, sections: Dict[str, List[Tuple[int, str]]]) -> Dict[str, Dict[int, str]]:
        section_to_kind = {
            "Pair Coeffs": "atom",
            "Bond Coeffs": "bond",
            "Angle Coeffs": "angle",
            "Dihedral Coeffs": "dihedral",
            "Improper Coeffs": "improper",
        }
        labels: Dict[str, Dict[int, str]] = {kind: {} for kind in section_to_kind.values()}
        for section_name, kind in section_to_kind.items():
            for _, raw in sections.get(section_name, []):
                main, comment = self._split_comment(raw)
                if not main.strip() or not comment.strip():
                    continue
                tokens = main.split()
                try:
                    type_id = int(tokens[0])
                except (ValueError, IndexError):
                    continue
                labels[kind][type_id] = comment[1:].strip() if comment.startswith("#") else comment.strip()
        return labels

    def _parse_coeff_values(self, sections: Dict[str, List[Tuple[int, str]]]) -> Dict[str, Dict[int, List[str]]]:
        section_to_kind = {
            "Pair Coeffs": "atom",
            "Bond Coeffs": "bond",
            "Angle Coeffs": "angle",
            "Dihedral Coeffs": "dihedral",
            "Improper Coeffs": "improper",
        }
        values: Dict[str, Dict[int, List[str]]] = {kind: {} for kind in section_to_kind.values()}
        for section_name, kind in section_to_kind.items():
            for _, raw in sections.get(section_name, []):
                main, _ = self._split_comment(raw)
                if not main.strip():
                    continue
                tokens = main.split()
                try:
                    type_id = int(tokens[0])
                except (ValueError, IndexError):
                    continue
                values[kind][type_id] = tokens[1:]
        return values

    def _parse_cs_info(self, cs_lines: List[Tuple[int, str]]) -> List[Tuple[int, int]]:
        records: List[Tuple[int, int]] = []
        for _, raw in cs_lines:
            main, _ = self._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            if len(tokens) < 2:
                continue
            try:
                records.append((int(tokens[0]), int(tokens[1])))
            except ValueError:
                continue
        return records


class DataModel:
    def __init__(self, parsed: ParsedDataFile):
        self.parsed = parsed
        self.atoms = dict(sorted(parsed.atoms.items()))
        self.bonds = sorted(parsed.bonds, key=lambda item: item.bond_id)
        self.angles = sorted(parsed.angles, key=lambda item: item.angle_id)
        self.dihedrals = sorted(parsed.dihedrals, key=lambda item: item.dihedral_id)
        self.impropers = sorted(parsed.impropers, key=lambda item: item.improper_id)
        self.masses = parsed.masses.copy()
        self.mass_names = parsed.mass_labels.copy()
        self.mass_labels = parsed.mass_labels.copy()
        self.coeff_names = {kind: labels.copy() for kind, labels in parsed.coeff_labels.items()}
        self.coeff_labels = {kind: labels.copy() for kind, labels in parsed.coeff_labels.items()}
        self.coeff_values = {kind: {type_id: list(tokens) for type_id, tokens in values.items()} for kind, values in parsed.coeff_values.items()}
        for kind in ("atom", "bond", "angle", "dihedral", "improper"):
            if kind != "atom":
                self.coeff_names.setdefault(kind, {})
            self.coeff_labels.setdefault(kind, {})
            self.coeff_values.setdefault(kind, {})
        self._seed_topology_labels_from_record_comments()
        self.type_colors: Dict[int, Tuple[int, int, int]] = {}
        self.cs_info = list(parsed.cs_info)
        self.cs_info_uids: List[Tuple[int, int]] = []
        self.atom_style = parsed.atom_style
        self.original_lines = parsed.original_lines
        self.dirty = False
        self.hidden_atoms: Set[int] = set()
        self.hidden_atom_uids: Set[int] = set()
        self.selection_kind: Optional[str] = None
        self.selected_atoms: Set[int] = set()
        self.selected_atom_uids: Set[int] = set()
        self.selected_bonds: Set[int] = set()
        self.selected_angles: Set[int] = set()
        self.selected_dihedrals: Set[int] = set()
        self.selected_impropers: Set[int] = set()
        self.type_id_maps = self._initial_type_id_maps()
        self.fragments, self.fragment_of_atom = self._compute_fragments()
        self._set_cs_info_uids_from_atom_ids(self.cs_info)

    def _next_atom_uid(self) -> int:
        return max((atom.uid for atom in self.atoms.values()), default=0) + 1

    def atom_uid_map(self) -> Dict[int, int]:
        return {atom.uid: atom.atom_id for atom in self.atoms.values()}

    def atom_id_to_uid_map(self) -> Dict[int, int]:
        return {atom.atom_id: atom.uid for atom in self.atoms.values()}

    def atom_id_from_uid(self, uid: int) -> Optional[int]:
        return self.atom_uid_map().get(int(uid))

    def atom_uid_from_id(self, atom_id: int) -> Optional[int]:
        atom = self.atoms.get(int(atom_id))
        return None if atom is None else atom.uid

    def _set_cs_info_uids_from_atom_ids(self, cs_info: Sequence[Tuple[int, int]]) -> None:
        self.cs_info_uids = []
        for atom_id, model_id in cs_info:
            atom = self.atoms.get(int(atom_id))
            if atom is not None:
                self.cs_info_uids.append((atom.uid, int(model_id)))
        self._sync_cs_info_ids_from_uids()

    def _sync_cs_info_ids_from_uids(self) -> None:
        uid_to_id = self.atom_uid_map()
        self.cs_info = [
            (uid_to_id[uid], model_id)
            for uid, model_id in self.cs_info_uids
            if uid in uid_to_id
        ]

    def _atom_ids_to_uids(self, atom_ids: Set[int]) -> Set[int]:
        return {self.atoms[atom_id].uid for atom_id in atom_ids if atom_id in self.atoms}

    def _atom_uids_to_ids(self, atom_uids: Set[int]) -> Set[int]:
        uid_to_id = self.atom_uid_map()
        return {uid_to_id[uid] for uid in atom_uids if uid in uid_to_id}

    def _set_selected_atom_ids(self, atom_ids: Set[int]) -> None:
        self.selected_atoms = {atom_id for atom_id in atom_ids if atom_id in self.atoms}
        self.selected_atom_uids = self._atom_ids_to_uids(self.selected_atoms)

    def _sync_selected_atom_ids_from_uids(self) -> None:
        self.selected_atoms = self._atom_uids_to_ids(self.selected_atom_uids)
        if not self.selected_atoms and self.selection_kind == "atom":
            self.selection_kind = None

    def _set_hidden_atom_ids(self, atom_ids: Set[int]) -> None:
        self.hidden_atoms = {atom_id for atom_id in atom_ids if atom_id in self.atoms}
        self.hidden_atom_uids = self._atom_ids_to_uids(self.hidden_atoms)

    def _sync_hidden_atom_ids_from_uids(self) -> None:
        self.hidden_atoms = self._atom_uids_to_ids(self.hidden_atom_uids)

    def _cs_pair_groups(self) -> Dict[int, List[int]]:
        groups: Dict[int, List[int]] = defaultdict(list)
        for atom_id, model_id in self.cs_info:
            if atom_id in self.atoms and atom_id not in groups[model_id]:
                groups[model_id].append(atom_id)
        return groups

    def _is_shell_atom(self, atom_id: int) -> bool:
        atom = self.atoms.get(atom_id)
        if atom is None:
            return False
        label = self.mass_labels.get(atom.atom_type, "")
        return "shell" in label.lower()

    def core_shell_maps(self) -> Tuple[Dict[int, int], Dict[int, int]]:
        core_to_shell: Dict[int, int] = {}
        shell_to_core: Dict[int, int] = {}
        for atom_ids in self._cs_pair_groups().values():
            if len(atom_ids) != 2:
                continue
            shell_candidates = [atom_id for atom_id in atom_ids if self._is_shell_atom(atom_id)]
            if len(shell_candidates) == 1:
                shell_id = shell_candidates[0]
                core_id = atom_ids[0] if atom_ids[1] == shell_id else atom_ids[1]
            else:
                core_id, shell_id = atom_ids[0], atom_ids[1]
            if core_id in self.atoms and shell_id in self.atoms and core_id != shell_id:
                core_to_shell[core_id] = shell_id
                shell_to_core[shell_id] = core_id
        return core_to_shell, shell_to_core

    def shell_for_core(self, atom_id: int) -> Optional[int]:
        core_to_shell, _ = self.core_shell_maps()
        return core_to_shell.get(int(atom_id))

    def core_for_shell(self, atom_id: int) -> Optional[int]:
        _, shell_to_core = self.core_shell_maps()
        return shell_to_core.get(int(atom_id))

    def _replace_atoms_in_record(self, record: object, replacements: Dict[int, int]) -> int:
        fields = [field for field in ("atom1", "atom2", "atom3", "atom4") if hasattr(record, field)]
        original_ids = [int(getattr(record, field)) for field in fields]
        changed = 0
        for field, atom_id in zip(fields, original_ids):
            shell_id = replacements.get(atom_id)
            if shell_id is None:
                continue
            if shell_id in original_ids:
                continue
            setattr(record, field, shell_id)
            changed += 1
        return changed

    def replace_selected_topology_cores_with_shells(self) -> Dict[str, int]:
        if self.selection_kind not in TOPOLOGY_KIND_TO_SECTION:
            return {"records": 0, "atoms": 0, "skipped": 0}
        kind = self.selection_kind
        selected_ids = set(self.topology_id_set(kind))
        core_to_shell, _ = self.core_shell_maps()
        changed_records = 0
        changed_atoms = 0
        skipped_records = 0
        for record in self.topology_records(kind):
            if record.record_id not in selected_ids:
                continue
            before = tuple(record.atom_ids)
            changed = self._replace_atoms_in_record(record, core_to_shell)
            if changed:
                changed_records += 1
                changed_atoms += changed
            elif any(atom_id in core_to_shell for atom_id in before):
                skipped_records += 1
        if changed_records:
            self.fragments, self.fragment_of_atom = self._compute_fragments()
            self.dirty = True
        return {"records": changed_records, "atoms": changed_atoms, "skipped": skipped_records}

    def replace_core_with_shell_everywhere(self, shell_id: int) -> Dict[str, int]:
        core_id = self.core_for_shell(shell_id)
        if core_id is None:
            return {"records": 0, "atoms": 0, "skipped": 0}
        replacements = {core_id: int(shell_id)}
        changed_records = 0
        changed_atoms = 0
        skipped_records = 0
        for kind in ("bond", "angle", "dihedral", "improper"):
            for record in self.topology_records(kind):
                before = tuple(record.atom_ids)
                changed = self._replace_atoms_in_record(record, replacements)
                if changed:
                    changed_records += 1
                    changed_atoms += changed
                elif core_id in before:
                    skipped_records += 1
        if changed_records:
            self.fragments, self.fragment_of_atom = self._compute_fragments()
            self.dirty = True
        return {"records": changed_records, "atoms": changed_atoms, "skipped": skipped_records}

    def _seed_topology_labels_from_record_comments(self) -> None:
        for kind in ("bond", "angle", "dihedral", "improper"):
            names = self.coeff_names.setdefault(kind, {})
            labels = self.coeff_labels.setdefault(kind, {})
            for record in self.topology_records(kind):
                comment = record.raw_comment.strip()
                if comment.startswith("#"):
                    comment = comment[1:].strip()
                if comment:
                    names.setdefault(record.record_type, comment)
                    labels.setdefault(record.record_type, comment)

    def _initial_type_id_maps(self) -> Dict[str, Dict[int, int]]:
        return {
            "atom": {type_id: type_id for type_id in sorted(set(self.masses) | {atom.atom_type for atom in self.atoms.values()})},
            "bond": {type_id: type_id for type_id in sorted({bond.bond_type for bond in self.bonds})},
            "angle": {type_id: type_id for type_id in sorted({angle.angle_type for angle in self.angles})},
            "dihedral": {type_id: type_id for type_id in sorted({record.dihedral_type for record in self.dihedrals})},
            "improper": {type_id: type_id for type_id in sorted({record.improper_type for record in self.impropers})},
        }

    def _bond_adjacency(self) -> Dict[int, Set[int]]:
        adjacency: Dict[int, Set[int]] = {aid: set() for aid in self.atoms}
        for bond in self.bonds:
            if bond.atom1 not in self.atoms or bond.atom2 not in self.atoms:
                continue
            adjacency[bond.atom1].add(bond.atom2)
            adjacency[bond.atom2].add(bond.atom1)
        return adjacency

    def bonded_component_atoms(self, seed_atom_ids: Set[int]) -> Set[int]:
        adjacency = self._bond_adjacency()
        seed_ids = {aid for aid in seed_atom_ids if aid in adjacency}
        if not seed_ids:
            return set()

        visited: Set[int] = set()
        stack = sorted(seed_ids, reverse=True)
        while stack:
            node = stack.pop()
            if node in visited:
                continue
            visited.add(node)
            for neighbor in sorted(adjacency[node], reverse=True):
                if neighbor not in visited:
                    stack.append(neighbor)
        return visited

    def _compute_fragments(self) -> Tuple[Dict[int, List[int]], Dict[int, int]]:
        adjacency = self._bond_adjacency()
        fragment_of_atom: Dict[int, int] = {}
        fragments: Dict[int, List[int]] = {}
        current_frag = 0

        for aid in sorted(self.atoms):
            if aid in fragment_of_atom:
                continue
            current_frag += 1
            stack = [aid]
            component: Set[int] = set()
            while stack:
                node = stack.pop()
                if node in fragment_of_atom:
                    continue
                fragment_of_atom[node] = current_frag
                component.add(node)
                for neighbor in sorted(adjacency[node], reverse=True):
                    if neighbor not in fragment_of_atom:
                        stack.append(neighbor)
            fragments[current_frag] = sorted(component)

        return fragments, fragment_of_atom

    def visible_atom_ids(self) -> List[int]:
        self._sync_hidden_atom_ids_from_uids()
        return [aid for aid in sorted(self.atoms) if aid not in self.hidden_atoms]

    def visible_bonds(self) -> List[BondRecord]:
        self._sync_hidden_atom_ids_from_uids()
        hidden = self.hidden_atoms
        return [bond for bond in self.bonds if bond.atom1 not in hidden and bond.atom2 not in hidden]

    def atom_types(self) -> Dict[int, int]:
        return {aid: atom.atom_type for aid, atom in self.atoms.items()}

    def topology_records(self, kind: str) -> List[object]:
        if kind == "bond":
            return list(self.bonds)
        if kind == "angle":
            return list(self.angles)
        if kind == "dihedral":
            return list(self.dihedrals)
        if kind == "improper":
            return list(self.impropers)
        raise ValueError(f"unknown topology kind: {kind}")

    def topology_id_set(self, kind: str) -> Set[int]:
        if kind == "bond":
            return self.selected_bonds
        if kind == "angle":
            return self.selected_angles
        if kind == "dihedral":
            return self.selected_dihedrals
        if kind == "improper":
            return self.selected_impropers
        raise ValueError(f"unknown topology kind: {kind}")

    def topology_exists(self, kind: str) -> bool:
        return bool(self.topology_records(kind))

    def has_selection(self) -> bool:
        return bool(self.current_selection_ids())

    def current_selection_ids(self) -> Set[int]:
        if self.selection_kind == "atom":
            self._sync_selected_atom_ids_from_uids()
            return set(self.selected_atoms)
        if self.selection_kind in TOPOLOGY_KIND_TO_SECTION:
            return set(self.topology_id_set(self.selection_kind))
        return set()

    def current_pattern_atom_ids(self) -> Set[int]:
        if self.selection_kind == "atom":
            self._sync_selected_atom_ids_from_uids()
            return set(self.selected_atoms)
        atom_ids: Set[int] = set()
        if self.selection_kind in TOPOLOGY_KIND_TO_SECTION:
            selected_ids = self.topology_id_set(self.selection_kind)
            for record in self.topology_records(self.selection_kind):
                if record.record_id in selected_ids:
                    atom_ids.update(record.atom_ids)
        return atom_ids

    def current_selected_bonds(self) -> List[BondRecord]:
        if self.selection_kind != "bond":
            return []
        return [bond for bond in self.bonds if bond.bond_id in self.selected_bonds]

    def clear_topology_selection(self) -> None:
        self.selected_bonds.clear()
        self.selected_angles.clear()
        self.selected_dihedrals.clear()
        self.selected_impropers.clear()

    def select_one(self, atom_id: int, additive: bool = False, toggle: bool = False) -> None:
        if atom_id not in self.atoms:
            return
        if self.selection_kind != "atom":
            self.clear_selection()
            self.selection_kind = "atom"
        self._sync_selected_atom_ids_from_uids()
        if toggle:
            if atom_id in self.selected_atoms:
                self.selected_atoms.remove(atom_id)
            else:
                self.selected_atoms.add(atom_id)
            self.selected_atom_uids = self._atom_ids_to_uids(self.selected_atoms)
            self.selection_kind = "atom" if self.selected_atoms else None
            return
        if additive:
            self.selected_atoms.add(atom_id)
            self.selected_atom_uids = self._atom_ids_to_uids(self.selected_atoms)
        else:
            self._set_selected_atom_ids({atom_id})
        self.selection_kind = "atom"

    def select_atoms(self, atom_ids: Set[int]) -> None:
        filtered = {aid for aid in atom_ids if aid in self.atoms}
        self.clear_selection()
        if not filtered:
            return
        self._set_selected_atom_ids(filtered)
        self.selection_kind = "atom"

    def select_pattern_one(self, kind: str, record_id: int, additive: bool = False, toggle: bool = False) -> None:
        if kind not in TOPOLOGY_KIND_TO_SECTION:
            raise ValueError(f"unknown topology kind: {kind}")
        available_ids = {record.record_id for record in self.topology_records(kind)}
        if record_id not in available_ids:
            return
        if self.selection_kind != kind:
            self.clear_selection()
            self.selection_kind = kind
        selection = self.topology_id_set(kind)
        if toggle:
            if record_id in selection:
                selection.remove(record_id)
            else:
                selection.add(record_id)
            self.selection_kind = kind if selection else None
            return
        if additive:
            selection.add(record_id)
        else:
            selection.clear()
            selection.add(record_id)
        self.selection_kind = kind

    def select_same_type(self, atom_id: int) -> None:
        if atom_id not in self.atoms:
            return
        atom_type = self.atoms[atom_id].atom_type
        self.select_atoms({aid for aid, atom in self.atoms.items() if atom.atom_type == atom_type})

    def select_visible_same_element(self, atom_id: int) -> Tuple[Optional[str], int]:
        if atom_id not in self.atoms:
            return None, 0
        target_atom = self.atoms[atom_id]
        target_element = infer_element_from_mass(self.masses.get(target_atom.atom_type))
        visible_ids = set(self.visible_atom_ids())

        if target_element is None:
            selected_ids = {
                aid
                for aid in visible_ids
                if self.atoms[aid].atom_type == target_atom.atom_type
            }
            self.select_atoms(selected_ids)
            return None, len(selected_ids)

        target_symbol = target_element[0]
        selected_ids = {
            aid
            for aid in visible_ids
            if (infer_element_from_mass(self.masses.get(self.atoms[aid].atom_type)) or (None,))[0] == target_symbol
        }
        self.select_atoms(selected_ids)
        return target_symbol, len(selected_ids)

    def select_same_pattern_type(self, kind: str, record_id: int) -> None:
        if kind not in TOPOLOGY_KIND_TO_SECTION:
            raise ValueError(f"unknown topology kind: {kind}")
        target_record = next((record for record in self.topology_records(kind) if record.record_id == record_id), None)
        if target_record is None:
            return
        self.select_patterns(
            kind,
            {record.record_id for record in self.topology_records(kind) if record.record_type == target_record.record_type},
        )

    def select_fragments_of_selection(self) -> Set[int]:
        if self.selection_kind != "atom" or not self.selected_atoms:
            return set()
        new_selection = self.bonded_component_atoms(self.selected_atoms)
        self.select_atoms(new_selection)
        return new_selection

    def select_patterns(self, kind: str, record_ids: Set[int]) -> None:
        if kind not in TOPOLOGY_KIND_TO_SECTION:
            raise ValueError(f"unknown topology kind: {kind}")
        available = {record.record_id for record in self.topology_records(kind)}
        filtered = available & set(record_ids)
        self.clear_selection()
        if not filtered:
            return
        self.topology_id_set(kind).update(filtered)
        self.selection_kind = kind

    def select_current_pattern_atoms(self) -> Set[int]:
        atom_ids = self.current_pattern_atom_ids()
        self.select_atoms(atom_ids)
        return atom_ids

    def reveal_atoms(self, atom_ids: Set[int]) -> None:
        self._sync_hidden_atom_ids_from_uids()
        reveal_ids = {atom_id for atom_id in atom_ids if atom_id in self.atoms}
        self.hidden_atoms.difference_update(reveal_ids)
        self.hidden_atom_uids = self._atom_ids_to_uids(self.hidden_atoms)

    def _handle_atom_type_table_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_type_tables or item.column() not in (3, 4) or self.data_model is None:
            return

        type_item = self.atom_type_table.item(item.row(), 0)
        if type_item is None:
            self.update_summary_panel()
            return

        try:
            atom_type = int(type_item.text())
        except ValueError:
            self.update_summary_panel()
            return

        new_label = item.text().strip()
        if new_label == "未定义":
            new_label = ""
        is_name_column = item.column() == 3
        current_label = (self.data_model.mass_names if is_name_column else self.data_model.mass_labels).get(atom_type, "").strip()
        if new_label == current_label:
            return

        self._clear_left_drag_state()
        reply = QtWidgets.QMessageBox.question(
            self,
            "修改 Name" if is_name_column else "修改 ff-type",
            f"修改 {'Name' if is_name_column else 'ff-type'} 将同时更新全部 type={atom_type} 的原子，是否继续？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            self.update_summary_panel()
            return

        if is_name_column:
            self.data_model.set_name_for_type(atom_type, new_label)
            synced = self._sync_atom_type_property(atom_type, "name", new_label)
        else:
            self.data_model.set_ff_type_for_type(atom_type, new_label)
            synced = self._sync_atom_type_property(atom_type, "ff", new_label)
        self.update_summary_panel()
        self.update_properties_panel()
        self.statusBar().showMessage(f"已更新 type={atom_type} 的 {'Name' if is_name_column else 'ff-type'}，同步 {synced} 个模型")

    def _handle_atom_type_table_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_type_tables or item.column() not in (3, 4) or self.data_model is None:
            return

        type_item = self.atom_type_table.item(item.row(), 0)
        if type_item is None:
            self.update_summary_panel()
            return

        try:
            atom_type = int(type_item.text())
        except ValueError:
            self.update_summary_panel()
            return

        new_label = item.text().strip()
        if new_label == "未定义":
            new_label = ""
        is_name_column = item.column() == 3
        current_label = (self.data_model.mass_names if is_name_column else self.data_model.mass_labels).get(atom_type, "").strip()
        if new_label == current_label:
            return

        self._clear_left_drag_state()
        reply = QtWidgets.QMessageBox.question(
            self,
            "修改 Name" if is_name_column else "修改 ff-type",
            f"修改 {'Name' if is_name_column else 'ff-type'} 将同时更新全部 type={atom_type} 的原子，是否继续？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            self.update_summary_panel()
            return

        if is_name_column:
            self.data_model.set_name_for_type(atom_type, new_label)
        else:
            self.data_model.set_ff_type_for_type(atom_type, new_label)
        self.update_summary_panel()
        self.update_properties_panel()
        self.statusBar().showMessage(f"已更新 type={atom_type} 的 {'Name' if is_name_column else 'ff-type'}")

    def toggle_box_visibility(self) -> None:
        self._show_box = bool(self.toggle_box_action.isChecked())
        self.refresh_scene()

    def toggle_cross_boundary_bonds(self) -> None:
        self._hide_cross_boundary_bonds = bool(self.toggle_cross_boundary_bonds_action.isChecked())
        self.refresh_scene()

    def hide_selected(self) -> None:
        atom_ids = self.current_pattern_atom_ids()
        if not atom_ids:
            return
        self._sync_hidden_atom_ids_from_uids()
        self._set_hidden_atom_ids(self.hidden_atoms | atom_ids)
        self.clear_selection()

    def isolate_selected(self) -> None:
        atom_ids = self.current_pattern_atom_ids()
        if not atom_ids:
            return
        self._set_hidden_atom_ids({aid for aid in self.atoms if aid not in atom_ids})

    def show_all(self) -> None:
        self.hidden_atoms.clear()
        self.hidden_atom_uids.clear()

    def clear_selection(self) -> None:
        self.selected_atoms.clear()
        self.selected_atom_uids.clear()
        self.clear_topology_selection()
        self.selection_kind = None

    def find_atoms(self, field_name: str, value: float) -> Set[int]:
        if field_name == "id":
            atom_id = int(value)
            return {atom_id} if atom_id in self.atoms else set()
        if field_name == "type":
            atom_types = value if isinstance(value, set) else {int(value)}
            return {aid for aid, atom in self.atoms.items() if atom.atom_type in atom_types}
        if field_name == "mass":
            return {
                aid
                for aid, atom in self.atoms.items()
                if math.isclose(self.masses.get(atom.atom_type, float("nan")), float(value), rel_tol=1e-6, abs_tol=1e-6)
            }
        raise ValueError(f"unsupported atom find field: {field_name}")

    def find_topology(self, kind: str, field_name: str, value) -> Set[int]:
        records = self.topology_records(kind)
        if field_name == "id":
            record_id = int(value)
            return {record_id} if any(record.record_id == record_id for record in records) else set()
        if field_name == "type":
            record_types = value if isinstance(value, set) else {int(value)}
            return {record.record_id for record in records if record.record_type in record_types}
        if kind == "bond" and field_name == "atom_types":
            type1, type2 = value
            target = tuple(sorted((int(type1), int(type2))))
            matches: Set[int] = set()
            for bond in self.bonds:
                atom1 = self.atoms.get(bond.atom1)
                atom2 = self.atoms.get(bond.atom2)
                if atom1 is None or atom2 is None:
                    continue
                bond_types = tuple(sorted((atom1.atom_type, atom2.atom_type)))
                if bond_types == target:
                    matches.add(bond.bond_id)
            return matches
        raise ValueError(f"unsupported topology find field: {field_name}")

    def change_selected_type(self, new_type: int, insert_existing_type: bool = False) -> None:
        if self.selection_kind == "atom":
            selected_ids = {aid for aid in self.selected_atoms if aid in self.atoms}
            if not selected_ids:
                return
            if insert_existing_type:
                self._insert_atom_type_for_selection(int(new_type), selected_ids)
                self.normalize_type_ids(["atom"])
                self.dirty = True
                return
            for aid in self.selected_atoms:
                atom = self.atoms[aid]
                self._set_atom_type(atom, int(new_type))
            self.normalize_type_ids(["atom"])
            self.dirty = True
            return

        if self.selection_kind in TOPOLOGY_KIND_TO_SECTION:
            kind = self.selection_kind
            selected_ids = set(self.topology_id_set(kind))
            if not selected_ids:
                return
            if insert_existing_type:
                self._insert_topology_type_for_selection(kind, int(new_type), selected_ids)
                self.normalize_type_ids([kind])
                self.dirty = True
                return
            for record in self.topology_records(kind):
                if record.record_id in selected_ids:
                    self._set_topology_record_type(kind, record, int(new_type))
            self.normalize_type_ids([kind])
            self.dirty = True

    def _insert_atom_type_for_selection(self, target_type: int, selected_ids: Set[int]) -> None:
        target_type = int(target_type)
        selected_ids = {aid for aid in selected_ids if aid in self.atoms}
        if not selected_ids:
            return

        source_type = self.atoms[min(selected_ids)].atom_type
        source_mass = self.masses.get(source_type)
        source_name = self.mass_names.get(source_type, "").strip()
        source_label = self.mass_labels.get(source_type, "").strip()
        source_color = self.type_colors.get(source_type)
        all_types = sorted(set(self.masses) | set(self.mass_labels) | set(self.type_colors) | {atom.atom_type for atom in self.atoms.values()} | {target_type})
        shift_mapping = {atom_type: atom_type + 1 if atom_type >= target_type else atom_type for atom_type in all_types}

        for aid, atom in self.atoms.items():
            if aid in selected_ids:
                self._set_atom_type(atom, target_type)
            elif atom.atom_type >= target_type:
                self._set_atom_type(atom, atom.atom_type + 1)

        new_masses: Dict[int, float] = {}
        for atom_type, mass in sorted(self.masses.items()):
            new_masses[shift_mapping.get(atom_type, atom_type)] = mass
        if source_mass is not None:
            new_masses[target_type] = source_mass
        self.masses = new_masses

        new_names: Dict[int, str] = {}
        for atom_type, name in sorted(self.mass_names.items()):
            new_names[shift_mapping.get(atom_type, atom_type)] = name
        if source_name:
            new_names[target_type] = source_name
        self.mass_names = new_names

        new_labels: Dict[int, str] = {}
        for atom_type, label in sorted(self.mass_labels.items()):
            new_labels[shift_mapping.get(atom_type, atom_type)] = label
        if source_label:
            new_labels[target_type] = source_label
        self.mass_labels = new_labels

        old_coeff_values = self.coeff_values.get("atom", {})
        self.coeff_values["atom"] = {
            shift_mapping.get(atom_type, atom_type): list(tokens)
            for atom_type, tokens in sorted(old_coeff_values.items())
        }

        new_colors: Dict[int, Tuple[int, int, int]] = {}
        for atom_type, color in sorted(self.type_colors.items()):
            new_colors[shift_mapping.get(atom_type, atom_type)] = color
        if source_color is not None:
            new_colors[target_type] = source_color
        self.type_colors = new_colors
        self._compose_type_id_map("atom", shift_mapping)

    def _set_topology_record_type(self, kind: str, record: object, new_type: int) -> None:
        attr = TOPOLOGY_TYPE_ATTR.get(kind)
        if attr is None:
            raise ValueError(f"unknown topology kind: {kind}")
        setattr(record, attr, int(new_type))

    def _insert_topology_type_for_selection(self, kind: str, target_type: int, selected_ids: Set[int]) -> None:
        if kind not in TOPOLOGY_TYPE_ATTR:
            raise ValueError(f"unknown topology kind: {kind}")
        target_type = int(target_type)
        records = self.topology_records(kind)
        available_ids = {record.record_id for record in records}
        selected_ids = set(selected_ids) & available_ids
        if not selected_ids:
            return

        source_record = min((record for record in records if record.record_id in selected_ids), key=lambda item: item.record_id)
        source_type = source_record.record_type
        source_name = self.coeff_names.get(kind, {}).get(source_type, "").strip()
        source_label = self.coeff_labels.get(kind, {}).get(source_type, "").strip()
        all_types = sorted(self._current_type_ids(kind) | set(self.coeff_labels.get(kind, {})) | {target_type})
        shift_mapping = {record_type: record_type + 1 if record_type >= target_type else record_type for record_type in all_types}

        for record in records:
            if record.record_id in selected_ids:
                self._set_topology_record_type(kind, record, target_type)
            elif record.record_type >= target_type:
                self._set_topology_record_type(kind, record, record.record_type + 1)

        old_names = self.coeff_names.get(kind, {})
        new_names: Dict[int, str] = {}
        for record_type, name in sorted(old_names.items()):
            new_names[shift_mapping.get(record_type, record_type)] = name
        if source_name:
            new_names[target_type] = source_name
        self.coeff_names[kind] = new_names

        old_labels = self.coeff_labels.get(kind, {})
        new_labels: Dict[int, str] = {}
        for record_type, label in sorted(old_labels.items()):
            new_labels[shift_mapping.get(record_type, record_type)] = label
        if source_label:
            new_labels[target_type] = source_label
        self.coeff_labels[kind] = new_labels
        old_coeff_values = self.coeff_values.get(kind, {})
        self.coeff_values[kind] = {
            shift_mapping.get(record_type, record_type): list(tokens)
            for record_type, tokens in sorted(old_coeff_values.items())
        }
        self._compose_type_id_map(kind, shift_mapping)

    def total_charge(self) -> float:
        total = 0.0
        for atom in self.atoms.values():
            if atom.charge is not None:
                total += atom.charge
        return total

    def set_mass_for_type(self, atom_type: int, new_mass: float) -> None:
        self.masses[int(atom_type)] = float(new_mass)
        self.dirty = True

    def set_ff_type_for_type(self, atom_type: int, label: str) -> None:
        self.mass_labels[int(atom_type)] = label.strip()
        self.dirty = True

    def set_name_for_type(self, atom_type: int, name: str) -> None:
        self.mass_names[int(atom_type)] = name.strip()
        self.dirty = True

    def set_coeff_label_for_type(self, kind: str, type_id: int, label: str) -> None:
        if kind not in ("bond", "angle", "dihedral", "improper"):
            raise ValueError(f"unknown topology kind: {kind}")
        self.coeff_labels.setdefault(kind, {})[int(type_id)] = label.strip()
        self.dirty = True

    def set_coeff_name_for_type(self, kind: str, type_id: int, name: str) -> None:
        if kind not in ("bond", "angle", "dihedral", "improper"):
            raise ValueError(f"unknown topology kind: {kind}")
        self.coeff_names.setdefault(kind, {})[int(type_id)] = name.strip()
        self.dirty = True

    def set_coeff_values_for_type(self, kind: str, type_id: int, tokens: Sequence[object]) -> None:
        if kind not in ("atom", "bond", "angle", "dihedral", "improper"):
            raise ValueError(f"unknown coeff kind: {kind}")
        self.coeff_values.setdefault(kind, {})[int(type_id)] = [_format_forcefield_value(token) for token in tokens]
        self.dirty = True

    def apply_forcefield(self, definition: ForceFieldDefinition) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "atom": 0,
            "bond": 0,
            "angle": 0,
            "dihedral": 0,
            "improper": 0,
            "missing": [],
            "unsupported": [],
        }
        token_index = self._charge_token_index()
        atom_types = sorted(set(self.masses) | {atom.atom_type for atom in self.atoms.values()})
        for atom_type in atom_types:
            label = self.mass_labels.get(atom_type, "").strip()
            if not label:
                continue
            entry = definition.atom_types.get(label)
            if entry is None:
                summary["missing"].append(f"atom type {atom_type}: {label}")
                continue
            charge = entry.get("charge")
            if charge is not None:
                try:
                    charge_value = float(charge)
                except (TypeError, ValueError):
                    charge_value = None
                if charge_value is not None:
                    for atom in self.atoms.values():
                        if atom.atom_type != atom_type:
                            continue
                        atom.charge = charge_value
                        if atom.raw_tokens and token_index is not None and len(atom.raw_tokens) > token_index:
                            atom.raw_tokens[token_index] = f"{charge_value:g}"
            mass = entry.get("mass")
            if mass is not None:
                try:
                    self.masses[atom_type] = float(mass)
                except (TypeError, ValueError):
                    pass
            coeff_tokens = _atom_pair_coeff_tokens(entry)
            if coeff_tokens is not None:
                self.set_coeff_values_for_type("atom", atom_type, coeff_tokens)
            else:
                summary["unsupported"].append(f"atom type {atom_type}: {label} pair coeff")
            summary["atom"] += 1

        topology_maps = {
            "bond": definition.bond_types,
            "angle": definition.angle_types,
            "dihedral": definition.dihedral_types,
            "improper": definition.improper_types,
        }
        for kind, entries in topology_maps.items():
            if not entries:
                continue
            used_types = sorted(self._current_type_ids(kind) | set(self.coeff_labels.get(kind, {})))
            for type_id in used_types:
                label = self.coeff_labels.get(kind, {}).get(type_id, "").strip()
                if not label:
                    continue
                entry = entries.get(label)
                if entry is None:
                    summary["missing"].append(f"{kind} type {type_id}: {label}")
                    continue
                coeff_tokens = _topology_coeff_tokens(entry, kind)
                if coeff_tokens is None:
                    summary["unsupported"].append(f"{kind} type {type_id}: {label} coeff")
                    continue
                self.set_coeff_values_for_type(kind, type_id, coeff_tokens)
                summary[kind] += 1
        if any(summary[kind] for kind in ("atom", "bond", "angle", "dihedral", "improper")):
            self.dirty = True
        return summary

    def set_color_for_type(self, atom_type: int, color: Tuple[int, int, int]) -> None:
        r, g, b = color
        self.type_colors[int(atom_type)] = (
            max(0, min(255, int(r))),
            max(0, min(255, int(g))),
            max(0, min(255, int(b))),
        )

    def set_charge_for_atoms(self, atom_ids: Set[int], new_charge: float) -> int:
        token_index = self._charge_token_index()
        changed = 0
        for aid in sorted(atom_ids):
            atom = self.atoms.get(aid)
            if atom is None:
                continue
            atom.charge = float(new_charge)
            if atom.raw_tokens and token_index is not None and len(atom.raw_tokens) > token_index:
                atom.raw_tokens[token_index] = f"{float(new_charge):g}"
            changed += 1
        if changed:
            self.dirty = True
        return changed

    def atom_ids_of_type(self, atom_type: int) -> Set[int]:
        return {aid for aid, atom in self.atoms.items() if atom.atom_type == int(atom_type)}

    def atom_type_count(self) -> int:
        return len({atom.atom_type for atom in self.atoms.values()})

    def topology_type_count(self, kind: str) -> int:
        return len(self._current_type_ids(kind))

    def _current_type_ids(self, kind: str) -> Set[int]:
        if kind == "atom":
            return {atom.atom_type for atom in self.atoms.values()}
        if kind == "bond":
            return {bond.bond_type for bond in self.bonds}
        if kind == "angle":
            return {angle.angle_type for angle in self.angles}
        if kind == "dihedral":
            return {record.dihedral_type for record in self.dihedrals}
        if kind == "improper":
            return {record.improper_type for record in self.impropers}
        raise ValueError(f"unknown type kind: {kind}")

    def _compose_type_id_map(self, kind: str, mapping: Dict[int, int]) -> None:
        old_aliases = self.type_id_maps.get(kind, {})
        self.type_id_maps[kind] = {
            original_type: mapping[current_type]
            for original_type, current_type in old_aliases.items()
            if current_type in mapping
        }

    def normalize_type_ids(self, kinds: Optional[Sequence[str]] = None) -> Dict[str, Dict[int, int]]:
        target_kinds = list(kinds) if kinds is not None else ["atom", "bond", "angle", "dihedral", "improper"]
        mappings: Dict[str, Dict[int, int]] = {}
        changed = False

        for kind in target_kinds:
            used_types = sorted(self._current_type_ids(kind))
            mapping = {old_type: new_type for new_type, old_type in enumerate(used_types, start=1)}
            mappings[kind] = mapping

            if kind == "atom":
                metadata_changed = bool((set(self.masses) | set(self.mass_names) | set(self.mass_labels) | set(self.type_colors) | set(self.coeff_values.get("atom", {}))) - set(used_types))
                for atom in self.atoms.values():
                    new_type = mapping.get(atom.atom_type)
                    if new_type is not None and new_type != atom.atom_type:
                        self._set_atom_type(atom, new_type)
                        changed = True

                new_masses: Dict[int, float] = {}
                new_names: Dict[int, str] = {}
                new_labels: Dict[int, str] = {}
                new_colors: Dict[int, Tuple[int, int, int]] = {}
                new_coeff_values: Dict[int, List[str]] = {}
                for old_type, new_type in mapping.items():
                    if old_type in self.masses:
                        new_masses[new_type] = self.masses[old_type]
                    if old_type in self.mass_names:
                        new_names[new_type] = self.mass_names[old_type]
                    if old_type in self.mass_labels:
                        new_labels[new_type] = self.mass_labels[old_type]
                    if old_type in self.type_colors:
                        new_colors[new_type] = self.type_colors[old_type]
                    if old_type in self.coeff_values.get("atom", {}):
                        new_coeff_values[new_type] = list(self.coeff_values["atom"][old_type])
                if new_masses != self.masses or new_names != self.mass_names or new_labels != self.mass_labels or new_colors != self.type_colors or new_coeff_values != self.coeff_values.get("atom", {}):
                    changed = True
                if metadata_changed:
                    changed = True
                self.masses = new_masses
                self.mass_names = new_names
                self.mass_labels = new_labels
                self.type_colors = new_colors
                self.coeff_values["atom"] = new_coeff_values
            elif kind == "bond":
                for bond in self.bonds:
                    new_type = mapping.get(bond.bond_type)
                    if new_type is not None and new_type != bond.bond_type:
                        bond.bond_type = new_type
                        changed = True
                old_labels = self.coeff_labels.get(kind, {})
                old_names = self.coeff_names.get(kind, {})
                new_names = {new_type: old_names[old_type] for old_type, new_type in mapping.items() if old_type in old_names}
                new_labels = {new_type: old_labels[old_type] for old_type, new_type in mapping.items() if old_type in old_labels}
                old_coeff_values = self.coeff_values.get(kind, {})
                new_coeff_values = {new_type: list(old_coeff_values[old_type]) for old_type, new_type in mapping.items() if old_type in old_coeff_values}
                if new_names != old_names or new_labels != old_labels or new_coeff_values != old_coeff_values:
                    changed = True
                self.coeff_names[kind] = new_names
                self.coeff_labels[kind] = new_labels
                self.coeff_values[kind] = new_coeff_values
            elif kind == "angle":
                for angle in self.angles:
                    new_type = mapping.get(angle.angle_type)
                    if new_type is not None and new_type != angle.angle_type:
                        angle.angle_type = new_type
                        changed = True
                old_labels = self.coeff_labels.get(kind, {})
                old_names = self.coeff_names.get(kind, {})
                new_names = {new_type: old_names[old_type] for old_type, new_type in mapping.items() if old_type in old_names}
                new_labels = {new_type: old_labels[old_type] for old_type, new_type in mapping.items() if old_type in old_labels}
                old_coeff_values = self.coeff_values.get(kind, {})
                new_coeff_values = {new_type: list(old_coeff_values[old_type]) for old_type, new_type in mapping.items() if old_type in old_coeff_values}
                if new_names != old_names or new_labels != old_labels or new_coeff_values != old_coeff_values:
                    changed = True
                self.coeff_names[kind] = new_names
                self.coeff_labels[kind] = new_labels
                self.coeff_values[kind] = new_coeff_values
            elif kind == "dihedral":
                for record in self.dihedrals:
                    new_type = mapping.get(record.dihedral_type)
                    if new_type is not None and new_type != record.dihedral_type:
                        record.dihedral_type = new_type
                        changed = True
                old_labels = self.coeff_labels.get(kind, {})
                old_names = self.coeff_names.get(kind, {})
                new_names = {new_type: old_names[old_type] for old_type, new_type in mapping.items() if old_type in old_names}
                new_labels = {new_type: old_labels[old_type] for old_type, new_type in mapping.items() if old_type in old_labels}
                old_coeff_values = self.coeff_values.get(kind, {})
                new_coeff_values = {new_type: list(old_coeff_values[old_type]) for old_type, new_type in mapping.items() if old_type in old_coeff_values}
                if new_names != old_names or new_labels != old_labels or new_coeff_values != old_coeff_values:
                    changed = True
                self.coeff_names[kind] = new_names
                self.coeff_labels[kind] = new_labels
                self.coeff_values[kind] = new_coeff_values
            elif kind == "improper":
                for record in self.impropers:
                    new_type = mapping.get(record.improper_type)
                    if new_type is not None and new_type != record.improper_type:
                        record.improper_type = new_type
                        changed = True
                old_labels = self.coeff_labels.get(kind, {})
                old_names = self.coeff_names.get(kind, {})
                new_names = {new_type: old_names[old_type] for old_type, new_type in mapping.items() if old_type in old_names}
                new_labels = {new_type: old_labels[old_type] for old_type, new_type in mapping.items() if old_type in old_labels}
                old_coeff_values = self.coeff_values.get(kind, {})
                new_coeff_values = {new_type: list(old_coeff_values[old_type]) for old_type, new_type in mapping.items() if old_type in old_coeff_values}
                if new_names != old_names or new_labels != old_labels or new_coeff_values != old_coeff_values:
                    changed = True
                self.coeff_names[kind] = new_names
                self.coeff_labels[kind] = new_labels
                self.coeff_values[kind] = new_coeff_values

            self._compose_type_id_map(kind, mapping)

        if changed:
            self.dirty = True
        return mappings

    def _set_atom_type(self, atom: AtomRecord, new_type: int) -> None:
        atom.atom_type = int(new_type)
        if atom.raw_tokens:
            atom.raw_tokens[self._type_token_index()] = str(int(new_type))

    def _shift_atom_types_after(self, after_type: int, delta: int) -> None:
        if delta <= 0:
            return
        shift_mapping = {
            atom_type: atom_type + delta if atom_type > after_type else atom_type
            for atom_type in sorted(set(self.masses) | set(self.mass_names) | set(self.mass_labels) | set(self.type_colors) | {atom.atom_type for atom in self.atoms.values()})
        }
        for atom in self.atoms.values():
            if atom.atom_type > after_type:
                self._set_atom_type(atom, atom.atom_type + delta)

        new_masses: Dict[int, float] = {}
        for atom_type, mass in sorted(self.masses.items()):
            new_type = atom_type + delta if atom_type > after_type else atom_type
            new_masses[new_type] = mass
        self.masses = new_masses

        new_names: Dict[int, str] = {}
        for atom_type, name in sorted(self.mass_names.items()):
            new_type = atom_type + delta if atom_type > after_type else atom_type
            new_names[new_type] = name
        self.mass_names = new_names

        new_labels: Dict[int, str] = {}
        for atom_type, label in sorted(self.mass_labels.items()):
            new_type = atom_type + delta if atom_type > after_type else atom_type
            new_labels[new_type] = label
        self.mass_labels = new_labels

        old_coeff_values = self.coeff_values.get("atom", {})
        self.coeff_values["atom"] = {
            atom_type + delta if atom_type > after_type else atom_type: list(tokens)
            for atom_type, tokens in sorted(old_coeff_values.items())
        }

        new_colors: Dict[int, Tuple[int, int, int]] = {}
        for atom_type, color in sorted(self.type_colors.items()):
            new_type = atom_type + delta if atom_type > after_type else atom_type
            new_colors[new_type] = color
        self.type_colors = new_colors
        self._compose_type_id_map("atom", shift_mapping)

    def _suggest_shell_label(self, base_label: str, base_type: int) -> str:
        cleaned = base_label.strip()
        if cleaned:
            return f"{cleaned}_shell"
        return f"type_{base_type}_shell"

    def add_shell_atoms(
        self,
        original_type: int,
        selected_atom_ids: Set[int],
        split_selected_atoms: bool = False,
        z_offset: float = 0.05,
    ) -> Dict[str, object]:
        original_type = int(original_type)
        selected_ids = {aid for aid in selected_atom_ids if aid in self.atoms and self.atoms[aid].atom_type == original_type}
        all_type_ids = self.atom_ids_of_type(original_type)
        if not selected_ids:
            raise ValueError("未选中可用于添加 shell 的原子。")
        if not all_type_ids:
            raise ValueError("目标 type 不存在。")

        original_mass = self.masses.get(original_type)
        original_name = self.mass_names.get(original_type, "").strip()
        original_label = self.mass_labels.get(original_type, "").strip()
        original_color = self.type_colors.get(original_type)
        next_atom_id = max(self.atoms) if self.atoms else 0

        if split_selected_atoms:
            self._shift_atom_types_after(original_type, 2)
            core_type = original_type + 1
            shell_type = original_type + 2
            if original_mass is not None:
                self.masses[core_type] = original_mass
            self.mass_names[core_type] = original_name
            self.mass_labels[core_type] = original_label
            if original_color is not None:
                self.type_colors[core_type] = original_color
            for atom_id in sorted(selected_ids):
                self._set_atom_type(self.atoms[atom_id], core_type)
            source_atom_ids = sorted(selected_ids)
        else:
            self._shift_atom_types_after(original_type, 1)
            core_type = original_type
            shell_type = original_type + 1
            source_atom_ids = sorted(all_type_ids)

        if original_mass is not None:
            self.masses[shell_type] = original_mass
        self.mass_names[shell_type] = self._suggest_shell_label(original_name, core_type)
        self.mass_labels[shell_type] = self._suggest_shell_label(original_label, core_type)
        if original_color is not None:
            self.type_colors[shell_type] = original_color

        shell_bond_type = max(self._current_type_ids("bond") | set(self.coeff_labels.get("bond", {})), default=0) + 1
        self.coeff_names.setdefault("bond", {})[shell_bond_type] = "core-shell"
        self.coeff_labels.setdefault("bond", {})[shell_bond_type] = "core-shell"
        next_bond_id = max((bond.bond_id for bond in self.bonds), default=0)

        new_shell_ids: List[int] = []
        new_shell_bond_ids: List[int] = []
        new_hidden_ids: Set[int] = set()
        next_model_id = max((model_id for _, model_id in self.cs_info), default=0)
        next_atom_uid = self._next_atom_uid()
        for atom_id in source_atom_ids:
            source_atom = self.atoms[atom_id]
            next_atom_id += 1
            next_bond_id += 1
            next_model_id += 1
            shell_atom = AtomRecord(
                uid=next_atom_uid,
                atom_id=next_atom_id,
                atom_type=shell_type,
                x=source_atom.x,
                y=source_atom.y,
                z=source_atom.z - float(z_offset),
                mol=source_atom.mol,
                charge=source_atom.charge,
                image=source_atom.image,
            )
            next_atom_uid += 1
            self.atoms[next_atom_id] = shell_atom
            self.bonds.append(
                BondRecord(
                    bond_id=next_bond_id,
                    bond_type=shell_bond_type,
                    atom1=atom_id,
                    atom2=next_atom_id,
                    raw_comment="# core-shell",
                )
            )
            new_shell_ids.append(next_atom_id)
            new_shell_bond_ids.append(next_bond_id)
            self.cs_info.append((atom_id, next_model_id))
            self.cs_info.append((next_atom_id, next_model_id))
            self.cs_info_uids.append((source_atom.uid, next_model_id))
            self.cs_info_uids.append((shell_atom.uid, next_model_id))
            if atom_id in self.hidden_atoms:
                new_hidden_ids.add(next_atom_id)

        self.hidden_atoms.update(new_hidden_ids)
        self.fragments, self.fragment_of_atom = self._compute_fragments()
        self.dirty = True
        return {
            "original_type": original_type,
            "core_type": core_type,
            "shell_type": shell_type,
            "source_atom_ids": source_atom_ids,
            "shell_atom_ids": new_shell_ids,
            "shell_bond_type": shell_bond_type,
            "shell_bond_ids": new_shell_bond_ids,
            "split_selected_atoms": split_selected_atoms,
        }

    def _next_mol_id(self) -> Optional[int]:
        if self.atom_style not in {"full", "molecular"}:
            return None
        return max((atom.mol or 0 for atom in self.atoms.values()), default=0) + 1

    def _representative_charge_for_type(self, atom_type: int) -> Optional[float]:
        charges = [atom.charge for atom in self.atoms.values() if atom.atom_type == int(atom_type) and atom.charge is not None]
        return charges[0] if charges else None

    def insert_atom(self, atom_type: int, position: Sequence[float]) -> int:
        atom_type = int(atom_type)
        next_atom_id = max(self.atoms, default=0) + 1
        charge = self._representative_charge_for_type(atom_type)
        x, y, z = (float(position[0]), float(position[1]), float(position[2]))
        atom = AtomRecord(
            uid=self._next_atom_uid(),
            atom_id=next_atom_id,
            atom_type=atom_type,
            x=x,
            y=y,
            z=z,
            mol=self._next_mol_id(),
            charge=charge,
        )
        self.atoms[next_atom_id] = atom
        self._after_structure_change()
        self.dirty = True
        return next_atom_id

    def insert_fragment_copy(self, fragment_id: int, target_position: Sequence[float]) -> Dict[str, object]:
        fragment_atoms = sorted(self.fragments.get(int(fragment_id), []))
        if not fragment_atoms:
            raise ValueError(f"fragment {fragment_id} 不存在。")

        coords = np.array([[self.atoms[aid].x, self.atoms[aid].y, self.atoms[aid].z] for aid in fragment_atoms], dtype=float)
        centroid = coords.mean(axis=0)
        target = np.array([float(target_position[0]), float(target_position[1]), float(target_position[2])], dtype=float)
        translation = target - centroid

        atom_id_map: Dict[int, int] = {}
        next_atom_id = max(self.atoms, default=0)
        next_uid = self._next_atom_uid()
        new_mol_id = self._next_mol_id()
        for atom_id in fragment_atoms:
            source = self.atoms[atom_id]
            next_atom_id += 1
            new_pos = np.array([source.x, source.y, source.z], dtype=float) + translation
            atom_id_map[atom_id] = next_atom_id
            self.atoms[next_atom_id] = AtomRecord(
                uid=next_uid,
                atom_id=next_atom_id,
                atom_type=source.atom_type,
                x=float(new_pos[0]),
                y=float(new_pos[1]),
                z=float(new_pos[2]),
                mol=new_mol_id if new_mol_id is not None else source.mol,
                charge=source.charge,
                image=source.image,
            )
            next_uid += 1

        atom_set = set(fragment_atoms)
        next_bond_id = max((bond.bond_id for bond in self.bonds), default=0)
        new_bond_ids: List[int] = []
        for bond in list(self.bonds):
            if set(bond.atom_ids) <= atom_set:
                next_bond_id += 1
                self.bonds.append(
                    BondRecord(
                        bond_id=next_bond_id,
                        bond_type=bond.bond_type,
                        atom1=atom_id_map[bond.atom1],
                        atom2=atom_id_map[bond.atom2],
                        raw_comment=bond.raw_comment,
                    )
                )
                new_bond_ids.append(next_bond_id)

        next_angle_id = max((angle.angle_id for angle in self.angles), default=0)
        new_angle_ids: List[int] = []
        for angle in list(self.angles):
            if set(angle.atom_ids) <= atom_set:
                next_angle_id += 1
                self.angles.append(
                    AngleRecord(
                        angle_id=next_angle_id,
                        angle_type=angle.angle_type,
                        atom1=atom_id_map[angle.atom1],
                        atom2=atom_id_map[angle.atom2],
                        atom3=atom_id_map[angle.atom3],
                        raw_comment=angle.raw_comment,
                    )
                )
                new_angle_ids.append(next_angle_id)

        next_dihedral_id = max((record.dihedral_id for record in self.dihedrals), default=0)
        new_dihedral_ids: List[int] = []
        for record in list(self.dihedrals):
            if set(record.atom_ids) <= atom_set:
                next_dihedral_id += 1
                self.dihedrals.append(
                    DihedralRecord(
                        dihedral_id=next_dihedral_id,
                        dihedral_type=record.dihedral_type,
                        atom1=atom_id_map[record.atom1],
                        atom2=atom_id_map[record.atom2],
                        atom3=atom_id_map[record.atom3],
                        atom4=atom_id_map[record.atom4],
                        raw_comment=record.raw_comment,
                    )
                )
                new_dihedral_ids.append(next_dihedral_id)

        next_improper_id = max((record.improper_id for record in self.impropers), default=0)
        new_improper_ids: List[int] = []
        for record in list(self.impropers):
            if set(record.atom_ids) <= atom_set:
                next_improper_id += 1
                self.impropers.append(
                    ImproperRecord(
                        improper_id=next_improper_id,
                        improper_type=record.improper_type,
                        atom1=atom_id_map[record.atom1],
                        atom2=atom_id_map[record.atom2],
                        atom3=atom_id_map[record.atom3],
                        atom4=atom_id_map[record.atom4],
                        raw_comment=record.raw_comment,
                    )
                )
                new_improper_ids.append(next_improper_id)

        self._after_structure_change()
        self.dirty = True
        return {
            "atom_ids": sorted(atom_id_map.values()),
            "bond_ids": new_bond_ids,
            "angle_ids": new_angle_ids,
            "dihedral_ids": new_dihedral_ids,
            "improper_ids": new_improper_ids,
        }

    def delete_current_selection(self) -> Dict[str, int]:
        if self.selection_kind == "atom":
            self._sync_selected_atom_ids_from_uids()
            return self.delete_atoms(self.selected_atoms)
        if self.selection_kind == "bond":
            return self.delete_bonds(self.selected_bonds)
        if self.selection_kind == "angle":
            return self.delete_angles(self.selected_angles)
        if self.selection_kind == "dihedral":
            return self.delete_dihedrals(self.selected_dihedrals)
        if self.selection_kind == "improper":
            return self.delete_impropers(self.selected_impropers)
        return {}

    def delete_atoms(self, atom_ids: Set[int]) -> Dict[str, int]:
        remove_atoms = {aid for aid in atom_ids if aid in self.atoms}
        if not remove_atoms:
            return {}
        remove_uids = self._atom_ids_to_uids(remove_atoms)
        remove_bonds = {bond.bond_id for bond in self.bonds if set(bond.atom_ids) & remove_atoms}
        remove_angles = {angle.angle_id for angle in self.angles if set(angle.atom_ids) & remove_atoms}
        remove_dihedrals = {record.dihedral_id for record in self.dihedrals if set(record.atom_ids) & remove_atoms}
        remove_impropers = {record.improper_id for record in self.impropers if set(record.atom_ids) & remove_atoms}

        self.atoms = {aid: atom for aid, atom in self.atoms.items() if aid not in remove_atoms}
        self.bonds = [bond for bond in self.bonds if bond.bond_id not in remove_bonds]
        self.angles = [angle for angle in self.angles if angle.angle_id not in remove_angles]
        self.dihedrals = [record for record in self.dihedrals if record.dihedral_id not in remove_dihedrals]
        self.impropers = [record for record in self.impropers if record.improper_id not in remove_impropers]
        self.cs_info = [(aid, model_id) for aid, model_id in self.cs_info if aid not in remove_atoms]
        self.cs_info_uids = [(uid, model_id) for uid, model_id in self.cs_info_uids if uid not in remove_uids]

        atom_mapping = self._renumber_atoms()
        self._remap_topology_atoms(atom_mapping)
        self._renumber_topology_ids()
        self.hidden_atom_uids.difference_update(remove_uids)
        self._sync_hidden_atom_ids_from_uids()
        self._sync_cs_info_ids_from_uids()
        self._after_structure_change()
        self.dirty = True
        return {
            "atom": len(remove_atoms),
            "bond": len(remove_bonds),
            "angle": len(remove_angles),
            "dihedral": len(remove_dihedrals),
            "improper": len(remove_impropers),
        }

    def delete_bonds(self, bond_ids: Set[int]) -> Dict[str, int]:
        selected_bonds = [bond for bond in self.bonds if bond.bond_id in bond_ids]
        if not selected_bonds:
            return {}
        remove_bonds = {bond.bond_id for bond in selected_bonds}
        self.bonds = [bond for bond in self.bonds if bond.bond_id not in remove_bonds]
        self._renumber_topology_ids()
        self._after_structure_change()
        self.dirty = True
        return {"bond": len(remove_bonds)}

    def delete_angles(self, angle_ids: Set[int]) -> Dict[str, int]:
        selected_angles = [angle for angle in self.angles if angle.angle_id in angle_ids]
        if not selected_angles:
            return {}
        remove_angles = {angle.angle_id for angle in selected_angles}
        self.angles = [angle for angle in self.angles if angle.angle_id not in remove_angles]
        self._renumber_topology_ids()
        self._after_structure_change()
        self.dirty = True
        return {"angle": len(remove_angles)}

    def delete_dihedrals(self, dihedral_ids: Set[int]) -> Dict[str, int]:
        selected_dihedrals = [record for record in self.dihedrals if record.dihedral_id in dihedral_ids]
        if not selected_dihedrals:
            return {}
        remove_dihedrals = {record.dihedral_id for record in selected_dihedrals}
        self.dihedrals = [record for record in self.dihedrals if record.dihedral_id not in remove_dihedrals]
        self._renumber_topology_ids()
        self._after_structure_change()
        self.dirty = True
        return {"dihedral": len(remove_dihedrals)}

    def delete_impropers(self, improper_ids: Set[int]) -> Dict[str, int]:
        remove_impropers = {record.improper_id for record in self.impropers if record.improper_id in improper_ids}
        if not remove_impropers:
            return {}
        self.impropers = [record for record in self.impropers if record.improper_id not in remove_impropers]
        self._renumber_topology_ids()
        self._after_structure_change()
        self.dirty = True
        return {"improper": len(remove_impropers)}

    def _after_structure_change(self) -> None:
        self.normalize_type_ids()
        self.clear_selection()
        self.fragments, self.fragment_of_atom = self._compute_fragments()
        self._sync_hidden_atom_ids_from_uids()

    def _renumber_atoms(self) -> Dict[int, int]:
        atom_mapping: Dict[int, int] = {}
        ordered_atoms = [self.atoms[aid] for aid in sorted(self.atoms)]
        new_atoms: Dict[int, AtomRecord] = {}
        for new_id, atom in enumerate(ordered_atoms, start=1):
            old_id = atom.atom_id
            atom_mapping[old_id] = new_id
            atom.atom_id = new_id
            if atom.raw_tokens:
                atom.raw_tokens[0] = str(new_id)
            new_atoms[new_id] = atom
        self.atoms = new_atoms
        return atom_mapping

    def _remap_topology_atoms(self, atom_mapping: Dict[int, int]) -> None:
        for bond in self.bonds:
            bond.atom1 = atom_mapping[bond.atom1]
            bond.atom2 = atom_mapping[bond.atom2]
        for angle in self.angles:
            angle.atom1 = atom_mapping[angle.atom1]
            angle.atom2 = atom_mapping[angle.atom2]
            angle.atom3 = atom_mapping[angle.atom3]
        for record in self.dihedrals:
            record.atom1 = atom_mapping[record.atom1]
            record.atom2 = atom_mapping[record.atom2]
            record.atom3 = atom_mapping[record.atom3]
            record.atom4 = atom_mapping[record.atom4]
        for record in self.impropers:
            record.atom1 = atom_mapping[record.atom1]
            record.atom2 = atom_mapping[record.atom2]
            record.atom3 = atom_mapping[record.atom3]
            record.atom4 = atom_mapping[record.atom4]
        self._sync_cs_info_ids_from_uids()

    def _renumber_topology_ids(self) -> None:
        for new_id, bond in enumerate(self.bonds, start=1):
            bond.bond_id = new_id
        for new_id, angle in enumerate(self.angles, start=1):
            angle.angle_id = new_id
        for new_id, record in enumerate(self.dihedrals, start=1):
            record.dihedral_id = new_id
        for new_id, record in enumerate(self.impropers, start=1):
            record.improper_id = new_id

    def _type_token_index(self) -> int:
        if self.atom_style == "full":
            return 2
        if self.atom_style == "charge":
            return 1
        if self.atom_style == "molecular":
            return 2
        if self.atom_style == "atomic":
            return 1
        raise ValueError(f"未知 atom_style: {self.atom_style}")

    def _charge_token_index(self) -> Optional[int]:
        if self.atom_style == "full":
            return 3
        if self.atom_style == "charge":
            return 2
        return None

    def _atom_tokens(self, atom: AtomRecord) -> List[str]:
        if atom.raw_tokens:
            return atom.raw_tokens[:]
        if self.atom_style == "full":
            tokens = [
                str(atom.atom_id),
                str(atom.mol if atom.mol is not None else 0),
                str(atom.atom_type),
                f"{atom.charge if atom.charge is not None else 0.0:g}",
                f"{atom.x:.10f}",
                f"{atom.y:.10f}",
                f"{atom.z:.10f}",
            ]
        elif self.atom_style == "charge":
            tokens = [
                str(atom.atom_id),
                str(atom.atom_type),
                f"{atom.charge if atom.charge is not None else 0.0:g}",
                f"{atom.x:.10f}",
                f"{atom.y:.10f}",
                f"{atom.z:.10f}",
            ]
        elif self.atom_style == "molecular":
            tokens = [
                str(atom.atom_id),
                str(atom.mol if atom.mol is not None else 0),
                str(atom.atom_type),
                f"{atom.x:.10f}",
                f"{atom.y:.10f}",
                f"{atom.z:.10f}",
            ]
        else:
            tokens = [
                str(atom.atom_id),
                str(atom.atom_type),
                f"{atom.x:.10f}",
                f"{atom.y:.10f}",
                f"{atom.z:.10f}",
            ]
        if atom.image != (0, 0, 0):
            tokens.extend(str(value) for value in atom.image)
        return tokens

    def _format_line(self, tokens: Sequence[object], raw_comment: str = "") -> str:
        main = " ".join(map(str, tokens))
        if raw_comment:
            return f"{main} {raw_comment}\n"
        return f"{main}\n"

    def _render_topology_section(self, section_name: str) -> List[str]:
        rendered = ["\n"]
        if section_name == "Masses":
            for atom_type in sorted(self.masses):
                label = self.mass_names.get(atom_type, "").strip()
                comment = f"# {label}" if label else ""
                rendered.append(self._format_line([atom_type, f"{self.masses[atom_type]:g}"], comment))
        elif section_name == "Atoms":
            for atom_id in sorted(self.atoms):
                atom = self.atoms[atom_id]
                label = self.mass_names.get(atom.atom_type, "").strip()
                comment = f"# {label}" if label else ""
                rendered.append(self._format_line(self._atom_tokens(atom), comment))
        elif section_name == "Bonds":
            for bond in self.bonds:
                comment = self._topology_record_comment("bond", bond)
                rendered.append(self._format_line([bond.bond_id, bond.bond_type, bond.atom1, bond.atom2], comment))
        elif section_name == "Angles":
            for angle in self.angles:
                comment = self._topology_record_comment("angle", angle)
                rendered.append(
                    self._format_line([angle.angle_id, angle.angle_type, angle.atom1, angle.atom2, angle.atom3], comment)
                )
        elif section_name == "Dihedrals":
            for record in self.dihedrals:
                comment = self._topology_record_comment("dihedral", record)
                rendered.append(
                    self._format_line(
                        [record.dihedral_id, record.dihedral_type, record.atom1, record.atom2, record.atom3, record.atom4],
                        comment,
                    )
                )
        elif section_name == "Impropers":
            for record in self.impropers:
                comment = self._topology_record_comment("improper", record)
                rendered.append(
                    self._format_line(
                        [record.improper_id, record.improper_type, record.atom1, record.atom2, record.atom3, record.atom4],
                        comment,
                    )
                )
        rendered.append("\n")
        return rendered

    def _topology_record_comment(self, kind: str, record: object) -> str:
        labels = self.coeff_names.get(kind, {})
        if record.record_type in labels:
            label = labels.get(record.record_type, "").strip()
            return f"# {label}" if label else ""
        return getattr(record, "raw_comment", "")

    def _render_cs_info_section(self) -> List[str]:
        self._sync_cs_info_ids_from_uids()
        rendered = ["\n"]
        for atom_id, model_id in sorted(self.cs_info, key=lambda item: (item[1], item[0])):
            if atom_id in self.atoms:
                rendered.append(self._format_line([atom_id, model_id]))
        rendered.append("\n")
        return rendered

    def _render_coeff_section(self, section_name: str, block: SectionBlock) -> List[str]:
        kind = COEFF_SECTION_TYPE_KIND.get(section_name)
        mapping = self.type_id_maps.get(kind or "", {})
        if not kind:
            return [raw for _, raw in block.lines]
        coeff_values = self.coeff_values.get(kind, {})
        if coeff_values:
            type_ids = sorted(self._current_type_ids(kind) | set(coeff_values))
            rendered = ["\n"]
            for type_id in type_ids:
                tokens = coeff_values.get(type_id)
                if not tokens:
                    continue
                label = self.mass_names.get(type_id, "").strip() if kind == "atom" else self.coeff_names.get(kind, {}).get(type_id, "").strip()
                comment = f"# {label}" if label else ""
                rendered.append(self._format_line([type_id, *tokens], comment))
            rendered.append("\n")
            return rendered
        if not mapping:
            return [raw for _, raw in block.lines]

        rendered = ["\n"]
        for _, raw in block.lines:
            main, comment = LammpsDataParser._split_comment(raw)
            if not main.strip():
                continue
            tokens = main.split()
            try:
                old_type = int(tokens[0])
            except (ValueError, IndexError):
                rendered.append(raw)
                continue
            new_type = mapping.get(old_type)
            if new_type is None:
                continue
            tokens[0] = str(new_type)
            if section_name == "Pair Coeffs":
                label = self.mass_names.get(new_type, "").strip()
                comment = f"# {label}" if label else comment
            elif PRIMARY_COEFF_SECTION_TYPE_KIND.get(section_name) == kind:
                labels = self.coeff_names.get(kind, {})
                if new_type in labels:
                    label = labels.get(new_type, "").strip()
                    comment = f"# {label}" if label else ""
            rendered.append(self._format_line(tokens, comment))
        rendered.append("\n")
        return rendered

    def _render_generated_coeff_section(self, section_name: str, kind: str) -> List[str]:
        coeff_values = self.coeff_values.get(kind, {})
        if not coeff_values:
            return []
        lines = [f"\n{section_name}\n"]
        lines.append("\n")
        for type_id in sorted(self._current_type_ids(kind) | set(coeff_values)):
            tokens = coeff_values.get(type_id)
            if not tokens:
                continue
            label = self.mass_names.get(type_id, "").strip() if kind == "atom" else self.coeff_names.get(kind, {}).get(type_id, "").strip()
            comment = f"# {label}" if label else ""
            lines.append(self._format_line([type_id, *tokens], comment))
        lines.append("\n")
        return lines

    def _update_preamble_counts(self, lines: Sequence[str]) -> List[str]:
        count_map = {
            "atoms": len(self.atoms),
            "atom types": self.atom_type_count(),
            "bonds": len(self.bonds),
            "bond types": self.topology_type_count("bond"),
            "angles": len(self.angles),
            "angle types": self.topology_type_count("angle"),
            "dihedrals": len(self.dihedrals),
            "dihedral types": self.topology_type_count("dihedral"),
            "impropers": len(self.impropers),
            "improper types": self.topology_type_count("improper"),
        }
        updated: List[str] = []
        for raw in lines:
            new_line = raw
            stripped = raw.rstrip("\n")
            for label, count in count_map.items():
                match = re.match(rf"^(\s*)(\d+)(\s+{label}\b.*)$", stripped)
                if match:
                    new_line = f"{match.group(1)}{count}{match.group(3)}\n"
                    break
            updated.append(new_line)
        return updated

    def save_as(self, output_path: Path) -> None:
        self.normalize_type_ids()
        if not self.parsed.section_order:
            Path(output_path).write_text("".join(self.original_lines), encoding="utf-8")
            return

        first_header_index = self.parsed.section_order[0].header_index
        out_lines: List[str] = []
        out_lines.extend(self._update_preamble_counts(self.original_lines[:first_header_index]))

        for block in self.parsed.section_order:
            if block.name == "CS-Info":
                continue
            out_lines.append(self.original_lines[block.header_index])
            if block.name == "Masses" or block.name in SECTION_COUNT_LABELS:
                out_lines.extend(self._render_topology_section(block.name))
            elif block.name in COEFF_SECTION_TYPE_KIND:
                out_lines.extend(self._render_coeff_section(block.name, block))
            else:
                out_lines.extend(raw for _, raw in block.lines)

        existing_sections = {block.name for block in self.parsed.section_order}
        generated_sections = [
            ("Pair Coeffs", "atom"),
            ("Bond Coeffs", "bond"),
            ("Angle Coeffs", "angle"),
            ("Dihedral Coeffs", "dihedral"),
            ("Improper Coeffs", "improper"),
        ]
        for section_name, kind in generated_sections:
            if section_name not in existing_sections and self.coeff_values.get(kind):
                out_lines.extend(self._render_generated_coeff_section(section_name, kind))

        if self.cs_info:
            out_lines.append("\nCS-Info\n")
            out_lines.extend(self._render_cs_info_section())

        Path(output_path).write_text("".join(out_lines), encoding="utf-8")


class AtomColorDialog(QtWidgets.QDialog):
    BASIC_COLORS: List[Tuple[int, int, int]] = [
        (255, 255, 255), (144, 144, 144), (0, 0, 0), (255, 13, 13), (61, 255, 0), (48, 80, 248),
        (240, 200, 160), (255, 255, 48), (255, 128, 0), (31, 240, 31), (171, 92, 242), (191, 166, 166),
        (179, 227, 245), (138, 255, 0), (144, 224, 80), (125, 128, 176), (224, 102, 51), (200, 128, 51),
        (255, 181, 181), (204, 128, 255), (217, 255, 255), (128, 209, 227), (143, 64, 212), (191, 194, 199),
        (230, 230, 230), (180, 180, 180), (120, 120, 120), (80, 80, 80), (255, 180, 180), (180, 220, 255),
        (180, 255, 180), (255, 220, 160), (220, 180, 255), (255, 160, 220), (160, 255, 230), (230, 230, 160),
    ]

    def __init__(self, initial_color: Tuple[int, int, int], parent: Optional[QtWidgets.QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("Atom color")
        self._updating = False
        self._color = tuple(max(0, min(255, int(value))) for value in initial_color)

        layout = QtWidgets.QVBoxLayout(self)

        grid = QtWidgets.QGridLayout()
        for index, color in enumerate(self.BASIC_COLORS):
            button = QtWidgets.QPushButton()
            button.setFixedSize(28, 24)
            button.setToolTip(f"RGB {color}")
            button.setStyleSheet(
                f"QPushButton {{ background-color: rgb({color[0]}, {color[1]}, {color[2]}); border: 1px solid #666; }}"
                "QPushButton:hover { border: 2px solid #111; }"
            )
            button.clicked.connect(lambda checked=False, c=color: self._set_color(c))
            grid.addWidget(button, index // 6, index % 6)
        layout.addLayout(grid)

        self.preview = QtWidgets.QLabel()
        self.preview.setFixedHeight(28)
        self.preview.setFrameShape(QtWidgets.QFrame.Box)
        layout.addWidget(self.preview)

        rgb_layout = QtWidgets.QHBoxLayout()
        self.r_spin = self._make_spinbox()
        self.g_spin = self._make_spinbox()
        self.b_spin = self._make_spinbox()
        for label, spinbox in (("R", self.r_spin), ("G", self.g_spin), ("B", self.b_spin)):
            rgb_layout.addWidget(QtWidgets.QLabel(label))
            rgb_layout.addWidget(spinbox)
        layout.addLayout(rgb_layout)

        picker_button = QtWidgets.QPushButton("打开取色器...")
        picker_button.clicked.connect(self._open_color_picker)
        layout.addWidget(picker_button)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._set_color(self._color)

    def _make_spinbox(self) -> QtWidgets.QSpinBox:
        spinbox = QtWidgets.QSpinBox()
        spinbox.setRange(0, 255)
        spinbox.valueChanged.connect(self._spinbox_changed)
        return spinbox

    def _set_color(self, color: Tuple[int, int, int]) -> None:
        self._color = tuple(max(0, min(255, int(value))) for value in color)
        self._updating = True
        try:
            self.r_spin.setValue(self._color[0])
            self.g_spin.setValue(self._color[1])
            self.b_spin.setValue(self._color[2])
        finally:
            self._updating = False
        self.preview.setStyleSheet(
            f"QLabel {{ background-color: rgb({self._color[0]}, {self._color[1]}, {self._color[2]}); border: 1px solid #666; }}"
        )

    def _spinbox_changed(self) -> None:
        if self._updating:
            return
        self._set_color((self.r_spin.value(), self.g_spin.value(), self.b_spin.value()))

    def _open_color_picker(self) -> None:
        current = QtGui.QColor(*self._color)
        color = QtWidgets.QColorDialog.getColor(current, self, "选择 atom 颜色")
        if color.isValid():
            self._set_color((color.red(), color.green(), color.blue()))

    def color(self) -> Tuple[int, int, int]:
        return self._color


class InsertAtomDialog(QtWidgets.QDialog):
    def __init__(self, data_model: DataModel, parent: Optional[QtWidgets.QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("插入原子")
        layout = QtWidgets.QFormLayout(self)

        self.type_combo = QtWidgets.QComboBox()
        for atom_type in sorted(data_model._current_type_ids("atom")):
            name = data_model.mass_names.get(atom_type, "").strip() or "未命名"
            element = infer_element_from_mass(data_model.masses.get(atom_type))
            element_text = element[0] if element else "?"
            self.type_combo.addItem(f"type {atom_type} | {element_text} | {name}", atom_type)
        layout.addRow("插入 atom type:", self.type_combo)

        self.radius_spin = QtWidgets.QDoubleSpinBox()
        self.radius_spin.setRange(0.0, 10000.0)
        self.radius_spin.setDecimals(4)
        self.radius_spin.setValue(1.0)
        layout.addRow("r 距离:", self.radius_spin)

        self.theta_spin = QtWidgets.QDoubleSpinBox()
        self.theta_spin.setRange(0.0, 180.0)
        self.theta_spin.setDecimals(3)
        self.theta_spin.setValue(90.0)
        layout.addRow("theta 极角(+Z=0):", self.theta_spin)

        self.phi_spin = QtWidgets.QDoubleSpinBox()
        self.phi_spin.setRange(-360.0, 360.0)
        self.phi_spin.setDecimals(3)
        self.phi_spin.setValue(0.0)
        layout.addRow("phi 方位角(+X=0):", self.phi_spin)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def atom_type(self) -> int:
        return int(self.type_combo.currentData())

    def spherical_values(self) -> Tuple[float, float, float]:
        return float(self.radius_spin.value()), float(self.theta_spin.value()), float(self.phi_spin.value())


class InsertFragmentDialog(QtWidgets.QDialog):
    def __init__(self, data_model: DataModel, parent: Optional[QtWidgets.QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("插入 fragment")
        layout = QtWidgets.QFormLayout(self)

        self.fragment_combo = QtWidgets.QComboBox()
        for fragment_id, atom_ids in sorted(data_model.fragments.items()):
            type_counts: Dict[int, int] = defaultdict(int)
            for atom_id in atom_ids:
                atom = data_model.atoms.get(atom_id)
                if atom is not None:
                    type_counts[atom.atom_type] += 1
            type_text = ", ".join(f"{type_id}:{count}" for type_id, count in sorted(type_counts.items())[:6])
            self.fragment_combo.addItem(f"fragment {fragment_id} | atoms {len(atom_ids)} | {type_text}", fragment_id)
        layout.addRow("复制已有 fragment:", self.fragment_combo)

        self.radius_spin = QtWidgets.QDoubleSpinBox()
        self.radius_spin.setRange(0.0, 10000.0)
        self.radius_spin.setDecimals(4)
        self.radius_spin.setValue(2.0)
        layout.addRow("r 距离:", self.radius_spin)

        self.theta_spin = QtWidgets.QDoubleSpinBox()
        self.theta_spin.setRange(0.0, 180.0)
        self.theta_spin.setDecimals(3)
        self.theta_spin.setValue(90.0)
        layout.addRow("theta 极角(+Z=0):", self.theta_spin)

        self.phi_spin = QtWidgets.QDoubleSpinBox()
        self.phi_spin.setRange(-360.0, 360.0)
        self.phi_spin.setDecimals(3)
        self.phi_spin.setValue(0.0)
        layout.addRow("phi 方位角(+X=0):", self.phi_spin)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def fragment_id(self) -> int:
        return int(self.fragment_combo.currentData())

    def spherical_values(self) -> Tuple[float, float, float]:
        return float(self.radius_spin.value()), float(self.theta_spin.value()), float(self.phi_spin.value())


class ForceFieldCandidateDialog(QtWidgets.QDialog):
    def __init__(
        self,
        forcefields: Sequence[ForceFieldDefinition],
        kind: str,
        target_element: str = "",
        current_symbol: str = "",
        parent: Optional[QtWidgets.QWidget] = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("候选力场参数")
        self.resize(920, 520)
        self._selected_symbol = ""
        self._selected_forcefield = ""

        layout = QtWidgets.QVBoxLayout(self)
        info = QtWidgets.QLabel(
            f"对象: {kind}"
            + (f"，元素: {target_element}" if target_element else "")
            + (f"，当前 ff-type: {current_symbol}" if current_symbol else "")
        )
        layout.addWidget(info)

        self.table = QtWidgets.QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["forcefield", "symbol", "species", "element/atoms", "charge", "mass", "style", "params"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.itemDoubleClicked.connect(lambda item: self.accept())
        layout.addWidget(self.table)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._populate(forcefields, kind, target_element)

    def _populate(self, forcefields: Sequence[ForceFieldDefinition], kind: str, target_element: str) -> None:
        rows: List[Tuple[str, str, Dict[str, Any]]] = []
        for definition in forcefields:
            entries = {
                "atom": definition.atom_types,
                "bond": definition.bond_types,
                "angle": definition.angle_types,
                "dihedral": definition.dihedral_types,
                "improper": definition.improper_types,
            }.get(kind, {})
            for symbol, entry in sorted(entries.items()):
                if kind == "atom" and target_element:
                    entry_element = _forcefield_entry_element(symbol, entry)
                    if entry_element and entry_element != target_element:
                        continue
                rows.append((definition.display_name, symbol, entry))
        self.table.setRowCount(len(rows))
        for row_index, (ff_name, symbol, entry) in enumerate(rows):
            if kind == "atom":
                element_or_atoms = _forcefield_entry_element(symbol, entry)
                charge = _format_forcefield_value(entry.get("charge"))
                mass = _format_forcefield_value(entry.get("mass"))
                pair = entry.get("pair") if isinstance(entry.get("pair"), dict) else {}
                style = str(pair.get("style") or "")
            else:
                atoms = entry.get("atoms", [])
                element_or_atoms = "-".join(map(str, atoms)) if isinstance(atoms, list) else str(atoms or "")
                charge = ""
                mass = ""
                style = str(entry.get("style") or "")
            values = [
                ff_name,
                symbol,
                str(entry.get("species") or ""),
                element_or_atoms,
                charge,
                mass,
                style,
                _forcefield_params_text(entry),
            ]
            for column, value in enumerate(values):
                item = QtWidgets.QTableWidgetItem(value)
                item.setData(QtCore.Qt.UserRole, symbol)
                item.setData(QtCore.Qt.UserRole + 1, ff_name)
                self.table.setItem(row_index, column, item)
        self.table.resizeColumnsToContents()
        if rows:
            self.table.selectRow(0)

    def selected_symbol(self) -> str:
        row = self.table.currentRow()
        if row < 0:
            return ""
        item = self.table.item(row, 1)
        return "" if item is None else item.text().strip()


OPEN_VIEWER_WINDOWS: List["ViewerMainWindow"] = []
SYNC_GROUP_WINDOWS: List["ViewerMainWindow"] = []


def open_viewer_windows() -> List["ViewerMainWindow"]:
    return [window for window in OPEN_VIEWER_WINDOWS if window is not None and not window.isHidden()]


class SyncGroupDialog(QtWidgets.QDialog):
    def __init__(self, windows: Sequence["ViewerMainWindow"], active: Sequence["ViewerMainWindow"], parent=None):
        super().__init__(parent)
        self.setWindowTitle("同步组管理")
        self.resize(720, 420)
        self._windows = list(windows)
        self._selected: List["ViewerMainWindow"] = []

        layout = QtWidgets.QVBoxLayout(self)
        info = QtWidgets.QLabel(
            "勾选需要同步处理的模型。同步组建立后，Name/ff-type/mass/color/赋予力场参数等 type 级操作会同步到组内模型。"
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        self.table = QtWidgets.QTableWidget(len(self._windows), 4)
        self.table.setHorizontalHeaderLabels(["同步", "窗口", "文件", "状态"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        active_set = set(active)
        for row, window in enumerate(self._windows):
            check_item = QtWidgets.QTableWidgetItem()
            check_item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsUserCheckable)
            check_item.setCheckState(QtCore.Qt.Checked if window in active_set else QtCore.Qt.Unchecked)
            self.table.setItem(row, 0, check_item)
            self.table.setItem(row, 1, QtWidgets.QTableWidgetItem(window.windowTitle()))
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(window.current_file or "-")))
            status = "已打开模型" if window.data_model is not None else "未加载 data"
            self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(status))
        self.table.resizeColumnsToContents()
        layout.addWidget(self.table)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._accept_checked)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _accept_checked(self) -> None:
        self._selected = []
        for row, window in enumerate(self._windows):
            item = self.table.item(row, 0)
            if item is not None and item.checkState() == QtCore.Qt.Checked and window.data_model is not None:
                self._selected.append(window)
        self.accept()

    def selected_windows(self) -> List["ViewerMainWindow"]:
        return list(self._selected)


class FindDialog(QtWidgets.QDialog):
    FIELD_OPTIONS = {
        "atom": [("id", "id"), ("type", "type"), ("mass", "mass")],
        "bond": [("id", "id"), ("type", "type"), ("atom_types", "atom1-type + atom2-type")],
        "angle": [("id", "id"), ("type", "type")],
        "dihedral": [("id", "id"), ("type", "type")],
        "improper": [("id", "id"), ("type", "type")],
    }

    def __init__(self, parent: "ViewerMainWindow"):
        super().__init__(parent)
        self.viewer = parent
        self.setWindowTitle("查找")
        self.resize(420, 180)

        layout = QtWidgets.QVBoxLayout(self)
        form = QtWidgets.QFormLayout()

        self.category_combo = QtWidgets.QComboBox(self)
        self.category_combo.addItems(["atom", "bond", "angle", "dihedral", "improper"])

        self.field_combo = QtWidgets.QComboBox(self)
        self.value_edit = QtWidgets.QLineEdit(self)
        self.value_edit.setPlaceholderText("输入搜索值")

        self.status_label = QtWidgets.QLabel("输入参数后自动搜索。", self)
        self.status_label.setWordWrap(True)

        close_button = QtWidgets.QPushButton("关闭", self)
        close_button.clicked.connect(self.close)

        form.addRow("类别", self.category_combo)
        form.addRow("条件", self.field_combo)
        form.addRow("值", self.value_edit)
        layout.addLayout(form)
        layout.addWidget(self.status_label)
        layout.addWidget(close_button, alignment=QtCore.Qt.AlignRight)

        self.category_combo.currentTextChanged.connect(self._refresh_fields)
        self.field_combo.currentTextChanged.connect(self._update_placeholder)
        self.field_combo.currentTextChanged.connect(self._apply_search)
        self.value_edit.textChanged.connect(self._apply_search)

        self._refresh_fields()

    def _refresh_fields(self) -> None:
        category = self.category_combo.currentText().strip().lower()
        options = self.FIELD_OPTIONS.get(category, [("id", "id")])
        current_value = self.field_combo.currentData()
        self.field_combo.blockSignals(True)
        self.field_combo.clear()
        for value, label in options:
            self.field_combo.addItem(label, value)
        if current_value is not None:
            idx = self.field_combo.findData(current_value)
            if idx >= 0:
                self.field_combo.setCurrentIndex(idx)
        self.field_combo.blockSignals(False)
        self._update_placeholder()
        self._apply_search()

    def _update_placeholder(self) -> None:
        field_name = self.field_combo.currentData()
        if field_name == "atom_types":
            self.value_edit.setPlaceholderText("例如: 1 2 或 1,2")
            return
        self.value_edit.setPlaceholderText("输入搜索值")

    def _apply_search(self) -> None:
        field_name = self.field_combo.currentData()
        if field_name is None:
            self.status_label.setText("请选择搜索条件。")
            return
        message = self.viewer.apply_find_request(
            self.category_combo.currentText().strip().lower(),
            str(field_name),
            self.value_edit.text(),
        )
        self.status_label.setText(message)

    def focus_input(self) -> None:
        self.show()
        self.raise_()
        self.activateWindow()
        self.value_edit.setFocus()
        self.value_edit.selectAll()


class ViewerMainWindow(QtWidgets.QMainWindow):
    def __init__(self, initial_file: Optional[Path] = None):
        super().__init__()
        OPEN_VIEWER_WINDOWS.append(self)
        self.setWindowTitle("LAMMPS Data Viewer")
        self.resize(1500, 920)

        self.frame = QtWidgets.QFrame(self)
        self.layout = QtWidgets.QVBoxLayout(self.frame)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.loading_label = QtWidgets.QLabel("正在初始化 3D 引擎...", self.frame)
        self.loading_label.setAlignment(QtCore.Qt.AlignCenter)
        self.layout.addWidget(self.loading_label)
        self.setCentralWidget(self.frame)
        self.setAcceptDrops(True)
        self.frame.setAcceptDrops(True)

        self.data_model: Optional[DataModel] = None
        self.current_file: Optional[Path] = None
        self.plotter = None
        self._plotter_ready = False
        self._pending_initial_file = Path(initial_file) if initial_file else None
        self.points_mesh: Optional[pv.PolyData] = None
        self.atom_mesh: Optional[pv.PolyData] = None
        self.bond_mesh: Optional[pv.PolyData] = None
        self.atom_actor = None
        self.bond_actor = None
        self.selected_actor = None
        self.selected_bond_actor = None
        self.selected_label_actor = None
        self.single_label_actor = None
        self.box_actor = None
        self.find_dialog: Optional[FindDialog] = None
        self.settings = QtCore.QSettings("CASH", "LAMMPSDataViewerV2")
        self.loaded_forcefields: Dict[str, ForceFieldDefinition] = {}
        self.forcefield_library_base: Optional[Path] = self._stored_forcefield_library_base()
        self.saved_forcefield_menu: Optional[QtWidgets.QMenu] = None
        self.last_picked_atom: Optional[int] = None
        self.last_picked_bond: Optional[int] = None
        self._right_press_pos: Optional[QtCore.QPoint] = None
        self._right_dragged = False
        self._left_press_pos: Optional[QtCore.QPoint] = None
        self._left_dragged = False
        self._selection_band = None
        self._updating_properties_table = False
        self._updating_type_tables = False
        self._properties_context_atom_type: Optional[int] = None
        self._type_table_anchor: Dict[str, int] = {}
        self._pick_tolerance_px = 5
        self._atom_radius_scale = 0.88
        self._bond_radius = 0.14
        self._selection_shell_scale = 1.18
        self._show_box = True
        self._display_mode = "in-cell"
        self._display_coord_cache: Dict[int, List[float]] = {}
        self._hide_cross_boundary_bonds = True
        self._scene_visible_bonds: List[BondRecord] = []
        self._sphere_geom = None
        self._sphere_geom_fast = None
        self._setup_ui()
        self.scan_forcefield_library(show_message=False)
        QtCore.QTimer.singleShot(0, self._initialize_3d_backend)

    def _initialize_3d_backend(self) -> None:
        if self._plotter_ready:
            return
        self.statusBar().showMessage("正在加载 PyVista/VTK 3D 引擎...")
        QtWidgets.QApplication.processEvents()
        try:
            ensure_3d_imports()
            self.plotter = QtInteractor(self.frame)
            self.plotter.interactor.setAcceptDrops(True)
            self.layout.replaceWidget(self.loading_label, self.plotter.interactor)
            self.loading_label.deleteLater()
            self._selection_band = QtWidgets.QRubberBand(QtWidgets.QRubberBand.Rectangle, self.plotter.interactor)
            self._selection_band.setStyleSheet(
                "QRubberBand {"
                "border: 2px solid rgba(0, 120, 215, 220);"
                "background-color: rgba(0, 120, 215, 0);"
                "}"
            )
            self._sphere_geom = pv.Sphere(radius=1.0, theta_resolution=28, phi_resolution=28)
            self._sphere_geom_fast = pv.Sphere(radius=1.0, theta_resolution=12, phi_resolution=12)
            self._setup_plotter()
            self._connect_events()
            self._plotter_ready = True
            self.statusBar().showMessage("3D 引擎初始化完成。")
            if self._pending_initial_file is not None:
                pending = self._pending_initial_file
                self._pending_initial_file = None
                self.load_data_file(pending)
        except Exception as exc:
            self.statusBar().showMessage("3D 引擎初始化失败。")
            QtWidgets.QMessageBox.critical(self, "3D 初始化失败", str(exc))

    def _setup_plotter(self) -> None:
        if self.plotter is None:
            return
        self.plotter.set_background("white")
        self.plotter.enable_anti_aliasing()
        self.plotter.show_axes()
        try:
            self.plotter.enable_custom_trackball_style(
                left="rotate",
                shift_left="pan",
                control_left="rotate",
                right="rotate",
                shift_right="rotate",
                control_right="rotate",
                middle="pan",
                shift_middle="pan",
                control_middle="pan",
            )
        except Exception:
            pass

    def _stored_forcefield_library_base(self) -> Optional[Path]:
        value = self.settings.value("forcefield/library_base", "", type=str)
        if not value:
            return None
        return Path(value)

    def forcefield_library_dir(self) -> Optional[Path]:
        if self.forcefield_library_base is None:
            return None
        return self.forcefield_library_base / "forcefields"

    def _is_forcefield_file(self, path: Path) -> bool:
        return path.suffix.lower() in FORCEFIELD_FILE_SUFFIXES

    def _is_msi2lmp_input_file(self, path: Path) -> bool:
        return path.suffix.lower() in MSI2LMP_INPUT_SUFFIXES

    def _stored_msi2lmp_path(self) -> str:
        stored = self.settings.value("msi2lmp/exe_path", "", type=str)
        if stored and Path(stored).exists():
            return stored
        found = shutil.which("msi2lmp.exe") or shutil.which("msi2lmp")
        return found or ""

    def _msi2lmp_args_template(self) -> str:
        template = self.settings.value("msi2lmp/args_template", "{stem} -class I -frc cvff -i >data.dat", type=str)
        if template.strip() in {"{stem} -class I -frc cvff", "{stem} -class 1 -frc cvff -i"}:
            return "{stem} -class I -frc cvff -i >data.dat"
        return template

    def _msi2lmp_output_template(self) -> str:
        return self.settings.value("msi2lmp/output_template", "data.dat", type=str)

    def _msi2lmp_work_dir(self, exe_path: str) -> Path:
        return Path(exe_path).parent

    def _choose_msi2lmp_exe(self) -> str:
        current = self._stored_msi2lmp_path()
        if current:
            return current
        exe_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "选择 msi2lmp.exe",
            str(Path.home()),
            "msi2lmp (msi2lmp.exe msi2lmp);;Executable (*.exe);;All files (*.*)",
        )
        if exe_path:
            self.settings.setValue("msi2lmp/exe_path", exe_path)
        return exe_path

    def configure_materials_studio_conversion_dialog(self) -> None:
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Materials Studio 转换设置")
        layout = QtWidgets.QVBoxLayout(dialog)
        form = QtWidgets.QFormLayout()

        exe_edit = QtWidgets.QLineEdit(self._stored_msi2lmp_path())
        browse_button = QtWidgets.QPushButton("浏览...")
        exe_row = QtWidgets.QHBoxLayout()
        exe_row.addWidget(exe_edit, 1)
        exe_row.addWidget(browse_button)
        exe_widget = QtWidgets.QWidget()
        exe_widget.setLayout(exe_row)
        form.addRow("msi2lmp.exe:", exe_widget)

        args_edit = QtWidgets.QLineEdit(self._msi2lmp_args_template())
        form.addRow("命令参数模板:", args_edit)

        output_edit = QtWidgets.QLineEdit(self._msi2lmp_output_template())
        form.addRow("无重定向时输出文件:", output_edit)
        layout.addLayout(form)

        note = QtWidgets.QLabel(
            "默认流程不会再弹出参数窗口。可用 {stem} 表示文件名主体；"
            "如果参数中包含 >data.dat，则输出文件设置会被忽略。"
        )
        note.setWordWrap(True)
        layout.addWidget(note)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def browse_exe() -> None:
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                dialog,
                "选择 msi2lmp.exe",
                str(Path(exe_edit.text()).parent if exe_edit.text().strip() else Path.home()),
                "msi2lmp (msi2lmp.exe msi2lmp);;Executable (*.exe);;All files (*.*)",
            )
            if path:
                exe_edit.setText(path)

        browse_button.clicked.connect(browse_exe)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        exe_path = exe_edit.text().strip()
        if exe_path:
            self.settings.setValue("msi2lmp/exe_path", exe_path)
        self.settings.setValue("msi2lmp/args_template", args_edit.text().strip() or "{stem} -class I -frc cvff -i >data.dat")
        self.settings.setValue("msi2lmp/output_template", output_edit.text().strip() or "data.dat")
        self.statusBar().showMessage("Materials Studio 转换设置已更新。")

    def _forcefield_key(self, definition: ForceFieldDefinition) -> str:
        return f"{definition.name}::{definition.version or '-'}::{definition.source_path}"

    def _refresh_saved_forcefield_menu(self) -> None:
        if self.saved_forcefield_menu is None:
            return
        self.saved_forcefield_menu.clear()
        library_dir = self.forcefield_library_dir()
        if library_dir is None:
            action = self.saved_forcefield_menu.addAction("未设置力场库文件夹")
            action.setEnabled(False)
            return
        if not library_dir.exists():
            action = self.saved_forcefield_menu.addAction(f"目录不存在: {library_dir}")
            action.setEnabled(False)
            return
        files = sorted(
            path
            for path in library_dir.iterdir()
            if path.is_file() and path.suffix.lower() in {".yaml", ".yml", ".json"}
        )
        if not files:
            action = self.saved_forcefield_menu.addAction("暂无已保存力场")
            action.setEnabled(False)
            return
        for path in files:
            action = self.saved_forcefield_menu.addAction(path.stem)
            action.setToolTip(str(path))
            action.triggered.connect(lambda checked=False, p=path: self.load_saved_forcefield(p))

    def scan_forcefield_library(self, show_message: bool = True) -> None:
        library_dir = self.forcefield_library_dir()
        loaded = 0
        if library_dir is not None and library_dir.exists():
            for path in sorted(library_dir.iterdir()):
                if not path.is_file() or path.suffix.lower() not in {".yaml", ".yml", ".json"}:
                    continue
                try:
                    definition = load_forcefield_definition(path)
                except Exception:
                    continue
                self.loaded_forcefields[self._forcefield_key(definition)] = definition
                loaded += 1
        self._refresh_saved_forcefield_menu()
        if show_message:
            if library_dir is None:
                self.statusBar().showMessage("尚未设置力场库文件夹。")
            else:
                self.statusBar().showMessage(f"已扫描力场库: {library_dir}，加载 {loaded} 个力场。")

    def choose_forcefield_library_folder(self) -> Optional[Path]:
        start_dir = str(self.forcefield_library_base or Path.home())
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "选择力场库根目录", start_dir)
        if not folder:
            return None
        base = Path(folder)
        library_dir = base / "forcefields"
        try:
            library_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "无法创建力场库", str(exc))
            return None
        self.forcefield_library_base = base
        self.settings.setValue("forcefield/library_base", str(base))
        self._refresh_saved_forcefield_menu()
        self.statusBar().showMessage(f"力场库已设置为: {library_dir}")
        return library_dir

    def import_forcefield_dialog(self) -> None:
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "导入力场文件",
            str(Path.home()),
            "Force field (*.yaml *.yml *.json *.xlsx);;All files (*.*)",
        )
        if file_path:
            self.import_forcefield_file(Path(file_path), ask_to_save=True)

    def import_materials_studio_dialog(self) -> None:
        file_paths, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "导入 Materials Studio car/cor/mdf",
            str(Path.cwd()),
            "Materials Studio (*.car *.cor *.mdf);;All files (*.*)",
        )
        if not file_paths:
            return
        data_paths = self.convert_materials_studio_files([Path(path) for path in file_paths])
        self.open_data_paths(data_paths, prefer_current=True)

    def _unique_msi2lmp_inputs(self, paths: Sequence[Path]) -> List[Path]:
        groups: Dict[Tuple[str, str], List[Path]] = {}
        for path in paths:
            if not self._is_msi2lmp_input_file(path):
                continue
            try:
                parent_key = str(path.parent.resolve()).lower()
            except Exception:
                parent_key = str(path.parent).lower()
            key = (parent_key, path.stem.lower())
            groups.setdefault(key, []).append(path)

        representatives: List[Path] = []
        priority = {".car": 0, ".mdf": 1, ".cor": 2}
        for group_paths in groups.values():
            representatives.append(sorted(group_paths, key=lambda item: priority.get(item.suffix.lower(), 99))[0])
        return representatives

    def convert_materials_studio_files(self, paths: Sequence[Path]) -> List[Path]:
        data_paths: List[Path] = []
        for path in self._unique_msi2lmp_inputs(paths):
            data_path = self.convert_materials_studio_file(path)
            if data_path is not None:
                data_paths.append(data_path)
        return data_paths

    def _looks_like_lammps_data_file(self, path: Path) -> bool:
        if not path.exists() or not path.is_file():
            return False
        try:
            if path.stat().st_size < 32:
                return False
            has_atom_count = False
            has_atoms_section = False
            has_box = False
            with path.open("r", encoding="utf-8", errors="replace") as handle:
                for line_index, line in enumerate(handle):
                    stripped = line.strip()
                    if re.match(r"^\d+\s+atoms\b", stripped):
                        has_atom_count = True
                    if any(token in stripped for token in ("xlo xhi", "ylo yhi", "zlo zhi")):
                        has_box = True
                    header = LammpsDataParser._canonical_header(line)
                    if header and header[0] == "Atoms":
                        has_atoms_section = True
                    if has_atom_count and has_atoms_section:
                        return True
                    if line_index > 10000 and has_atom_count and has_box:
                        return True
            return has_atom_count and has_atoms_section
        except Exception:
            return False

    def convert_materials_studio_file(self, input_path: Path) -> Optional[Path]:
        input_path = Path(input_path)
        if not input_path.exists():
            QtWidgets.QMessageBox.warning(self, "MS 导入失败", f"文件不存在:\n{input_path}")
            return None
        exe_path = self._choose_msi2lmp_exe()
        if not exe_path:
            QtWidgets.QMessageBox.warning(self, "MS 导入失败", "未指定 msi2lmp.exe。")
            return None

        source_dir = input_path.parent
        work_dir = self._msi2lmp_work_dir(exe_path)
        stem = input_path.stem
        try:
            work_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "msi2lmp 工作目录无效", str(exc))
            return None

        pair_files = [source_dir / f"{stem}{suffix}" for suffix in (".car", ".cor", ".mdf")]
        existing_pairs = [path.name for path in pair_files if path.exists()]
        if not existing_pairs:
            QtWidgets.QMessageBox.warning(self, "MS 导入失败", "未找到同名 .car/.cor/.mdf 文件。")
            return None
        if not any((source_dir / f"{stem}{suffix}").exists() for suffix in (".cor", ".mdf")):
            reply = QtWidgets.QMessageBox.question(
                self,
                "缺少拓扑文件提醒",
                f"仅找到 {existing_pairs}。\nmsi2lmp 通常需要同名 .car 和 .mdf/.cor 文件。是否仍尝试转换？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                return None

        try:
            for source in pair_files:
                if source.exists():
                    target = work_dir / source.name
                    if source.resolve() != target.resolve():
                        shutil.copy2(source, target)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "复制 MS 文件失败", str(exc))
            return None

        template = self._msi2lmp_args_template().strip() or "{stem} -class I -frc cvff -i >data.dat"
        stdout_target: Optional[Path] = None
        command_template = template
        redirect_match = re.search(r">\s*(\S+)\s*$", command_template)
        if redirect_match:
            stdout_target = work_dir / redirect_match.group(1).format(stem=stem, file=str(input_path), dir=str(work_dir))
            command_template = command_template[: redirect_match.start()].strip()
        else:
            output_template = self._msi2lmp_output_template().strip() or "data.dat"
            stdout_target = work_dir / output_template.format(stem=stem, file=str(input_path), dir=str(work_dir))
        args = [part.format(stem=stem, file=str(input_path), dir=str(work_dir)) for part in shlex.split(command_template, posix=False)]
        command_display = f"{exe_path} {' '.join(args)}"
        if stdout_target is not None:
            command_display = f"{command_display} > {stdout_target.name}"
            try:
                if stdout_target.exists() and stdout_target.parent.resolve() == work_dir.resolve():
                    stdout_target.unlink()
            except Exception:
                pass
        start_time = time.time()
        try:
            if stdout_target is not None:
                with stdout_target.open("w", encoding="utf-8", errors="replace") as stdout_handle:
                    completed = subprocess.run(
                        [exe_path, *args],
                        cwd=str(work_dir),
                        text=True,
                        stdout=stdout_handle,
                        stderr=subprocess.PIPE,
                        timeout=300,
                        check=False,
                    )
            else:
                completed = subprocess.run(
                    [exe_path, *args],
                    cwd=str(work_dir),
                    text=True,
                    capture_output=True,
                    timeout=300,
                    check=False,
                )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "msi2lmp 执行失败", str(exc))
            return None

        data_candidates = sorted(
            [path for path in list(work_dir.glob("*.data")) + list(work_dir.glob("*.dat")) if path.is_file()],
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        changed_candidates = [path for path in data_candidates if path.stat().st_mtime >= start_time - 2.0]
        preferred = []
        if stdout_target is not None:
            preferred.append(stdout_target)
        preferred.extend([work_dir / f"{stem}.data", work_dir / f"{stem}.dat"])
        preferred.extend([work_dir / "data.data", work_dir / "data.dat"])
        data_path = None
        for candidate in preferred:
            if (
                candidate.exists()
                and candidate.stat().st_mtime >= start_time - 2.0
                and self._looks_like_lammps_data_file(candidate)
            ):
                data_path = candidate
                break
        if data_path is None:
            data_path = next(
                (path for path in changed_candidates if path.exists() and self._looks_like_lammps_data_file(path)),
                None,
            )
        if data_path is not None:
            try:
                target_name = data_path.name
                if data_path.name.lower() in {"data.dat", "data.data"}:
                    target_name = f"{stem}{data_path.suffix or '.dat'}"
                target_in_source = source_dir / target_name
                if data_path.resolve() != target_in_source.resolve():
                    shutil.copy2(data_path, target_in_source)
                    data_path = target_in_source
            except Exception:
                pass
        if data_path is None and data_candidates:
            candidate_names = "\n".join(str(path) for path in data_candidates[:10])
            reply = QtWidgets.QMessageBox.question(
                self,
                "未自动定位 data 文件",
                "msi2lmp 已结束，但程序没有自动识别出本次生成的 data 文件。\n\n"
                f"当前目录中的候选文件:\n{candidate_names}\n\n是否手动选择一个 .data/.dat 文件加载？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes,
            )
            if reply == QtWidgets.QMessageBox.Yes:
                selected_path, _ = QtWidgets.QFileDialog.getOpenFileName(
                    self,
                    "选择 msi2lmp 生成的 data 文件",
                    str(work_dir),
                    "LAMMPS data (*.data *.dat *.lmp *.txt);;All files (*.*)",
                )
                if selected_path:
                    data_path = Path(selected_path)
        if data_path is None:
            log_text = (completed.stdout or "") + "\n" + (completed.stderr or "")
            QtWidgets.QMessageBox.critical(
                self,
                "msi2lmp 转换失败",
                f"命令返回码: {completed.returncode}\n"
                f"工作目录: {work_dir}\n"
                f"命令: {command_display}\n\n"
                f"输出:\n{log_text[-4000:]}",
            )
            return None
        if completed.returncode != 0:
            log_text = ((completed.stdout or "") + "\n" + (completed.stderr or "")).strip()
            QtWidgets.QMessageBox.warning(
                self,
                "msi2lmp 转换警告",
                f"msi2lmp 返回码为 {completed.returncode}，但已生成 data 文件，将继续加载：\n{data_path}\n\n"
                f"输出摘要:\n{log_text[-3000:]}",
            )
        self.statusBar().showMessage(f"msi2lmp 转换完成: {data_path}")
        return data_path

    def import_forcefield_file(self, path: Path, ask_to_save: bool = True) -> Optional[ForceFieldDefinition]:
        try:
            definition = load_forcefield_definition(path)
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "无法导入力场", f"{path}\n\n{exc}")
            return None
        self.loaded_forcefields[self._forcefield_key(definition)] = definition
        if ask_to_save:
            result = QtWidgets.QMessageBox.question(
                self,
                "保存力场",
                f"已导入力场: {definition.display_name}\n\n是否保存到程序力场库，方便下次直接加载？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Yes,
            )
            if result == QtWidgets.QMessageBox.Cancel:
                self.loaded_forcefields.pop(self._forcefield_key(definition), None)
                self.statusBar().showMessage("已取消导入力场。")
                return None
            if result == QtWidgets.QMessageBox.Yes:
                if self.save_forcefield_to_library(definition) is None:
                    return definition
        self.statusBar().showMessage(
            f"已导入力场 {definition.display_name}: "
            f"atom {len(definition.atom_types)}, bond {len(definition.bond_types)}, angle {len(definition.angle_types)}"
        )
        return definition

    def load_saved_forcefield(self, path: Path) -> None:
        definition = self.import_forcefield_file(path, ask_to_save=False)
        if definition is not None:
            QtWidgets.QMessageBox.information(
                self,
                "已加载力场",
                f"{definition.display_name}\n"
                f"atom types: {len(definition.atom_types)}\n"
                f"bond types: {len(definition.bond_types)}\n"
                f"angle types: {len(definition.angle_types)}",
            )

    def save_forcefield_to_library(self, definition: ForceFieldDefinition) -> Optional[Path]:
        if yaml is None:
            QtWidgets.QMessageBox.warning(self, "无法保存力场", "当前 Python 环境缺少 PyYAML，无法写入 YAML 力场库。")
            return None
        library_dir = self.forcefield_library_dir()
        if library_dir is None:
            library_dir = self.choose_forcefield_library_folder()
            if library_dir is None:
                return None
        try:
            library_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "无法创建力场库", str(exc))
            return None
        version_part = f"_{definition.version}" if definition.version else ""
        file_name = _safe_forcefield_filename(f"{definition.name}{version_part}") + ".yaml"
        target = library_dir / file_name
        if target.exists():
            prompt = QtWidgets.QMessageBox(self)
            prompt.setWindowTitle("力场已存在")
            prompt.setText(f"力场库中已存在同名文件:\n{target}")
            overwrite_button = prompt.addButton("覆盖", QtWidgets.QMessageBox.AcceptRole)
            copy_button = prompt.addButton("另存为新文件", QtWidgets.QMessageBox.ActionRole)
            prompt.addButton("仅本次使用", QtWidgets.QMessageBox.RejectRole)
            prompt.exec()
            clicked = prompt.clickedButton()
            if clicked == copy_button:
                stem = target.stem
                index = 2
                while (library_dir / f"{stem}_{index}.yaml").exists():
                    index += 1
                target = library_dir / f"{stem}_{index}.yaml"
            elif clicked != overwrite_button:
                return None
        try:
            with target.open("w", encoding="utf-8") as handle:
                yaml.safe_dump(definition.to_dict(), handle, allow_unicode=True, sort_keys=False)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "保存力场失败", str(exc))
            return None
        definition.source_path = str(target)
        self.loaded_forcefields[self._forcefield_key(definition)] = definition
        self._refresh_saved_forcefield_menu()
        self.statusBar().showMessage(f"力场已保存到: {target}")
        return target

    def _loaded_forcefield_list(self) -> List[ForceFieldDefinition]:
        seen: Set[str] = set()
        result: List[ForceFieldDefinition] = []
        for definition in self.loaded_forcefields.values():
            key = definition.display_name
            if key in seen:
                continue
            seen.add(key)
            result.append(definition)
        return sorted(result, key=lambda item: item.display_name.lower())

    def _choose_loaded_forcefield(self) -> Optional[ForceFieldDefinition]:
        forcefields = self._loaded_forcefield_list()
        if not forcefields:
            QtWidgets.QMessageBox.information(self, "未加载力场", "请先导入或加载一个力场文件。")
            return None
        if len(forcefields) == 1:
            return forcefields[0]
        names = [definition.display_name for definition in forcefields]
        selected, ok = QtWidgets.QInputDialog.getItem(self, "选择力场", "请选择用于赋值的单一力场:", names, 0, False)
        if not ok:
            return None
        for definition in forcefields:
            if definition.display_name == selected:
                return definition
        return None

    def _type_table_kind(self, table: object) -> Optional[str]:
        table_to_kind = {
            self.atom_type_table: "atom",
            self.bond_type_table: "bond",
            self.angle_type_table: "angle",
            self.dihedral_type_table: "dihedral",
            self.improper_type_table: "improper",
        }
        return table_to_kind.get(table)

    def _candidate_target_element(self, kind: str, type_id: Optional[int]) -> str:
        if self.data_model is None or kind != "atom" or type_id is None:
            return ""
        element_info = infer_element_from_mass(self.data_model.masses.get(type_id))
        return "" if element_info is None else element_info[0]

    def show_forcefield_candidates(self) -> None:
        if not self._loaded_forcefield_list():
            QtWidgets.QMessageBox.information(self, "未加载力场", "请先导入或加载一个力场文件。")
            return
        kind = self.data_model.selection_kind if self.data_model is not None else None
        if kind not in ("atom", "bond", "angle", "dihedral", "improper"):
            kind = "atom"
        type_id = None
        current_symbol = ""
        target_element = ""
        if self.data_model is not None:
            if kind == "atom" and self.data_model.selected_atoms:
                selected_types = sorted({self.data_model.atoms[aid].atom_type for aid in self.data_model.selected_atoms if aid in self.data_model.atoms})
                if len(selected_types) == 1:
                    type_id = selected_types[0]
                    current_symbol = self.data_model.mass_labels.get(type_id, "")
            elif kind in TOPOLOGY_KIND_TO_SECTION:
                selected_ids = self.data_model.topology_id_set(kind)
                selected_records = [record for record in self.data_model.topology_records(kind) if record.record_id in selected_ids]
                selected_types = sorted({record.record_type for record in selected_records})
                if len(selected_types) == 1:
                    type_id = selected_types[0]
                    current_symbol = self.data_model.coeff_labels.get(kind, {}).get(type_id, "")
            target_element = self._candidate_target_element(kind, type_id)
        dialog = ForceFieldCandidateDialog(self._loaded_forcefield_list(), kind, target_element, current_symbol, self)
        dialog.exec()

    def _choose_candidate_for_type(self, kind: str, type_id: int) -> None:
        if self.data_model is None:
            return
        forcefields = self._loaded_forcefield_list()
        if not forcefields:
            QtWidgets.QMessageBox.information(self, "未加载力场", "请先导入或加载一个力场文件。")
            return
        current_symbol = (
            self.data_model.mass_labels.get(type_id, "")
            if kind == "atom"
            else self.data_model.coeff_labels.get(kind, {}).get(type_id, "")
        )
        dialog = ForceFieldCandidateDialog(
            forcefields,
            kind,
            self._candidate_target_element(kind, type_id),
            current_symbol,
            self,
        )
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        symbol = dialog.selected_symbol()
        if not symbol:
            return
        if kind == "atom":
            self.data_model.set_ff_type_for_type(type_id, symbol)
            synced = self._sync_atom_type_property(type_id, "ff", symbol)
        else:
            self.data_model.set_coeff_label_for_type(kind, type_id, symbol)
            synced = self._sync_topology_type_property(kind, type_id, "ff", symbol)
        self.update_summary_panel()
        self.update_properties_panel()
        self.statusBar().showMessage(f"已将 {kind} type={type_id} 的 ff-type 设置为 {symbol}，同步 {synced} 个模型")

    def import_ff_mapping_dialog(self) -> None:
        if self.data_model is None:
            QtWidgets.QMessageBox.information(self, "未打开模型", "请先打开 data 文件。")
            return
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "导入 ff-type 映射",
            str(Path.home()),
            "Mapping (*.mod *.txt *.ini);;All files (*.*)",
        )
        if not file_path:
            return
        try:
            mappings = self._read_ff_mapping_file(Path(file_path))
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "映射文件无效", str(exc))
            return
        changed = self._apply_ff_mapping(mappings)
        self.update_summary_panel()
        self.update_properties_panel()
        self.statusBar().showMessage(f"已应用 ff-type 映射，更新 {changed} 个 type 标签。")

    def _read_ff_mapping_file(self, path: Path) -> List[Tuple[str, str]]:
        mappings: List[Tuple[str, str]] = []
        for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith(";"):
                continue
            if line.startswith("[") and line.endswith("]"):
                continue
            if "=" not in line:
                continue
            left, right = line.split("=", 1)
            pattern = left.strip()
            symbol = right.strip()
            if pattern and symbol:
                mappings.append((pattern, symbol))
        if not mappings:
            raise ValueError("未读取到有效映射。请使用 pattern=symbol 格式，例如 Hw*=h*。")
        return mappings

    def _apply_ff_mapping(self, mappings: Sequence[Tuple[str, str]]) -> int:
        if self.data_model is None:
            return 0
        changed = 0
        for atom_type, label in list(self.data_model.mass_labels.items()):
            for pattern, symbol in mappings:
                if fnmatch.fnmatchcase(label, pattern) and label != symbol:
                    self.data_model.mass_labels[atom_type] = symbol
                    changed += 1
                    break
        for kind in ("bond", "angle", "dihedral", "improper"):
            for type_id, label in list(self.data_model.coeff_labels.get(kind, {}).items()):
                for pattern, symbol in mappings:
                    if fnmatch.fnmatchcase(label, pattern) and label != symbol:
                        self.data_model.coeff_labels[kind][type_id] = symbol
                        changed += 1
                        break
        if changed:
            self.data_model.dirty = True
        return changed

    def apply_forcefield_dialog(self) -> None:
        if self.data_model is None:
            QtWidgets.QMessageBox.information(self, "未打开模型", "请先打开 data 文件。")
            return
        definition = self._choose_loaded_forcefield()
        if definition is None:
            return
        reply = QtWidgets.QMessageBox.question(
            self,
            "赋予力场参数",
            f"将按当前 ff-type 使用 {definition.display_name} 写入 charge/mass/coeff 参数。\n"
            "未匹配到的 ff-type 会保留原状。是否继续？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return
        summary = self.data_model.apply_forcefield(definition)
        synced = self._sync_apply_forcefield(definition)
        self.update_summary_panel()
        self.update_selection_panel()
        missing = summary.get("missing", [])
        unsupported = summary.get("unsupported", [])
        message = (
            f"已赋予力场参数:\n"
            f"atom: {summary['atom']}\n"
            f"bond: {summary['bond']}\n"
            f"angle: {summary['angle']}\n"
            f"dihedral: {summary['dihedral']}\n"
            f"improper: {summary['improper']}\n"
            f"未匹配: {len(missing)}\n"
            f"暂不支持写入: {len(unsupported)}\n"
            f"同步模型: {synced}"
        )
        details = "\n".join([*missing[:30], *unsupported[:30]])
        if details:
            message += "\n\n前 30 项:\n" + details
        QtWidgets.QMessageBox.information(self, "赋予力场完成", message)

    def _setup_ui(self) -> None:
        self.statusBar().showMessage("打开一个 LAMMPS data 文件开始。左键单击选择，左键拖动框选，右键拖动旋转，中键平移。")
        self._make_actions()
        self._make_menus()
        self._make_toolbar()
        self._make_side_panel()

    def _make_actions(self) -> None:
        self.open_action = QtGui.QAction("打开", self)
        self.open_action.setShortcut("Ctrl+O")
        self.open_action.triggered.connect(self.open_file_dialog)

        self.open_new_window_action = QtGui.QAction("在新窗口打开...", self)
        self.open_new_window_action.triggered.connect(self.open_file_new_window_dialog)

        self.import_ms_action = QtGui.QAction("导入 Materials Studio car/cor...", self)
        self.import_ms_action.triggered.connect(self.import_materials_studio_dialog)

        self.configure_ms_action = QtGui.QAction("Materials Studio 转换设置...", self)
        self.configure_ms_action.triggered.connect(self.configure_materials_studio_conversion_dialog)

        self.find_action = QtGui.QAction("Find", self)
        self.find_action.setShortcut("Ctrl+F")
        self.find_action.setShortcutContext(QtCore.Qt.WindowShortcut)
        self.find_action.triggered.connect(self.open_find_dialog)

        self.save_as_action = QtGui.QAction("另存为", self)
        self.save_as_action.setShortcut("Ctrl+Shift+S")
        self.save_as_action.triggered.connect(self.save_as_dialog)

        self.exit_action = QtGui.QAction("退出", self)
        self.exit_action.setShortcut("Ctrl+Q")
        self.exit_action.triggered.connect(self.close)

        self.tile_windows_action = QtGui.QAction("平铺已打开窗口", self)
        self.tile_windows_action.triggered.connect(self.tile_open_windows)

        self.sync_group_action = QtGui.QAction("同步组管理...", self)
        self.sync_group_action.triggered.connect(self.open_sync_group_dialog)

        self.hide_selected_action = QtGui.QAction("隐藏选中原子", self)
        self.hide_selected_action.triggered.connect(self.hide_selected)

        self.isolate_selected_action = QtGui.QAction("仅显示选中原子", self)
        self.isolate_selected_action.triggered.connect(self.isolate_selected)

        self.show_all_action = QtGui.QAction("显示全部", self)
        self.show_all_action.triggered.connect(self.show_all)

        self.toggle_box_action = QtGui.QAction("显示盒子", self)
        self.toggle_box_action.setCheckable(True)
        self.toggle_box_action.setChecked(True)
        self.toggle_box_action.triggered.connect(self.toggle_box_visibility)

        self.toggle_cross_boundary_bonds_action = QtGui.QAction("隐藏跨边界长键", self)
        self.toggle_cross_boundary_bonds_action.setCheckable(True)
        self.toggle_cross_boundary_bonds_action.setChecked(True)
        self.toggle_cross_boundary_bonds_action.triggered.connect(self.toggle_cross_boundary_bonds)

        self.display_in_cell_action = QtGui.QAction("显示模式: in-cell", self)
        self.display_in_cell_action.setCheckable(True)
        self.display_in_cell_action.setChecked(True)
        self.display_in_cell_action.triggered.connect(lambda: self.set_display_mode("in-cell"))

        self.display_default_action = QtGui.QAction("显示模式: default", self)
        self.display_default_action.setCheckable(True)
        self.display_default_action.triggered.connect(lambda: self.set_display_mode("default"))

        self.display_mode_group = QtGui.QActionGroup(self)
        self.display_mode_group.setExclusive(True)
        self.display_mode_group.addAction(self.display_in_cell_action)
        self.display_mode_group.addAction(self.display_default_action)

        self.clear_selection_action = QtGui.QAction("清空选择", self)
        self.clear_selection_action.triggered.connect(self.clear_selection)

        self.select_fragment_action = QtGui.QAction("选择所在 fragment", self)
        self.select_fragment_action.triggered.connect(self.select_fragment)

        self.change_type_action = QtGui.QAction("修改选中 type", self)
        self.change_type_action.triggered.connect(self.change_selected_type)

        self.add_shell_action = QtGui.QAction("添加 Shell 原子", self)
        self.add_shell_action.triggered.connect(self.add_shell_atoms)

        self.insert_atom_action = QtGui.QAction("插入原子...", self)
        self.insert_atom_action.triggered.connect(self.insert_atom)

        self.insert_fragment_action = QtGui.QAction("插入 fragment...", self)
        self.insert_fragment_action.triggered.connect(self.insert_fragment)

        self.copy_ids_action = QtGui.QAction("复制选中 ID", self)
        self.copy_ids_action.triggered.connect(self.copy_selected_ids)

        self.select_pattern_atoms_action = QtGui.QAction("选中 pattern 包含的原子", self)
        self.select_pattern_atoms_action.triggered.connect(self.select_pattern_atoms)

        self.select_shell_atom_action = QtGui.QAction("选中对应 shell 原子", self)
        self.select_shell_atom_action.triggered.connect(self.select_shell_atom_for_selected_core)

        self.replace_core_with_shell_action = QtGui.QAction("用该 shell 替换对应 core 的拓扑", self)
        self.replace_core_with_shell_action.triggered.connect(self.replace_core_with_selected_shell)

        self.replace_topology_cores_action = QtGui.QAction("选中拓扑 core->shell", self)
        self.replace_topology_cores_action.triggered.connect(self.replace_selected_topology_cores_with_shells)

        self.delete_selected_action = QtGui.QAction("删除选中对象", self)
        self.delete_selected_action.setShortcut(QtGui.QKeySequence.Delete)
        self.delete_selected_action.setShortcutContext(QtCore.Qt.WindowShortcut)
        self.delete_selected_action.triggered.connect(self.delete_selected_items)

        self.reset_camera_action = QtGui.QAction("重置视角", self)
        self.reset_camera_action.triggered.connect(self.reset_camera)

        self.import_forcefield_action = QtGui.QAction("导入力场文件...", self)
        self.import_forcefield_action.triggered.connect(self.import_forcefield_dialog)

        self.set_forcefield_library_action = QtGui.QAction("设置力场库文件夹...", self)
        self.set_forcefield_library_action.triggered.connect(self.choose_forcefield_library_folder)

        self.reload_forcefield_library_action = QtGui.QAction("重新扫描力场库", self)
        self.reload_forcefield_library_action.triggered.connect(self.scan_forcefield_library)

        self.show_forcefield_candidates_action = QtGui.QAction("候选参数表...", self)
        self.show_forcefield_candidates_action.triggered.connect(self.show_forcefield_candidates)

        self.import_ff_mapping_action = QtGui.QAction("导入 ff-type 映射...", self)
        self.import_ff_mapping_action.triggered.connect(self.import_ff_mapping_dialog)

        self.apply_forcefield_action = QtGui.QAction("赋予力场参数...", self)
        self.apply_forcefield_action.triggered.connect(self.apply_forcefield_dialog)

    def _make_menus(self) -> None:
        menu = self.menuBar()
        file_menu = menu.addMenu("文件")
        file_menu.addAction(self.open_action)
        file_menu.addAction(self.open_new_window_action)
        file_menu.addAction(self.import_ms_action)
        file_menu.addAction(self.configure_ms_action)
        file_menu.addAction(self.find_action)
        file_menu.addAction(self.save_as_action)
        file_menu.addSeparator()
        file_menu.addAction(self.tile_windows_action)
        file_menu.addAction(self.sync_group_action)
        file_menu.addSeparator()
        file_menu.addAction(self.exit_action)

        view_menu = menu.addMenu("显示")
        view_menu.addAction(self.hide_selected_action)
        view_menu.addAction(self.isolate_selected_action)
        view_menu.addAction(self.show_all_action)
        view_menu.addSeparator()
        view_menu.addAction(self.toggle_box_action)
        view_menu.addAction(self.toggle_cross_boundary_bonds_action)
        view_menu.addSeparator()
        view_menu.addAction(self.display_in_cell_action)
        view_menu.addAction(self.display_default_action)
        view_menu.addSeparator()
        view_menu.addAction(self.reset_camera_action)

        select_menu = menu.addMenu("选择")
        select_menu.addAction(self.find_action)
        select_menu.addAction(self.select_fragment_action)
        select_menu.addAction(self.select_pattern_atoms_action)
        select_menu.addAction(self.clear_selection_action)
        select_menu.addAction(self.copy_ids_action)
        select_menu.addAction(self.change_type_action)
        select_menu.addAction(self.replace_topology_cores_action)
        select_menu.addAction(self.add_shell_action)
        select_menu.addAction(self.insert_atom_action)
        select_menu.addAction(self.insert_fragment_action)
        select_menu.addAction(self.delete_selected_action)

        forcefield_menu = menu.addMenu("力场")
        forcefield_menu.addAction(self.import_forcefield_action)
        forcefield_menu.addAction(self.set_forcefield_library_action)
        forcefield_menu.addAction(self.reload_forcefield_library_action)
        forcefield_menu.addSeparator()
        forcefield_menu.addAction(self.show_forcefield_candidates_action)
        forcefield_menu.addAction(self.import_ff_mapping_action)
        forcefield_menu.addAction(self.apply_forcefield_action)
        forcefield_menu.addSeparator()
        self.saved_forcefield_menu = forcefield_menu.addMenu("已保存力场")
        self._refresh_saved_forcefield_menu()

    def _make_toolbar(self) -> None:
        toolbar = self.addToolBar("主工具")
        toolbar.setMovable(False)
        for action in [
            self.open_action,
            self.open_new_window_action,
            self.import_ms_action,
            self.find_action,
            self.save_as_action,
            self.import_forcefield_action,
            self.apply_forcefield_action,
            self.hide_selected_action,
            self.isolate_selected_action,
            self.show_all_action,
            self.toggle_box_action,
            self.select_fragment_action,
            self.change_type_action,
            self.replace_topology_cores_action,
            self.add_shell_action,
            self.insert_atom_action,
            self.insert_fragment_action,
            self.delete_selected_action,
            self.reset_camera_action,
        ]:
            toolbar.addAction(action)

    def _make_side_panel(self) -> None:
        dock = QtWidgets.QDockWidget("信息", self)
        dock.setAllowedAreas(QtCore.Qt.LeftDockWidgetArea | QtCore.Qt.RightDockWidgetArea)
        panel = QtWidgets.QWidget(dock)
        layout = QtWidgets.QVBoxLayout(panel)

        self.summary_label = QtWidgets.QLabel("未加载文件")
        self.summary_label.setWordWrap(True)
        self.summary_label.setAlignment(QtCore.Qt.AlignTop)

        self.selection_text = QtWidgets.QPlainTextEdit()
        self.selection_text.setReadOnly(True)
        self.selection_text.setPlaceholderText("左键单击选择原子；左键拖动框选原子；Shift+左键追加选择；Ctrl+左键切换选择；Alt+双击左键选择当前显示的同元素原子；右键单击打开菜单；右键拖动旋转；中键拖动平移。")

        self.charge_label = QtWidgets.QLabel("No model loaded.")
        self.charge_label.setWordWrap(True)
        self.charge_label.setAlignment(QtCore.Qt.AlignTop)

        self.properties_hint_label = QtWidgets.QLabel("请选择单个原子或同 type 原子以查看属性。")
        self.properties_hint_label.setWordWrap(True)
        self.properties_hint_label.setAlignment(QtCore.Qt.AlignTop)

        self.properties_table = QtWidgets.QTableWidget(6, 2)
        self.properties_table.setHorizontalHeaderLabels(["Property", "Value"])
        self.properties_table.horizontalHeader().setStretchLastSection(True)
        self.properties_table.verticalHeader().setVisible(False)
        self.properties_table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.properties_table.setEditTriggers(
            QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed
        )
        self.properties_table.setFixedHeight(210)
        self.properties_table.itemChanged.connect(self._handle_property_item_changed)
        self.properties_table.cellDoubleClicked.connect(self._handle_property_cell_double_clicked)

        self.type_tabs = QtWidgets.QTabWidget()
        self.atom_type_table = self._make_type_stats_table(["type", "数量", "质量", "Name", "ff-type"])
        self.bond_type_table = self._make_type_stats_table(["type", "数量", "Name", "ff-type"])
        self.angle_type_table = self._make_type_stats_table(["type", "数量", "Name", "ff-type"])
        self.dihedral_type_table = self._make_type_stats_table(["type", "数量", "Name", "ff-type"])
        self.improper_type_table = self._make_type_stats_table(["type", "数量", "Name", "ff-type"])
        self.atom_type_table.setEditTriggers(
            QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed
        )
        self.atom_type_table.itemChanged.connect(self._handle_atom_type_table_item_changed)
        for table in (self.bond_type_table, self.angle_type_table, self.dihedral_type_table, self.improper_type_table):
            table.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed)
            table.itemChanged.connect(self._handle_topology_type_table_item_changed)
        for table in (
            self.atom_type_table,
            self.bond_type_table,
            self.angle_type_table,
            self.dihedral_type_table,
            self.improper_type_table,
        ):
            table.itemDoubleClicked.connect(self._handle_type_table_double_clicked)
        self.type_tabs.addTab(self.atom_type_table, "atom")
        self.type_tabs.addTab(self.bond_type_table, "bond")
        self.type_tabs.addTab(self.angle_type_table, "angle")
        self.type_tabs.addTab(self.dihedral_type_table, "dihedral")
        self.type_tabs.addTab(self.improper_type_table, "improper")

        layout.addWidget(self.summary_label)
        layout.addWidget(QtWidgets.QLabel("当前选择"))
        layout.addWidget(self.selection_text, stretch=1)
        layout.addWidget(QtWidgets.QLabel("Charge"))
        layout.addWidget(self.charge_label)
        layout.addWidget(QtWidgets.QLabel("Atom Properties"))
        layout.addWidget(self.properties_hint_label)
        layout.addWidget(self.properties_table)
        layout.addWidget(QtWidgets.QLabel("type 统计"))
        layout.addWidget(self.type_tabs, stretch=1)
        panel.setLayout(layout)
        dock.setWidget(panel)
        self.addDockWidget(QtCore.Qt.RightDockWidgetArea, dock)

        self.update_properties_panel()
        self.update_charge_panel()

    def _make_type_stats_table(self, headers: Sequence[str]) -> QtWidgets.QTableWidget:
        table = QtWidgets.QTableWidget(0, len(headers))
        table.setHorizontalHeaderLabels(list(headers))
        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        return table

    def _populate_type_stats_table(
        self,
        table: QtWidgets.QTableWidget,
        items: Sequence[Tuple[int, int]],
        masses: Optional[Dict[int, float]] = None,
        labels: Optional[Dict[int, str]] = None,
        ff_labels: Optional[Dict[int, str]] = None,
    ) -> None:
        self._updating_type_tables = True
        try:
            table.setRowCount(len(items))
            for row, (type_id, count) in enumerate(items):
                type_item = QtWidgets.QTableWidgetItem(str(type_id))
                type_item.setFlags(type_item.flags() & ~QtCore.Qt.ItemIsEditable)
                table.setItem(row, 0, type_item)

                count_item = QtWidgets.QTableWidgetItem(str(count))
                count_item.setFlags(count_item.flags() & ~QtCore.Qt.ItemIsEditable)
                table.setItem(row, 1, count_item)

                if masses is not None and table.columnCount() >= 3:
                    mass = None if masses is None else masses.get(type_id)
                    mass_item = QtWidgets.QTableWidgetItem("" if mass is None else f"{mass:g}")
                    mass_item.setFlags(mass_item.flags() & ~QtCore.Qt.ItemIsEditable)
                    table.setItem(row, 2, mass_item)

                if labels is not None and masses is None and table.columnCount() >= 3:
                    label_text = labels.get(type_id, "").strip()
                    label_item = QtWidgets.QTableWidgetItem(label_text or "未定义")
                    table.setItem(row, 2, label_item)

                if table.columnCount() >= 4:
                    label_text = ""
                    if masses is None and ff_labels is not None:
                        label_text = ff_labels.get(type_id, "").strip()
                    elif labels is not None:
                        label_text = labels.get(type_id, "").strip()
                    label_item = QtWidgets.QTableWidgetItem(label_text or "未定义")
                    table.setItem(row, 3, label_item)
                if table.columnCount() >= 5:
                    ff_text = ""
                    if ff_labels is not None:
                        ff_text = ff_labels.get(type_id, "").strip()
                    ff_item = QtWidgets.QTableWidgetItem(ff_text or "未定义")
                    table.setItem(row, 4, ff_item)
            table.resizeColumnsToContents()
        finally:
            self._updating_type_tables = False

    def _handle_type_table_double_clicked(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self.data_model is None:
            return
        table = self.sender()
        try:
            type_item = table.item(item.row(), 0) if hasattr(table, "item") else None
            type_id = int(type_item.text() if type_item is not None else item.text())
        except ValueError:
            return

        kind = self._type_table_kind(table)
        if kind is None:
            return
        ff_column = 4 if kind == "atom" else 3
        if item.column() == ff_column:
            self._choose_candidate_for_type(kind, type_id)
            return
        if item.column() != 0:
            return

        modifiers = QtWidgets.QApplication.keyboardModifiers()
        additive = bool(modifiers & QtCore.Qt.ControlModifier)
        ranged = bool(modifiers & QtCore.Qt.ShiftModifier) and kind in self._type_table_anchor
        if ranged:
            anchor_type = self._type_table_anchor[kind]
            low, high = sorted((anchor_type, type_id))
            type_ids = set(range(low, high + 1))
        else:
            type_ids = {type_id}
            self._type_table_anchor[kind] = type_id

        if kind == "atom":
            atom_ids = {aid for aid, atom in self.data_model.atoms.items() if atom.atom_type in type_ids}
            if additive and self.data_model.selection_kind == "atom":
                atom_ids |= set(self.data_model.selected_atoms)
            hidden_hit = bool(atom_ids & self.data_model.hidden_atoms)
            self.data_model.reveal_atoms(atom_ids)
            self.data_model.select_atoms(atom_ids)
            self.last_picked_atom = min(atom_ids) if len(atom_ids) == 1 else None
            self.last_picked_bond = None
            if hidden_hit:
                self.refresh_scene()
            else:
                self._update_selection_visuals()
                self.update_summary_panel()
            self.update_selection_panel()
            type_text = ",".join(str(value) for value in sorted(type_ids))
            self.statusBar().showMessage(f"已选择全部 atom type={type_text}，共 {len(atom_ids)} 个。")
            return

        record_ids = {
            record.record_id
            for record in self.data_model.topology_records(kind)
            if record.record_type in type_ids
        }
        if additive and self.data_model.selection_kind == kind:
            record_ids |= set(self.data_model.topology_id_set(kind))
        atom_ids: Set[int] = set()
        for record in self.data_model.topology_records(kind):
            if record.record_id in record_ids:
                atom_ids.update(record.atom_ids)
        hidden_hit = bool(atom_ids & self.data_model.hidden_atoms)
        self.data_model.reveal_atoms(atom_ids)
        self.data_model.select_patterns(kind, record_ids)
        self.last_picked_atom = None
        self.last_picked_bond = min(record_ids) if kind == "bond" and len(record_ids) == 1 else None
        if hidden_hit:
            self.refresh_scene()
        else:
            self._update_selection_visuals()
            self.update_summary_panel()
        self.update_selection_panel()
        type_text = ",".join(str(value) for value in sorted(type_ids))
        self.statusBar().showMessage(f"已选择全部 {kind} type={type_text}，共 {len(record_ids)} 个。")

    def _connect_events(self) -> None:
        if self.plotter is None:
            return
        self.plotter.interactor.installEventFilter(self)

    def _clear_left_drag_state(self) -> None:
        self._left_press_pos = None
        self._left_dragged = False
        if self._selection_band is not None:
            self._selection_band.hide()

    def eventFilter(self, obj, event):  # noqa: N802
        if self.plotter is not None and obj is self.plotter.interactor:
            if event.type() == QtCore.QEvent.DragEnter:
                return self._handle_drag_enter(event)
            if event.type() == QtCore.QEvent.Drop:
                return self._handle_drop(event)
            if event.type() == QtCore.QEvent.MouseButtonPress:
                return self._handle_mouse_press(event)
            if event.type() == QtCore.QEvent.MouseMove:
                return self._handle_mouse_move(event)
            if event.type() == QtCore.QEvent.MouseButtonRelease:
                return self._handle_mouse_release(event)
            if event.type() == QtCore.QEvent.MouseButtonDblClick:
                return self._handle_mouse_double_click(event)
        return super().eventFilter(obj, event)

    def dragEnterEvent(self, event: QtGui.QDragEnterEvent) -> None:  # noqa: N802
        if not self._handle_drag_enter(event):
            super().dragEnterEvent(event)

    def dropEvent(self, event: QtGui.QDropEvent) -> None:  # noqa: N802
        if not self._handle_drop(event):
            super().dropEvent(event)

    def _first_local_drop_path(self, event) -> Optional[Path]:
        paths = self._local_drop_paths(event)
        return paths[0] if paths else None

    def _local_drop_paths(self, event) -> List[Path]:
        mime = event.mimeData()
        if not mime.hasUrls():
            return []
        paths: List[Path] = []
        for url in mime.urls():
            if url.isLocalFile():
                paths.append(Path(url.toLocalFile()))
        return paths

    def _handle_drag_enter(self, event) -> bool:
        path = self._first_local_drop_path(event)
        if path is not None and path.is_file():
            event.acceptProposedAction()
            return True
        return False

    def _handle_drop(self, event) -> bool:
        paths = [path for path in self._local_drop_paths(event) if path.is_file()]
        if not paths:
            QtWidgets.QMessageBox.warning(self, "无法打开", "拖入的内容不是可读取的本地 data 文件。")
            return True
        event.acceptProposedAction()
        data_paths: List[Path] = []
        ms_input_paths: List[Path] = []
        for path in paths:
            if self._is_forcefield_file(path):
                self.import_forcefield_file(path, ask_to_save=True)
            elif self._is_msi2lmp_input_file(path):
                ms_input_paths.append(path)
            else:
                data_paths.append(path)
        if ms_input_paths:
            data_paths.extend(self.convert_materials_studio_files(ms_input_paths))
        self.open_data_paths(data_paths, prefer_current=True)
        return True

    def _handle_mouse_press(self, event: QtGui.QMouseEvent) -> bool:
        if event.button() == QtCore.Qt.RightButton:
            self._right_press_pos = event.position().toPoint()
            self._right_dragged = False
            return False
        if event.button() == QtCore.Qt.LeftButton:
            self._left_press_pos = event.position().toPoint()
            self._left_dragged = False
            if self._selection_band is not None:
                self._selection_band.hide()
            return True
        return False

    def _handle_mouse_move(self, event: QtGui.QMouseEvent) -> bool:
        if self._right_press_pos is not None:
            delta = event.position().toPoint() - self._right_press_pos
            if delta.manhattanLength() > self._pick_tolerance_px:
                self._right_dragged = True
            return False
        if self._left_press_pos is not None:
            if not (event.buttons() & QtCore.Qt.LeftButton):
                self._clear_left_drag_state()
                return False
            current_pos = event.position().toPoint()
            delta = current_pos - self._left_press_pos
            if delta.manhattanLength() > self._pick_tolerance_px:
                self._left_dragged = True
                selection_rect = QtCore.QRect(self._left_press_pos, current_pos).normalized()
                if self._selection_band is not None:
                    self._selection_band.setGeometry(selection_rect)
                    self._selection_band.show()
            return True
        if event.buttons() == QtCore.Qt.NoButton and event.modifiers() & QtCore.Qt.ControlModifier:
            return True
        return False

    def _handle_mouse_release(self, event: QtGui.QMouseEvent) -> bool:
        if event.button() == QtCore.Qt.LeftButton:
            if self._left_press_pos is None:
                return True
            start_pos = QtCore.QPoint(self._left_press_pos)
            end_pos = event.position().toPoint()
            delta = end_pos - start_pos
            left_dragged = self._left_dragged
            self._clear_left_drag_state()
            if left_dragged and delta.manhattanLength() > self._pick_tolerance_px:
                self._select_atoms_in_rect(start_pos, end_pos, event.modifiers())
            elif delta.manhattanLength() <= self._pick_tolerance_px:
                self._pick_from_click(event)
            return True

        if event.button() == QtCore.Qt.RightButton:
            if self._right_press_pos is None:
                return False
            pos = event.position().toPoint()
            delta = pos - self._right_press_pos
            is_click = (delta.manhattanLength() <= self._pick_tolerance_px) and (not self._right_dragged)
            self._right_press_pos = None
            self._right_dragged = False
            if is_click:
                self._show_context_menu(pos)
            return False
        return False

    def _handle_mouse_double_click(self, event: QtGui.QMouseEvent) -> bool:
        if event.button() == QtCore.Qt.LeftButton and event.modifiers() & QtCore.Qt.AltModifier:
            if self.data_model is None:
                return True
            pos = event.position().toPoint()
            picked_kind, picked_id = self._pick_front_target(pos)
            if picked_kind == "atom" and picked_id is not None:
                atom_id = picked_id
                element_symbol, selected_count = self.data_model.select_visible_same_element(atom_id)
                self.last_picked_atom = atom_id
                self.last_picked_bond = None
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
                atom = self.data_model.atoms[atom_id]
                if element_symbol is None:
                    self.statusBar().showMessage(f"已选择当前显示的 type={atom.atom_type} 原子，共 {selected_count} 个。")
                else:
                    self.statusBar().showMessage(f"已选择当前显示的 {element_symbol} 元素原子，共 {selected_count} 个。")
                return True
            if picked_kind == "bond" and picked_id is not None:
                bond_id = picked_id
                self.data_model.select_same_pattern_type("bond", bond_id)
                self.last_picked_atom = None
                self.last_picked_bond = bond_id
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
                bond = next(bond for bond in self.data_model.bonds if bond.bond_id == bond_id)
                self.statusBar().showMessage(
                    f"已选择全部 bond type={bond.bond_type} 的键，共 {len(self.data_model.selected_bonds)} 个。"
                )
            return True
        return False

    def _pick_from_click(self, event: QtGui.QMouseEvent) -> None:
        if self.data_model is None:
            return
        pos = event.position().toPoint()
        picked_kind, picked_id = self._pick_front_target(pos)
        if picked_kind == "bond" and picked_id is not None and event.modifiers() & QtCore.Qt.AltModifier:
            bond_id = picked_id
            self.data_model.select_same_pattern_type("bond", bond_id)
            self.last_picked_atom = None
            self.last_picked_bond = bond_id
            self._update_selection_visuals()
            self.update_summary_panel()
            self.update_selection_panel()
            bond = next(bond for bond in self.data_model.bonds if bond.bond_id == bond_id)
            self.statusBar().showMessage(
                f"已选择全部 bond type={bond.bond_type} 的键，共 {len(self.data_model.selected_bonds)} 个。"
            )
            return
        if picked_kind == "bond" and picked_id is not None:
            bond_id = picked_id
            additive = bool(event.modifiers() & QtCore.Qt.ShiftModifier)
            toggle = bool(event.modifiers() & QtCore.Qt.ControlModifier)
            self.data_model.select_pattern_one("bond", bond_id, additive=additive, toggle=toggle)
            self.last_picked_atom = None
            self.last_picked_bond = bond_id
            bond = next(bond for bond in self.data_model.bonds if bond.bond_id == bond_id)
            self.statusBar().showMessage(f"Bond {bond_id}: type={bond.bond_type}, atoms=({bond.atom1}, {bond.atom2})")
            self._update_selection_visuals()
            self.update_summary_panel()
            self.update_selection_panel()
            return
        if picked_kind != "atom" or picked_id is None:
            if not (event.modifiers() & QtCore.Qt.ControlModifier):
                changed = self.data_model.has_selection() or self.last_picked_atom is not None or self.last_picked_bond is not None
                self.data_model.clear_selection()
                self.last_picked_atom = None
                self.last_picked_bond = None
                if changed:
                    self._update_selection_visuals()
                    self.update_summary_panel()
                    self.update_selection_panel()
            return

        additive = bool(event.modifiers() & QtCore.Qt.ShiftModifier)
        toggle = bool(event.modifiers() & QtCore.Qt.ControlModifier)
        atom_id = picked_id
        self.data_model.select_one(atom_id, additive=additive, toggle=toggle)
        self.last_picked_atom = atom_id
        self.last_picked_bond = None
        atom = self.data_model.atoms[atom_id]
        frag = self.data_model.fragment_of_atom.get(atom_id, "-")
        msg = f"Atom {atom_id}: type={atom.atom_type}"
        if atom.mol is not None:
            msg += f", mol={atom.mol}"
        msg += f", fragment={frag}, xyz=({atom.x:.4f}, {atom.y:.4f}, {atom.z:.4f})"
        self.statusBar().showMessage(msg)
        self._update_selection_visuals()
        self.update_summary_panel()
        self.update_selection_panel()

    def _select_atoms_in_rect(
        self,
        start_pos: QtCore.QPoint,
        end_pos: QtCore.QPoint,
        modifiers: QtCore.Qt.KeyboardModifiers,
    ) -> None:
        if self.data_model is None or self.points_mesh is None:
            return

        selection_rect = QtCore.QRect(start_pos, end_pos).normalized()
        selected_ids: Set[int] = set()
        for atom_id in map(int, np.asarray(self.points_mesh["orig_id"])):
            screen_pos = self._project_atom_to_widget(atom_id)
            if screen_pos is not None and selection_rect.contains(screen_pos):
                selected_ids.add(atom_id)

        additive = bool(modifiers & QtCore.Qt.ShiftModifier)
        toggle = bool(modifiers & QtCore.Qt.ControlModifier)
        current_atoms = set(self.data_model.selected_atoms) if self.data_model.selection_kind == "atom" else set()

        if toggle:
            new_selection = current_atoms.symmetric_difference(selected_ids)
        elif additive:
            new_selection = current_atoms | selected_ids
        else:
            new_selection = selected_ids

        if not new_selection and not (additive or toggle):
            changed = self.data_model.has_selection() or self.last_picked_atom is not None
            self.data_model.clear_selection()
            self.last_picked_atom = None
            if changed:
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
            self.statusBar().showMessage("框选范围内未找到原子。")
            return

        self.data_model.select_atoms(new_selection)
        self.last_picked_atom = min(new_selection) if len(new_selection) == 1 else None
        self._update_selection_visuals()
        self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(f"框选到 {len(selected_ids)} 个原子，当前选中 {len(new_selection)} 个。")

    def _show_context_menu(self, pos: QtCore.QPoint) -> None:
        if self.data_model is None:
            return
        picked_kind, picked_id = self._pick_front_target(pos)
        if self.data_model.selection_kind in (None, "atom"):
            if picked_kind == "atom" and picked_id is not None and picked_id not in self.data_model.selected_atoms:
                self.data_model.select_one(picked_id)
                self.last_picked_atom = picked_id
                self.last_picked_bond = None
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
            elif picked_kind == "bond" and picked_id is not None and picked_id not in self.data_model.selected_bonds:
                self.data_model.select_pattern_one("bond", picked_id)
                self.last_picked_atom = None
                self.last_picked_bond = picked_id
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
        elif self.data_model.selection_kind == "bond":
            if picked_kind == "bond" and picked_id is not None and picked_id not in self.data_model.selected_bonds:
                self.data_model.select_pattern_one("bond", picked_id)
                self.last_picked_atom = None
                self.last_picked_bond = picked_id
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()

        if not self.data_model.has_selection():
            return

        menu = QtWidgets.QMenu(self)
        menu.addAction(self.hide_selected_action)
        menu.addAction(self.isolate_selected_action)
        menu.addAction(self.show_all_action)
        menu.addSeparator()
        if self.data_model.selection_kind == "atom":
            menu.addAction(self.select_fragment_action)
            menu.addAction(self.add_shell_action)
            selected_atoms = sorted(self.data_model.selected_atoms)
            if len(selected_atoms) == 1:
                atom_id = selected_atoms[0]
                if self.data_model.shell_for_core(atom_id) is not None:
                    menu.addAction(self.select_shell_atom_action)
                if self.data_model.core_for_shell(atom_id) is not None:
                    menu.addAction(self.replace_core_with_shell_action)
        elif self.data_model.selection_kind in TOPOLOGY_KIND_TO_SECTION:
            menu.addAction(self.select_pattern_atoms_action)
            menu.addAction(self.replace_topology_cores_action)
        menu.addAction(self.copy_ids_action)
        if self.data_model.selection_kind == "atom" or self.data_model.selection_kind in TOPOLOGY_KIND_TO_SECTION:
            menu.addAction(self.change_type_action)
        menu.addAction(self.delete_selected_action)
        menu.addSeparator()
        menu.addAction(self.clear_selection_action)
        menu.exec(self.plotter.interactor.mapToGlobal(pos))

    def open_file_dialog(self) -> None:
        file_paths, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "打开 LAMMPS data 文件",
            "",
            "LAMMPS data (*.data *.dat *.lmp *.txt);;All files (*.*)",
        )
        self.open_data_paths([Path(file_path) for file_path in file_paths], prefer_current=True)

    def open_file_new_window_dialog(self) -> None:
        file_paths, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self,
            "在新窗口打开 LAMMPS data 文件",
            "",
            "LAMMPS data (*.data *.dat *.lmp *.txt);;All files (*.*)",
        )
        self.open_data_paths([Path(file_path) for file_path in file_paths], prefer_current=False)

    def open_data_paths(self, paths: Sequence[Path], prefer_current: bool = True) -> None:
        data_paths = [Path(path) for path in paths if path]
        if not data_paths:
            return
        if prefer_current and self.data_model is None and self.current_file is None:
            self.load_data_file(data_paths[0])
            remaining = data_paths[1:]
        else:
            remaining = data_paths
        for path in remaining:
            self.open_new_viewer_window(path)

    def open_new_viewer_window(self, path: Optional[Path] = None) -> "ViewerMainWindow":
        window = ViewerMainWindow(path)
        window.resize(self.size())
        window.show()
        return window

    def tile_open_windows(self) -> None:
        windows = open_viewer_windows()
        if not windows:
            return
        screen = QtWidgets.QApplication.primaryScreen()
        if screen is None:
            return
        area = screen.availableGeometry()
        columns = max(1, math.ceil(math.sqrt(len(windows))))
        rows = max(1, math.ceil(len(windows) / columns))
        cell_width = max(420, area.width() // columns)
        cell_height = max(320, area.height() // rows)
        for index, window in enumerate(windows):
            row = index // columns
            column = index % columns
            window.showNormal()
            window.setGeometry(
                area.x() + column * cell_width,
                area.y() + row * cell_height,
                cell_width,
                cell_height,
            )

    def _sync_type_signature(self) -> Dict[str, Dict[int, Tuple[str, str]]]:
        if self.data_model is None:
            return {}
        atom_sig: Dict[int, Tuple[str, str]] = {}
        for atom_type in sorted(self.data_model._current_type_ids("atom")):
            element_info = infer_element_from_mass(self.data_model.masses.get(atom_type))
            element = "" if element_info is None else element_info[0]
            atom_sig[atom_type] = (element, self.data_model.mass_names.get(atom_type, "").strip())
        signature: Dict[str, Dict[int, Tuple[str, str]]] = {"atom": atom_sig}
        for kind in ("bond", "angle", "dihedral", "improper"):
            signature[kind] = {
                type_id: ("", self.data_model.coeff_names.get(kind, {}).get(type_id, "").strip())
                for type_id in sorted(self.data_model._current_type_ids(kind))
            }
        return signature

    def _sync_consistency_issues(self, windows: Sequence["ViewerMainWindow"]) -> List[str]:
        loaded = [window for window in windows if window.data_model is not None]
        if len(loaded) < 2:
            return ["同步组至少需要两个已加载 data 的窗口。"]
        base = loaded[0]._sync_type_signature()
        issues: List[str] = []
        for window in loaded[1:]:
            other = window._sync_type_signature()
            for kind in ("atom", "bond", "angle", "dihedral", "improper"):
                base_types = set(base.get(kind, {}))
                other_types = set(other.get(kind, {}))
                if base_types != other_types:
                    issues.append(f"{window.current_file or window.windowTitle()}: {kind} type 编号不一致")
                    continue
                for type_id in sorted(base_types):
                    if base[kind][type_id] != other[kind][type_id]:
                        issues.append(f"{window.current_file or window.windowTitle()}: {kind} type={type_id} 的元素/Name 不一致")
        return issues

    def open_sync_group_dialog(self) -> None:
        global SYNC_GROUP_WINDOWS
        windows = [window for window in open_viewer_windows() if window.data_model is not None]
        if not windows:
            QtWidgets.QMessageBox.information(self, "同步组", "当前没有已加载 data 的窗口。")
            return
        dialog = SyncGroupDialog(windows, SYNC_GROUP_WINDOWS, self)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        selected = dialog.selected_windows()
        if not selected:
            SYNC_GROUP_WINDOWS = []
            self.statusBar().showMessage("同步组已清空。")
            return
        issues = self._sync_consistency_issues(selected)
        if issues:
            QtWidgets.QMessageBox.warning(
                self,
                "同步组检查未通过",
                "以下差异会导致同步操作不安全，已取消建立同步组：\n\n" + "\n".join(issues[:40]),
            )
            return
        SYNC_GROUP_WINDOWS = list(selected)
        for window in SYNC_GROUP_WINDOWS:
            window.statusBar().showMessage(f"同步组已启用，共 {len(SYNC_GROUP_WINDOWS)} 个模型。")

    def _sync_peers(self) -> List["ViewerMainWindow"]:
        if self not in SYNC_GROUP_WINDOWS:
            return []
        return [window for window in SYNC_GROUP_WINDOWS if window is not self and window.data_model is not None]

    def _sync_atom_type_property(self, atom_type: int, field_name: str, value: object) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None or int(atom_type) not in window.data_model._current_type_ids("atom"):
                continue
            if field_name == "name":
                window.data_model.set_name_for_type(atom_type, str(value))
            elif field_name == "ff":
                window.data_model.set_ff_type_for_type(atom_type, str(value))
            elif field_name == "mass":
                window.data_model.set_mass_for_type(atom_type, float(value))
                window.refresh_scene()
            elif field_name == "color":
                window.data_model.set_color_for_type(atom_type, value)  # type: ignore[arg-type]
                window.refresh_scene()
            else:
                continue
            window.update_summary_panel()
            window.update_properties_panel()
            window.update_selection_panel()
            changed += 1
        return changed

    def _sync_topology_type_property(self, kind: str, type_id: int, field_name: str, value: str) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None or int(type_id) not in window.data_model._current_type_ids(kind):
                continue
            if field_name == "name":
                window.data_model.set_coeff_name_for_type(kind, type_id, value)
            elif field_name == "ff":
                window.data_model.set_coeff_label_for_type(kind, type_id, value)
            else:
                continue
            window.update_summary_panel()
            window.update_selection_panel()
            changed += 1
        return changed

    def _sync_apply_forcefield(self, definition: ForceFieldDefinition) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None:
                continue
            window.data_model.apply_forcefield(definition)
            window.update_summary_panel()
            window.update_selection_panel()
            changed += 1
        return changed

    def _select_types_in_model(self, model: DataModel, kind: str, type_ids: Set[int]) -> bool:
        if kind == "atom":
            atom_ids = {aid for aid, atom in model.atoms.items() if atom.atom_type in type_ids}
            if not atom_ids:
                return False
            model.select_atoms(atom_ids)
            return True
        if kind in TOPOLOGY_KIND_TO_SECTION:
            record_ids = {record.record_id for record in model.topology_records(kind) if record.record_type in type_ids}
            if not record_ids:
                return False
            model.select_patterns(kind, record_ids)
            return True
        return False

    def _current_selection_covers_full_types(self, kind: str, type_ids: Set[int]) -> bool:
        if self.data_model is None or not type_ids:
            return False
        if kind == "atom":
            expected = {aid for aid, atom in self.data_model.atoms.items() if atom.atom_type in type_ids}
            return expected == set(self.data_model.selected_atoms)
        if kind in TOPOLOGY_KIND_TO_SECTION:
            expected = {record.record_id for record in self.data_model.topology_records(kind) if record.record_type in type_ids}
            return expected == set(self.data_model.topology_id_set(kind))
        return False

    def _sync_change_selected_type_by_types(
        self,
        kind: str,
        source_types: Set[int],
        new_type: int,
        insert_existing_type: bool,
    ) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None:
                continue
            if not self._select_types_in_model(window.data_model, kind, source_types):
                continue
            window.data_model.change_selected_type(new_type, insert_existing_type=insert_existing_type)
            window.refresh_scene()
            window.update_summary_panel()
            window.update_selection_panel()
            changed += 1
        return changed

    def _sync_delete_selected_by_types(self, kind: str, source_types: Set[int]) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None:
                continue
            if not self._select_types_in_model(window.data_model, kind, source_types):
                continue
            counts = window.data_model.delete_current_selection()
            if counts.get("atom"):
                window.refresh_scene_actor_swap()
            elif counts.get("bond"):
                window.refresh_bond_actor()
                window._update_selection_visuals(render=False)
                if window.plotter is not None:
                    window.plotter.render()
            else:
                window._update_selection_visuals()
            window.update_summary_panel()
            window.update_selection_panel()
            changed += 1
        return changed

    def _sync_add_shell_by_type(self, atom_type: int) -> int:
        changed = 0
        for window in self._sync_peers():
            if window.data_model is None or atom_type not in window.data_model._current_type_ids("atom"):
                continue
            atom_ids = window.data_model.atom_ids_of_type(atom_type)
            if not atom_ids:
                continue
            result = window.data_model.add_shell_atoms(atom_type, atom_ids, split_selected_atoms=False, z_offset=0.05)
            shell_ids = set(result["shell_atom_ids"])
            window.data_model.reveal_atoms(shell_ids)
            window.data_model.select_atoms(shell_ids)
            window.last_picked_atom = min(shell_ids) if len(shell_ids) == 1 else None
            window.last_picked_bond = None
            window.refresh_scene()
            window.update_selection_panel()
            changed += 1
        return changed

    def save_as_dialog(self) -> bool:
        if self.data_model is None:
            return False
        self._clear_left_drag_state()
        total_charge = self.data_model.total_charge()
        if not math.isclose(total_charge, 0.0, abs_tol=1e-6):
            QtWidgets.QMessageBox.warning(self, "电中性提醒", "模型未保持电中性")
        default_name = self.current_file.stem + "_edited.data" if self.current_file else "edited.data"
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "另存为",
            default_name,
            "LAMMPS data (*.data *.dat *.lmp *.txt);;All files (*.*)",
        )
        if not file_path:
            return False
        try:
            self.data_model.save_as(Path(file_path))
            self.data_model.dirty = False
            self.statusBar().showMessage(f"已保存到: {file_path}")
            return True
        except Exception as exc:  # pragma: no cover - GUI path
            QtWidgets.QMessageBox.critical(self, "保存失败", str(exc))
            return False

    def open_find_dialog(self) -> None:
        self._clear_left_drag_state()
        if self.find_dialog is None:
            self.find_dialog = FindDialog(self)
        self.find_dialog._refresh_fields()
        self.find_dialog.focus_input()

    def apply_find_request(self, category: str, field_name: str, raw_value: str) -> str:
        if self.data_model is None:
            return "请先打开一个 data 文件。"

        category = category.strip().lower()
        field_name = field_name.strip().lower()
        raw_value = raw_value.strip()
        if category != "atom" and not self.data_model.topology_exists(category):
            self.data_model.clear_selection()
            self.last_picked_atom = None
            self.last_picked_bond = None
            self._update_selection_visuals()
            self.update_summary_panel()
            self.update_selection_panel()
            return "当前data不含此信息"
        if not raw_value:
            return "输入参数后自动搜索。"

        if category == "bond" and field_name == "atom_types":
            tokens = [token for token in re.split(r"[\s,;/]+", raw_value) if token]
            if len(tokens) != 2:
                return "请输入两个原子 type，例如: 1 2"
            try:
                value = (int(tokens[0]), int(tokens[1]))
            except ValueError:
                return "请输入两个整数 type，例如: 1 2"
        else:
            try:
                if field_name == "type":
                    value = parse_integer_selector(raw_value)
                else:
                    value = float(raw_value) if category == "atom" and field_name == "mass" else int(raw_value)
            except ValueError:
                if field_name == "type":
                    return "请输入有效的 type 表达式，例如 1,3~5。"
                return "请输入有效的数字。"

        try:
            if category == "atom":
                result_ids = self.data_model.find_atoms(field_name, value)
                if not result_ids:
                    self.data_model.clear_selection()
                    self.last_picked_atom = None
                    self.last_picked_bond = None
                    self._update_selection_visuals()
                    self.update_summary_panel()
                    self.update_selection_panel()
                    return "未找到符合要求的原子"
                hidden_hit = bool(result_ids & self.data_model.hidden_atoms)
                self.data_model.reveal_atoms(result_ids)
                self.data_model.select_atoms(result_ids)
                self.last_picked_atom = min(result_ids) if len(result_ids) == 1 else None
                self.last_picked_bond = None
                if hidden_hit:
                    self.refresh_scene()
                else:
                    self._update_selection_visuals()
                    self.update_summary_panel()
                self.update_selection_panel()
                return f"已选中 {len(result_ids)} 个 atom。"

            result_ids = self.data_model.find_topology(category, field_name, value)
            if not result_ids:
                self.data_model.clear_selection()
                self.last_picked_atom = None
                self.last_picked_bond = None
                self._update_selection_visuals()
                self.update_summary_panel()
                self.update_selection_panel()
                return f"未找到符合要求的{category}"

            atom_ids: Set[int] = set()
            for record in self.data_model.topology_records(category):
                if record.record_id in result_ids:
                    atom_ids.update(record.atom_ids)
            hidden_hit = bool(atom_ids & self.data_model.hidden_atoms)
            self.data_model.reveal_atoms(atom_ids)
            self.data_model.select_patterns(category, result_ids)
            self.last_picked_atom = None
            self.last_picked_bond = min(result_ids) if category == "bond" and len(result_ids) == 1 else None
            if hidden_hit:
                self.refresh_scene()
            else:
                self._update_selection_visuals()
                self.update_summary_panel()
            self.update_selection_panel()
            return f"已选中 {len(result_ids)} 个{category}。"
        except ValueError:
            return "搜索条件无效。"

    def select_pattern_atoms(self) -> None:
        if self.data_model is None or self.data_model.selection_kind not in TOPOLOGY_KIND_TO_SECTION:
            return
        atom_ids = self.data_model.current_pattern_atom_ids()
        if not atom_ids:
            self.statusBar().showMessage("当前 pattern 不包含可选原子。")
            return
        hidden_hit = bool(atom_ids & self.data_model.hidden_atoms)
        self.data_model.reveal_atoms(atom_ids)
        self.data_model.select_atoms(atom_ids)
        self.last_picked_atom = min(atom_ids) if len(atom_ids) == 1 else None
        self.last_picked_bond = None
        if hidden_hit:
            self.refresh_scene()
        else:
            self._update_selection_visuals()
            self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(f"已选中 pattern 包含的 {len(atom_ids)} 个原子。")

    def select_shell_atom_for_selected_core(self) -> None:
        if self.data_model is None or self.data_model.selection_kind != "atom" or len(self.data_model.selected_atoms) != 1:
            return
        core_id = next(iter(self.data_model.selected_atoms))
        shell_id = self.data_model.shell_for_core(core_id)
        if shell_id is None:
            self.statusBar().showMessage("当前原子没有明确的 shell 原子。")
            return
        self.data_model.reveal_atoms({shell_id})
        self.data_model.select_atoms({shell_id})
        self.last_picked_atom = shell_id
        self.last_picked_bond = None
        self.refresh_scene()
        self.update_selection_panel()
        self.statusBar().showMessage(f"已选中 core atom {core_id} 对应的 shell atom {shell_id}。")

    def replace_core_with_selected_shell(self) -> None:
        if self.data_model is None or self.data_model.selection_kind != "atom" or len(self.data_model.selected_atoms) != 1:
            return
        shell_id = next(iter(self.data_model.selected_atoms))
        core_id = self.data_model.core_for_shell(shell_id)
        if core_id is None:
            self.statusBar().showMessage("当前原子不是明确的 shell 原子。")
            return
        result = self.data_model.replace_core_with_shell_everywhere(shell_id)
        self.refresh_scene()
        self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(
            f"已用 shell atom {shell_id} 替换 core atom {core_id}: {result['records']} 条拓扑记录，{result['atoms']} 处替换，跳过 {result['skipped']} 条。"
        )

    def replace_selected_topology_cores_with_shells(self) -> None:
        if self.data_model is None or self.data_model.selection_kind not in TOPOLOGY_KIND_TO_SECTION:
            self.statusBar().showMessage("请先选中 bond/angle/dihedral/improper。")
            return
        kind = self.data_model.selection_kind
        result = self.data_model.replace_selected_topology_cores_with_shells()
        self.refresh_scene()
        self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(
            f"已处理选中 {kind}: {result['records']} 条记录，{result['atoms']} 处 core->shell，跳过 {result['skipped']} 条。"
        )

    def _format_deletion_summary(self, counts: Dict[str, int]) -> str:
        order = ["atom", "bond", "angle", "dihedral", "improper"]
        parts = [f"{kind}:{counts[kind]}" for kind in order if counts.get(kind)]
        return "已删除 " + ", ".join(parts) if parts else "没有删除任何对象。"

    def delete_selected_items(self) -> None:
        if self.data_model is None or not self.data_model.has_selection():
            return
        selection_kind = self.data_model.selection_kind
        source_types: Set[int] = set()
        if selection_kind == "atom":
            source_types = {self.data_model.atoms[aid].atom_type for aid in self.data_model.selected_atoms if aid in self.data_model.atoms}
        elif selection_kind in TOPOLOGY_KIND_TO_SECTION:
            selected_ids = set(self.data_model.topology_id_set(selection_kind))
            source_types = {
                record.record_type
                for record in self.data_model.topology_records(selection_kind)
                if record.record_id in selected_ids
            }
        counts = self.data_model.delete_current_selection()
        can_sync_by_type = self._current_selection_covers_full_types(selection_kind or "", source_types)
        synced = self._sync_delete_selected_by_types(selection_kind or "", source_types) if can_sync_by_type else 0
        self.last_picked_atom = None
        self.last_picked_bond = None
        if counts.get("atom"):
            self.refresh_scene_actor_swap()
        elif counts.get("bond"):
            self.refresh_bond_actor()
            self._update_selection_visuals(render=False)
            self.plotter.render()
        else:
            self._update_selection_visuals()
        self.update_summary_panel()
        self.update_selection_panel()
        suffix = f"；同步删除 {synced} 个模型" if can_sync_by_type else "；当前选择不是完整 type，未同步删除"
        self.statusBar().showMessage(f"{self._format_deletion_summary(counts)}{suffix}")

    def load_data_file(self, path: Path) -> None:
        if not self._plotter_ready:
            self._pending_initial_file = Path(path)
            self.statusBar().showMessage("3D 引擎尚未初始化，文件将在初始化完成后打开。")
            return
        try:
            parsed = LammpsDataParser(path).parse()
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "读取失败", f"无法解析文件:\n{path}\n\n{exc}")
            return

        self.data_model = DataModel(parsed)
        self.current_file = Path(path)
        self.last_picked_atom = None
        self.last_picked_bond = None
        self.refresh_scene(reset_camera=True)
        self.update_summary_panel()
        self.update_selection_panel()
        if self.find_dialog is not None:
            self.find_dialog._refresh_fields()
        self.setWindowTitle(f"LAMMPS Data Viewer - {self.current_file.name}")
        self.statusBar().showMessage(f"已加载: {path}")

    def set_display_mode(self, mode: str) -> None:
        if mode not in {"in-cell", "default"}:
            return
        if self._display_mode == mode:
            return
        self._display_mode = mode
        self._display_coord_cache = {}
        self.display_in_cell_action.setChecked(mode == "in-cell")
        self.display_default_action.setChecked(mode == "default")
        self.toggle_cross_boundary_bonds_action.setEnabled(mode == "in-cell")
        self.refresh_scene()
        self.update_summary_panel()
        self.statusBar().showMessage(f"Display mode: {mode}")

    def _box_lengths(self) -> Optional[Tuple[float, float, float]]:
        if self.data_model is None:
            return None
        box = self.data_model.parsed.box
        if not all(axis in box for axis in ("x", "y", "z")):
            return None
        lengths = tuple(float(box[axis][1] - box[axis][0]) for axis in ("x", "y", "z"))
        if any(length <= 0 for length in lengths):
            return None
        return lengths

    def _base_default_coord_for_atom(self, atom: AtomRecord, lengths: Optional[Tuple[float, float, float]]) -> List[float]:
        coord = [float(atom.x), float(atom.y), float(atom.z)]
        if lengths is None:
            return coord
        return [coord[i] + float(atom.image[i]) * lengths[i] for i in range(3)]

    def _wrapped_coord_for_atom(self, atom: AtomRecord) -> List[float]:
        coord = [float(atom.x), float(atom.y), float(atom.z)]
        box = self.data_model.parsed.box if self.data_model is not None else {}
        if not all(axis in box for axis in ("x", "y", "z")):
            return coord
        wrapped = coord[:]
        for i, axis in enumerate(("x", "y", "z")):
            lo, hi = box[axis]
            length = float(hi - lo)
            if length > 0:
                wrapped[i] = float(lo) + ((wrapped[i] - float(lo)) % length)
        return wrapped

    def _topology_display_edges(self) -> List[Tuple[int, int]]:
        if self.data_model is None:
            return []
        edges: List[Tuple[int, int]] = []

        def add_chain(atom_ids: Sequence[int]) -> None:
            for atom1, atom2 in zip(atom_ids, atom_ids[1:]):
                if atom1 in self.data_model.atoms and atom2 in self.data_model.atoms:
                    edges.append((int(atom1), int(atom2)))

        for bond in self.data_model.bonds:
            if bond.atom1 in self.data_model.atoms and bond.atom2 in self.data_model.atoms:
                edges.append((bond.atom1, bond.atom2))
        for angle in self.data_model.angles:
            add_chain(angle.atom_ids)
        for dihedral in self.data_model.dihedrals:
            add_chain(dihedral.atom_ids)
        for improper in self.data_model.impropers:
            ids = tuple(atom_id for atom_id in improper.atom_ids if atom_id in self.data_model.atoms)
            if len(ids) >= 2:
                for atom_id in ids[1:]:
                    edges.append((ids[0], atom_id))
        return edges

    def _build_default_display_coord_cache(self) -> Dict[int, List[float]]:
        if self.data_model is None:
            return {}
        lengths = self._box_lengths()
        base_coords = {
            atom_id: self._base_default_coord_for_atom(atom, lengths)
            for atom_id, atom in self.data_model.atoms.items()
        }
        if lengths is None:
            return base_coords

        adjacency: Dict[int, List[int]] = defaultdict(list)
        for atom1, atom2 in self._topology_display_edges():
            adjacency[atom1].append(atom2)
            adjacency[atom2].append(atom1)

        display_coords: Dict[int, List[float]] = {}
        for root in sorted(self.data_model.atoms):
            if root in display_coords:
                continue
            display_coords[root] = list(base_coords[root])
            queue = deque([root])
            while queue:
                current = queue.popleft()
                current_coord = display_coords[current]
                for neighbor in adjacency.get(current, []):
                    if neighbor in display_coords:
                        continue
                    neighbor_coord = list(base_coords[neighbor])
                    for i, length in enumerate(lengths):
                        if length > 0:
                            neighbor_coord[i] += round((current_coord[i] - neighbor_coord[i]) / length) * length
                    display_coords[neighbor] = neighbor_coord
                    queue.append(neighbor)
        return display_coords

    def _display_coord_for_atom(self, atom: AtomRecord) -> List[float]:
        box = self.data_model.parsed.box if self.data_model is not None else {}
        if self._display_mode != "default":
            return self._wrapped_coord_for_atom(atom) if all(axis in box for axis in ("x", "y", "z")) else [float(atom.x), float(atom.y), float(atom.z)]
        if not self._display_coord_cache:
            self._display_coord_cache = self._build_default_display_coord_cache()
        return self._display_coord_cache.get(atom.atom_id, [float(atom.x), float(atom.y), float(atom.z)])

    def _display_coord_for_atom_id(self, atom_id: int) -> List[float]:
        return self._display_coord_for_atom(self.data_model.atoms[atom_id])

    def _fast_render_mode(self, visible_count: Optional[int] = None) -> bool:
        if visible_count is None:
            if self.points_mesh is not None:
                visible_count = int(self.points_mesh.n_points)
            elif self.data_model is not None:
                visible_count = len(self.data_model.visible_atom_ids())
            else:
                visible_count = 0
        return visible_count >= 10000

    def _tube_sides_for_scene(self, visible_count: Optional[int] = None) -> int:
        return 8 if self._fast_render_mode(visible_count) else 18

    def refresh_scene(self, reset_camera: bool = False) -> None:
        if self.plotter is None:
            return
        if self.data_model is None:
            self.plotter.clear()
            return

        self._display_coord_cache = {}
        camera_position = None if reset_camera else self.plotter.camera_position
        self.plotter.clear()
        self.plotter.show_axes()
        self.atom_mesh = None
        self.atom_actor = None
        self.bond_actor = None
        self._scene_visible_bonds = []
        self.selected_actor = None
        self.selected_bond_actor = None
        self.selected_label_actor = None
        self.single_label_actor = None
        self.box_actor = None

        if self._show_box:
            self._add_box_actor()

        visible_ids = self.data_model.visible_atom_ids()
        if not visible_ids:
            self.plotter.render()
            self.update_summary_panel()
            return

        coords = np.array([self._display_coord_for_atom_id(aid) for aid in visible_ids], dtype=float)
        types = np.array([self.data_model.atoms[aid].atom_type for aid in visible_ids], dtype=int)
        colors = np.array([self._color_for_type(t) for t in types], dtype=np.uint8)
        radii = np.array([self._radius_for_type(t) for t in types], dtype=float)
        fast_render = self._fast_render_mode(len(visible_ids))

        self.points_mesh = pv.PolyData(coords)
        self.points_mesh["orig_id"] = np.array(visible_ids, dtype=np.int32)
        self.points_mesh["atom_uid"] = np.array([self.data_model.atoms[aid].uid for aid in visible_ids], dtype=np.int64)
        self.points_mesh["atype"] = types
        self.points_mesh["rgb"] = colors
        self.points_mesh["radius"] = radii

        atom_mesh = self._build_atom_mesh(self.points_mesh)
        self.atom_mesh = atom_mesh
        self.atom_actor = self.plotter.add_mesh(
            atom_mesh,
            scalars="rgb",
            rgb=True,
            smooth_shading=not fast_render,
            ambient=0.22,
            diffuse=0.72,
            specular=0.18,
            specular_power=20,
            pickable=True,
        )

        self._scene_visible_bonds = self._visible_bonds_for_scene(visible_ids)
        bond_mesh = self._build_bond_mesh(visible_ids, bonds=self._scene_visible_bonds)
        if bond_mesh is not None:
            tube_radius = self._bond_radius_for_scene(radii)
            try:
                bond_render_mesh = bond_mesh.tube(radius=tube_radius, n_sides=self._tube_sides_for_scene(len(visible_ids)), capping=True)
                self.bond_actor = self.plotter.add_mesh(
                    bond_render_mesh,
                    color="#7a7a7a",
                    smooth_shading=not fast_render,
                    ambient=0.18,
                    diffuse=0.75,
                    specular=0.08,
                    pickable=True,
                )
            except Exception:
                self.bond_actor = self.plotter.add_mesh(
                    bond_mesh,
                    color="#6f6f6f",
                    line_width=max(2, int(round(tube_radius * 12))),
                    pickable=True,
                )
        else:
            self.bond_actor = None

        self._update_selection_visuals(render=False)

        if camera_position is not None:
            try:
                self.plotter.camera_position = camera_position
            except Exception:
                pass
        else:
            self.plotter.reset_camera()

        self.plotter.render()
        self.update_summary_panel()

    def _build_points_mesh_for_visible(self, visible_ids: List[int]) -> Tuple[pv.PolyData, np.ndarray]:
        coords = np.array([self._display_coord_for_atom_id(aid) for aid in visible_ids], dtype=float)
        types = np.array([self.data_model.atoms[aid].atom_type for aid in visible_ids], dtype=int)
        colors = np.array([self._color_for_type(t) for t in types], dtype=np.uint8)
        radii = np.array([self._radius_for_type(t) for t in types], dtype=float)
        points_mesh = pv.PolyData(coords)
        points_mesh["orig_id"] = np.array(visible_ids, dtype=np.int32)
        points_mesh["atom_uid"] = np.array([self.data_model.atoms[aid].uid for aid in visible_ids], dtype=np.int64)
        points_mesh["atype"] = types
        points_mesh["rgb"] = colors
        points_mesh["radius"] = radii
        return points_mesh, radii

    def _add_atom_actor_for_mesh(self, points_mesh: pv.PolyData, visible_count: int):
        atom_mesh = self._build_atom_mesh(points_mesh)
        fast_render = self._fast_render_mode(visible_count)
        actor = self.plotter.add_mesh(
            atom_mesh,
            scalars="rgb",
            rgb=True,
            smooth_shading=not fast_render,
            ambient=0.22,
            diffuse=0.72,
            specular=0.18,
            specular_power=20,
            pickable=True,
            reset_camera=False,
            render=False,
        )
        return atom_mesh, actor

    def _add_bond_actor_for_visible(self, visible_ids: List[int], radii: np.ndarray):
        self._scene_visible_bonds = self._visible_bonds_for_scene(visible_ids)
        bond_mesh = self._build_bond_mesh(visible_ids, bonds=self._scene_visible_bonds)
        if bond_mesh is None:
            return None
        tube_radius = self._bond_radius_for_scene(radii)
        fast_render = self._fast_render_mode(len(visible_ids))
        try:
            bond_render_mesh = bond_mesh.tube(radius=tube_radius, n_sides=self._tube_sides_for_scene(len(visible_ids)), capping=True)
            return self.plotter.add_mesh(
                bond_render_mesh,
                color="#7a7a7a",
                smooth_shading=not fast_render,
                ambient=0.18,
                diffuse=0.75,
                specular=0.08,
                pickable=True,
                reset_camera=False,
                render=False,
            )
        except Exception:
            return self.plotter.add_mesh(
                bond_mesh,
                color="#6f6f6f",
                line_width=max(2, int(round(tube_radius * 12))),
                pickable=True,
                reset_camera=False,
                render=False,
            )

    def refresh_bond_actor(self) -> None:
        if self.plotter is None or self.data_model is None or self.points_mesh is None:
            return
        visible_ids = list(map(int, np.asarray(self.points_mesh["orig_id"])))
        radii = np.asarray(self.points_mesh["radius"], dtype=float)
        old_bond_actor = self.bond_actor
        self._display_coord_cache = {}
        new_bond_actor = self._add_bond_actor_for_visible(visible_ids, radii)
        self.bond_actor = new_bond_actor
        self._remove_actor(old_bond_actor)

    def refresh_scene_actor_swap(self) -> None:
        if self.plotter is None:
            return
        if self.data_model is None:
            self.refresh_scene()
            return

        self._display_coord_cache = {}
        visible_ids = self.data_model.visible_atom_ids()
        old_atom_actor = self.atom_actor
        old_bond_actor = self.bond_actor
        old_selected_actor = self.selected_actor
        old_selected_bond_actor = self.selected_bond_actor
        old_selected_label_actor = self.selected_label_actor
        old_single_label_actor = self.single_label_actor

        if not visible_ids:
            self.points_mesh = None
            self.atom_mesh = None
            self.atom_actor = None
            self.bond_actor = None
            self._scene_visible_bonds = []
            for actor in (old_atom_actor, old_bond_actor, old_selected_actor, old_selected_bond_actor, old_selected_label_actor, old_single_label_actor):
                self._remove_actor(actor)
            self.plotter.render()
            return

        points_mesh, radii = self._build_points_mesh_for_visible(visible_ids)
        atom_mesh, new_atom_actor = self._add_atom_actor_for_mesh(points_mesh, len(visible_ids))
        new_bond_actor = self._add_bond_actor_for_visible(visible_ids, radii)

        self.points_mesh = points_mesh
        self.atom_mesh = atom_mesh
        self.atom_actor = new_atom_actor
        self.bond_actor = new_bond_actor
        self.selected_actor = old_selected_actor
        self.selected_bond_actor = old_selected_bond_actor
        self.selected_label_actor = old_selected_label_actor
        self.single_label_actor = old_single_label_actor

        for actor in (old_atom_actor, old_bond_actor):
            self._remove_actor(actor)
        self._update_selection_visuals(render=False)
        self.plotter.render()

    def _visible_bonds_for_scene(self, visible_ids: List[int], bonds: Optional[Sequence[BondRecord]] = None) -> List[BondRecord]:
        visible_set = set(visible_ids)
        source_bonds = self.data_model.bonds if bonds is None else list(bonds)
        visible_bonds = [b for b in source_bonds if b.atom1 in visible_set and b.atom2 in visible_set]
        if self._display_mode == "in-cell" and self._hide_cross_boundary_bonds:
            visible_bonds = [b for b in visible_bonds if not self._is_cross_boundary_bond(b)]
        return visible_bonds

    def _build_bond_mesh(self, visible_ids: List[int], bonds: Optional[Sequence[BondRecord]] = None) -> Optional[pv.PolyData]:
        visible_bonds = self._visible_bonds_for_scene(visible_ids, bonds=bonds)
        if not visible_bonds:
            return None

        idx_map = {aid: i for i, aid in enumerate(visible_ids)}
        points = np.array([self._display_coord_for_atom_id(aid) for aid in visible_ids], dtype=float)
        lines: List[int] = []
        for bond in visible_bonds:
            lines.extend([2, idx_map[bond.atom1], idx_map[bond.atom2]])
        mesh = pv.PolyData(points)
        mesh.lines = np.array(lines, dtype=np.int64)
        return mesh

    def _is_cross_boundary_bond(self, bond: BondRecord) -> bool:
        box = self.data_model.parsed.box if self.data_model is not None else {}
        if not all(axis in box for axis in ("x", "y", "z")):
            return False
        a1 = self.data_model.atoms[bond.atom1]
        a2 = self.data_model.atoms[bond.atom2]
        lengths = {axis: box[axis][1] - box[axis][0] for axis in ("x", "y", "z")}
        deltas = {
            "x": abs(a1.x - a2.x),
            "y": abs(a1.y - a2.y),
            "z": abs(a1.z - a2.z),
        }
        return any(lengths[axis] > 0 and deltas[axis] > 0.5 * lengths[axis] for axis in ("x", "y", "z"))

    def _add_box_actor(self) -> None:
        box = self.data_model.parsed.box if self.data_model is not None else {}
        if not all(axis in box for axis in ("x", "y", "z")):
            self.box_actor = None
            return
        bounds = (
            box["x"][0], box["x"][1],
            box["y"][0], box["y"][1],
            box["z"][0], box["z"][1],
        )
        try:
            outline = pv.Box(bounds=bounds).outline()
            self.box_actor = self.plotter.add_mesh(outline, color="#202020", line_width=1.5, pickable=False)
        except Exception:
            self.box_actor = None

    def _build_atom_mesh(self, points_mesh: pv.PolyData) -> pv.PolyData:
        if points_mesh.n_points == 0:
            return pv.PolyData()
        geom = self._sphere_geom_fast if self._fast_render_mode(int(points_mesh.n_points)) else self._sphere_geom
        return points_mesh.glyph(scale="radius", geom=geom, orient=False, factor=1.0)

    def _remove_actor(self, actor) -> None:
        if actor is None:
            return
        try:
            self.plotter.remove_actor(actor, reset_camera=False, render=False)
        except Exception:
            pass

    def _update_selection_visuals(self, render: bool = True) -> None:
        self._remove_actor(self.selected_actor)
        self.selected_actor = None
        self._remove_actor(self.selected_bond_actor)
        self.selected_bond_actor = None
        self._remove_actor(self.selected_label_actor)
        self.selected_label_actor = None
        self._remove_actor(self.single_label_actor)
        self.single_label_actor = None

        if self.data_model is None or self.points_mesh is None:
            if render:
                self.plotter.render()
            return

        visible_ids = set(map(int, np.asarray(self.points_mesh["orig_id"])))
        fast_render = self._fast_render_mode(len(visible_ids))
        selected_visible = [aid for aid in sorted(self.data_model.current_pattern_atom_ids()) if aid in visible_ids]
        if selected_visible:
            sel_pts = np.array([self._display_coord_for_atom_id(aid) for aid in selected_visible], dtype=float)
            sel_mesh = pv.PolyData(sel_pts)
            sel_mesh["radius"] = np.array(
                [self._radius_for_type(self.data_model.atoms[aid].atom_type) * self._selection_shell_scale for aid in selected_visible],
                dtype=float,
            )
            sel_shell = self._build_atom_mesh(sel_mesh)
            self.selected_actor = self.plotter.add_mesh(
                sel_shell,
                color="#ff4d4d",
                opacity=0.22,
                smooth_shading=not fast_render,
                pickable=False,
                reset_camera=False,
                render=False,
            )

        if self.data_model.selection_kind == "bond":
            visible_list = list(map(int, np.asarray(self.points_mesh["orig_id"])))
            selected_bond_mesh = self._build_bond_mesh(visible_list, bonds=self.data_model.current_selected_bonds())
            if selected_bond_mesh is not None:
                tube_radius = max(self._bond_radius_for_scene(np.asarray(self.points_mesh["radius"])), 0.14) * 1.35
                try:
                    selected_render_mesh = selected_bond_mesh.tube(radius=tube_radius, n_sides=self._tube_sides_for_scene(), capping=True)
                    self.selected_bond_actor = self.plotter.add_mesh(
                        selected_render_mesh,
                        color="#ff6b57",
                        smooth_shading=not fast_render,
                        ambient=0.2,
                        diffuse=0.8,
                        specular=0.15,
                        pickable=False,
                        reset_camera=False,
                        render=False,
                    )
                except Exception:
                    self.selected_bond_actor = self.plotter.add_mesh(
                        selected_bond_mesh,
                        color="#ff6b57",
                        line_width=5,
                        pickable=False,
                        reset_camera=False,
                        render=False,
                    )

        if self.last_picked_bond is not None and self.data_model.selection_kind == "bond":
            picked_bond = next((bond for bond in self._scene_visible_bonds if bond.bond_id == self.last_picked_bond), None)
            if picked_bond is not None:
                coord1 = self._display_coord_for_atom_id(picked_bond.atom1)
                coord2 = self._display_coord_for_atom_id(picked_bond.atom2)
                mid_point = np.array(
                    [[(coord1[0] + coord2[0]) * 0.5, (coord1[1] + coord2[1]) * 0.5, (coord1[2] + coord2[2]) * 0.5]],
                    dtype=float,
                )
                label_text = [f"id={picked_bond.bond_id}\ntype={picked_bond.bond_type}"]
                self.single_label_actor = self.plotter.add_point_labels(
                    mid_point,
                    label_text,
                    point_size=0,
                    font_size=12,
                    always_visible=True,
                    pickable=False,
                    shape_opacity=0.7,
                    show_points=False,
                    reset_camera=False,
                    render=False,
                )

        elif self.last_picked_atom is not None and self.last_picked_atom in visible_ids:
            atom = self.data_model.atoms[self.last_picked_atom]
            label_pts = np.array([self._display_coord_for_atom(atom)], dtype=float)
            label_text = [f"id={atom.atom_id}\ntype={atom.atom_type}"]
            self.single_label_actor = self.plotter.add_point_labels(
                label_pts,
                label_text,
                point_size=0,
                font_size=12,
                always_visible=True,
                pickable=False,
                shape_opacity=0.7,
                show_points=False,
                reset_camera=False,
                render=False,
            )

        if render:
            self.plotter.render()

    def _bond_radius_for_scene(self, atom_radii: np.ndarray) -> float:
        if atom_radii.size == 0:
            return self._bond_radius
        return float(min(0.18, max(0.10, np.median(atom_radii) * 0.28)))

    def _color_for_type(self, atom_type: int) -> Tuple[int, int, int]:
        if self.data_model is not None:
            custom_color = self.data_model.type_colors.get(int(atom_type))
            if custom_color is not None:
                return custom_color
            element_info = infer_element_from_mass(self.data_model.masses.get(int(atom_type)))
            if element_info is not None:
                element_color = MATERIAL_STUDIO_ELEMENT_COLORS.get(element_info[0])
                if element_color is not None:
                    return element_color
        # Fallback for undefined masses/elements.
        hue = (atom_type * 0.61803398875) % 1.0
        sat = 0.55
        val = 0.92
        r, g, b = hsv_to_rgb(hue, sat, val)
        return int(r * 255), int(g * 255), int(b * 255)

    def _radius_for_type(self, atom_type: int) -> float:
        if self.data_model is None:
            return 0.5
        mass = self.data_model.masses.get(atom_type)
        if mass is None or mass <= 0:
            return 0.52 * self._atom_radius_scale
        radius = (0.25 + 0.10 * (float(mass) ** (1.0 / 3.0))) * self._atom_radius_scale
        return float(min(0.82, max(0.32, radius)))

    def _pixel_ratio(self) -> float:
        scale = 1.0
        if hasattr(self.plotter.interactor, "_getPixelRatio"):
            try:
                scale = float(self.plotter.interactor._getPixelRatio())
            except Exception:
                scale = 1.0
        return scale

    def _project_atom_to_widget(self, atom_id: int) -> Optional[QtCore.QPoint]:
        if self.data_model is None or atom_id not in self.data_model.atoms:
            return None
        coord = self._display_coord_for_atom_id(atom_id)
        renderer = self.plotter.renderer
        try:
            renderer.SetWorldPoint(coord[0], coord[1], coord[2], 1.0)
            renderer.WorldToDisplay()
            display_x, display_y, display_z = renderer.GetDisplayPoint()
        except Exception:
            return None
        if not all(math.isfinite(value) for value in (display_x, display_y, display_z)):
            return None
        scale = self._pixel_ratio()
        x = display_x / scale
        y = float(self.plotter.interactor.height()) - 1.0 - (display_y / scale)
        if not all(math.isfinite(value) for value in (x, y)):
            return None
        return QtCore.QPoint(int(round(x)), int(round(y)))

    @staticmethod
    def _distance_point_to_segment(point: np.ndarray, start: np.ndarray, end: np.ndarray) -> float:
        segment = end - start
        length_sq = float(np.dot(segment, segment))
        if length_sq <= 1e-18:
            return float(np.linalg.norm(point - start))
        projection = float(np.dot(point - start, segment) / length_sq)
        projection = max(0.0, min(1.0, projection))
        closest = start + projection * segment
        return float(np.linalg.norm(point - closest))

    def _camera_depth(self, world_point: np.ndarray) -> float:
        camera = self.plotter.renderer.GetActiveCamera()
        camera_pos = np.asarray(camera.GetPosition(), dtype=float)
        direction = np.asarray(camera.GetDirectionOfProjection(), dtype=float)
        norm = float(np.linalg.norm(direction))
        if norm <= 1e-12 or not np.all(np.isfinite(direction)):
            return float(np.linalg.norm(np.asarray(world_point, dtype=float) - camera_pos))
        direction /= norm
        return float(np.dot(np.asarray(world_point, dtype=float) - camera_pos, direction))

    def _pick_atom_hit(self, pos: QtCore.QPoint) -> Tuple[Optional[int], Optional[float]]:
        if self.data_model is None or self.atom_mesh is None:
            return None, None
        if self.atom_actor is None:
            return None, None
        scale = self._pixel_ratio()
        x = int(round(float(pos.x()) * scale))
        y = int(round((float(self.plotter.interactor.height()) - float(pos.y()) - 1.0) * scale))
        picker = vtkCellPicker()
        picker.SetTolerance(0.005)
        try:
            picker.PickFromListOn()
            picker.AddPickList(self.atom_actor)
            picker.Pick(x, y, 0, self.plotter.renderer)
        except Exception:
            return None, None
        cell_id = picker.GetCellId()
        if cell_id is None or cell_id < 0 or cell_id >= self.atom_mesh.n_cells:
            return None, None
        try:
            cell = self.atom_mesh.get_cell(cell_id)
        except Exception:
            return None, None
        orig_ids = np.asarray(self.atom_mesh["orig_id"])[cell.point_ids]
        if orig_ids.size == 0:
            return None, None
        pick_position = np.asarray(picker.GetPickPosition(), dtype=float)
        uid_map = self.data_model.atom_uid_map()
        if "atom_uid" in self.atom_mesh.array_names:
            atom_uids = np.asarray(self.atom_mesh["atom_uid"])[cell.point_ids]
            for atom_uid in atom_uids:
                atom_id = uid_map.get(int(atom_uid))
                if atom_id is not None:
                    return atom_id, self._camera_depth(pick_position)
        return int(orig_ids[0]), self._camera_depth(pick_position)

    def _pick_bond_hit(self, pos: QtCore.QPoint) -> Tuple[Optional[int], Optional[float]]:
        if self.data_model is None or self.bond_actor is None or not self._scene_visible_bonds:
            return None, None
        scale = self._pixel_ratio()
        x = int(round(float(pos.x()) * scale))
        y = int(round((float(self.plotter.interactor.height()) - float(pos.y()) - 1.0) * scale))
        picker = vtkCellPicker()
        picker.SetTolerance(0.005)
        try:
            picker.PickFromListOn()
            picker.AddPickList(self.bond_actor)
            picker.Pick(x, y, 0, self.plotter.renderer)
        except Exception:
            return None, None
        cell_id = picker.GetCellId()
        if cell_id is None or cell_id < 0:
            return None, None
        pick_position = np.asarray(picker.GetPickPosition(), dtype=float)
        best_bond_id: Optional[int] = None
        best_distance = float("inf")
        for bond in self._scene_visible_bonds:
            start = np.array(self._display_coord_for_atom_id(bond.atom1), dtype=float)
            end = np.array(self._display_coord_for_atom_id(bond.atom2), dtype=float)
            distance = self._distance_point_to_segment(pick_position, start, end)
            if distance < best_distance:
                best_distance = distance
                best_bond_id = bond.bond_id
        if best_bond_id is None:
            return None, None
        tolerance = max(self._bond_radius * 2.5, 0.4)
        if best_distance > tolerance:
            return None, None
        return best_bond_id, self._camera_depth(pick_position)

    def _pick_front_target(self, pos: QtCore.QPoint) -> Tuple[Optional[str], Optional[int]]:
        atom_id, atom_depth = self._pick_atom_hit(pos)
        bond_id, bond_depth = self._pick_bond_hit(pos)
        if atom_id is None and bond_id is None:
            return None, None
        if atom_id is None:
            return "bond", bond_id
        if bond_id is None:
            return "atom", atom_id
        if bond_depth is not None and atom_depth is not None and bond_depth <= atom_depth + 1e-4:
            return "bond", bond_id
        return "atom", atom_id

    def pick_atom_at(self, pos: QtCore.QPoint) -> Optional[int]:
        atom_id, _ = self._pick_atom_hit(pos)
        return atom_id

    def pick_bond_at(self, pos: QtCore.QPoint) -> Optional[int]:
        bond_id, _ = self._pick_bond_hit(pos)
        return bond_id

    def update_summary_panel(self) -> None:
        if self.data_model is None:
            self.summary_label.setText("未加载文件")
            self._populate_type_stats_table(self.atom_type_table, [])
            self._populate_type_stats_table(self.bond_type_table, [])
            self._populate_type_stats_table(self.angle_type_table, [])
            self._populate_type_stats_table(self.dihedral_type_table, [])
            self._populate_type_stats_table(self.improper_type_table, [])
            return

        total_atoms = len(self.data_model.atoms)
        total_bonds = len(self.data_model.bonds)
        total_angles = len(self.data_model.angles)
        total_dihedrals = len(self.data_model.dihedrals)
        total_impropers = len(self.data_model.impropers)
        visible_atoms = len(self.data_model.visible_atom_ids())
        selected = len(self.data_model.current_selection_ids())
        selection_kind = self.data_model.selection_kind or "-"
        fragments = len(self.data_model.fragments)
        box_text = []
        for axis in ("x", "y", "z"):
            if axis in self.data_model.parsed.box:
                lo, hi = self.data_model.parsed.box[axis]
                box_text.append(f"{axis}: [{lo:.3f}, {hi:.3f}]")
        box_summary = " | ".join(box_text) if box_text else "未解析到盒子边界"
        self.summary_label.setText(
            f"文件: {self.current_file.name if self.current_file else '-'}\n"
            f"atom_style: {self.data_model.atom_style}\n"
            f"display mode: {self._display_mode}\n"
            f"原子: {total_atoms}（当前可见 {visible_atoms}）\n"
            f"键/角/二面角/反二面角: {total_bonds}/{total_angles}/{total_dihedrals}/{total_impropers}\n"
            f"fragment: {fragments}\n"
            f"当前选择: {selection_kind} × {selected}\n"
            f"盒子: {box_summary}"
        )

        atom_type_counts: Dict[int, int] = defaultdict(int)
        for atom in self.data_model.atoms.values():
            atom_type_counts[atom.atom_type] += 1
        self._populate_type_stats_table(
            self.atom_type_table,
            sorted(atom_type_counts.items(), key=lambda item: item[0]),
            masses=self.data_model.masses,
            labels=self.data_model.mass_names,
            ff_labels=self.data_model.mass_labels,
        )

        bond_type_counts: Dict[int, int] = defaultdict(int)
        for bond in self.data_model.bonds:
            bond_type_counts[bond.record_type] += 1
        self._populate_type_stats_table(
            self.bond_type_table,
            sorted(bond_type_counts.items(), key=lambda item: item[0]),
            labels=self.data_model.coeff_names.get("bond", {}),
            ff_labels=self.data_model.coeff_labels.get("bond", {}),
        )

        angle_type_counts: Dict[int, int] = defaultdict(int)
        for angle in self.data_model.angles:
            angle_type_counts[angle.record_type] += 1
        self._populate_type_stats_table(
            self.angle_type_table,
            sorted(angle_type_counts.items(), key=lambda item: item[0]),
            labels=self.data_model.coeff_names.get("angle", {}),
            ff_labels=self.data_model.coeff_labels.get("angle", {}),
        )

        dihedral_type_counts: Dict[int, int] = defaultdict(int)
        for dihedral in self.data_model.dihedrals:
            dihedral_type_counts[dihedral.record_type] += 1
        self._populate_type_stats_table(
            self.dihedral_type_table,
            sorted(dihedral_type_counts.items(), key=lambda item: item[0]),
            labels=self.data_model.coeff_names.get("dihedral", {}),
            ff_labels=self.data_model.coeff_labels.get("dihedral", {}),
        )

        improper_type_counts: Dict[int, int] = defaultdict(int)
        for improper in self.data_model.impropers:
            improper_type_counts[improper.record_type] += 1
        self._populate_type_stats_table(
            self.improper_type_table,
            sorted(improper_type_counts.items(), key=lambda item: item[0]),
            labels=self.data_model.coeff_names.get("improper", {}),
            ff_labels=self.data_model.coeff_labels.get("improper", {}),
        )

        type_tab_index = {
            "atom": 0,
            "bond": 1,
            "angle": 2,
            "dihedral": 3,
            "improper": 4,
        }
        if selection_kind in type_tab_index:
            self.type_tabs.setCurrentIndex(type_tab_index[selection_kind])

    def update_selection_panel(self) -> None:
        self.update_properties_panel()
        self.update_charge_panel()
        if self.data_model is None:
            self.selection_text.setPlainText("")
            return
        if not self.data_model.has_selection():
            self.selection_text.setPlainText("无选择\n\n操作提示：左键单击选择；左键拖动框选；Shift 追加；Ctrl 切换；Alt+双击选择当前显示的同元素原子；右键拖动旋转；中键拖动平移。")
            return

        if self.data_model.selection_kind == "atom":
            ids = sorted(self.data_model.selected_atoms)
            preview = ids[:20]
            lines = [f"选中原子数: {len(ids)}", f"ID 列表(前20个): {' '.join(map(str, preview))}"]
            if len(ids) == 1:
                atom = self.data_model.atoms[ids[0]]
                frag = self.data_model.fragment_of_atom.get(ids[0], "-")
                lines.extend(
                    [
                        f"atom id = {atom.atom_id}",
                        f"type = {atom.atom_type}",
                        f"mol = {atom.mol if atom.mol is not None else '-'}",
                        f"charge = {atom.charge if atom.charge is not None else '-'}",
                        f"fragment = {frag}",
                        f"xyz = ({atom.x:.6f}, {atom.y:.6f}, {atom.z:.6f})",
                    ]
                )
            else:
                type_counts: Dict[int, int] = defaultdict(int)
                frag_counts: Dict[int, int] = defaultdict(int)
                for aid in ids:
                    type_counts[self.data_model.atoms[aid].atom_type] += 1
                    frag = self.data_model.fragment_of_atom.get(aid)
                    if frag is not None:
                        frag_counts[frag] += 1
                lines.append("type 统计: " + ", ".join(f"{k}:{v}" for k, v in sorted(type_counts.items())))
                lines.append("fragment 涉及: " + ", ".join(f"{k}({v})" for k, v in sorted(frag_counts.items())[:10]))
            self.selection_text.setPlainText("\n".join(lines))
            return

        kind = self.data_model.selection_kind or "-"
        ids = sorted(self.data_model.current_selection_ids())
        atom_ids = sorted(self.data_model.current_pattern_atom_ids())
        selected_id_set = set(ids)
        records = [record for record in self.data_model.topology_records(kind) if record.record_id in selected_id_set]
        type_counts: Dict[int, int] = defaultdict(int)
        for record in records:
            type_counts[record.record_type] += 1
        type_text = ", ".join(f"{type_id}({count})" for type_id, count in sorted(type_counts.items())) or "-"
        lines = [
            f"选中{kind}数: {len(ids)}",
            f"{kind} ID 列表(前20个): {' '.join(map(str, ids[:20]))}",
            f"{kind} type: {type_text}",
            f"包含原子数: {len(atom_ids)}",
            f"原子 ID 列表(前20个): {' '.join(map(str, atom_ids[:20]))}",
        ]
        if len(records) == 1:
            record = records[0]
            lines.extend(
                [
                    f"{kind} id = {record.record_id}",
                    f"type = {record.record_type}",
                    f"atoms = {' '.join(map(str, record.atom_ids))}",
                ]
            )
        self.selection_text.setPlainText("\n".join(lines))

    def _charge_sum_for_atom_ids(self, atom_ids: Set[int]) -> Tuple[float, int]:
        if self.data_model is None:
            return 0.0, 0
        total = 0.0
        missing = 0
        for atom_id in atom_ids:
            atom = self.data_model.atoms.get(atom_id)
            if atom is None:
                continue
            if atom.charge is None:
                missing += 1
                continue
            total += float(atom.charge)
        return total, missing

    def update_charge_panel(self) -> None:
        if not hasattr(self, "charge_label"):
            return
        if self.data_model is None:
            self.charge_label.setText("No model loaded.")
            return

        all_ids = set(self.data_model.atoms)
        visible_ids = set(self.data_model.visible_atom_ids())
        selected_ids = self.data_model.current_pattern_atom_ids()
        model_charge, model_missing = self._charge_sum_for_atom_ids(all_ids)
        visible_charge, visible_missing = self._charge_sum_for_atom_ids(visible_ids)
        selected_charge, selected_missing = self._charge_sum_for_atom_ids(selected_ids)

        selected_line = (
            f"Selected ({len(selected_ids)} atoms): {selected_charge:.6g}"
            if selected_ids
            else "Selected (0 atoms): -"
        )
        missing_parts = []
        if selected_missing:
            missing_parts.append(f"selected missing q: {selected_missing}")
        if visible_missing:
            missing_parts.append(f"visible missing q: {visible_missing}")
        if model_missing:
            missing_parts.append(f"model missing q: {model_missing}")
        missing_text = "\n" + "\n".join(missing_parts) if missing_parts else ""
        self.charge_label.setText(
            f"{selected_line}\n"
            f"Visible ({len(visible_ids)} atoms): {visible_charge:.6g}\n"
            f"Model ({len(all_ids)} atoms): {model_charge:.6g}"
            f"{missing_text}"
        )

    def _selected_atom_property_context(self) -> Optional[Tuple[int, List[AtomRecord]]]:
        if self.data_model is None or self.data_model.selection_kind != "atom" or not self.data_model.selected_atoms:
            return None
        selected_atoms = [self.data_model.atoms[aid] for aid in sorted(self.data_model.selected_atoms) if aid in self.data_model.atoms]
        atom_types = {atom.atom_type for atom in selected_atoms}
        if len(atom_types) != 1:
            return None
        return next(iter(atom_types)), selected_atoms

    def _set_property_row(self, row: int, name: str, value: str, editable: bool = False) -> None:
        name_item = self.properties_table.item(row, 0)
        if name_item is None:
            name_item = QtWidgets.QTableWidgetItem()
            self.properties_table.setItem(row, 0, name_item)
        name_item.setText(name)
        name_item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)

        value_item = self.properties_table.item(row, 1)
        if value_item is None:
            value_item = QtWidgets.QTableWidgetItem()
            self.properties_table.setItem(row, 1, value_item)
        value_item.setText(value)
        value_item.setBackground(QtGui.QBrush())
        value_item.setForeground(QtGui.QBrush())
        value_item.setToolTip("")
        flags = QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable
        if editable:
            flags |= QtCore.Qt.ItemIsEditable
        value_item.setFlags(flags)

    def _set_property_color_swatch(self, row: int, color: Tuple[int, int, int]) -> None:
        value_item = self.properties_table.item(row, 1)
        if value_item is None:
            return
        qcolor = QtGui.QColor(*color)
        value_item.setBackground(QtGui.QBrush(qcolor))
        value_item.setToolTip("Double click to edit atom type color")
        luminance = 0.2126 * color[0] + 0.7152 * color[1] + 0.0722 * color[2]
        text_color = QtGui.QColor(0, 0, 0) if luminance >= 150 else QtGui.QColor(255, 255, 255)
        value_item.setForeground(QtGui.QBrush(text_color))

    def update_properties_panel(self) -> None:
        if not hasattr(self, "properties_table"):
            return

        self._updating_properties_table = True
        try:
            self._properties_context_atom_type = None
            rows = [
                ("charge", "-", False),
                ("element", "未定义", False),
                ("mass", "未定义", False),
                ("Name", "未定义", False),
                ("ff-type", "未定义", False),
            ]
            hint_text = "请选择单个原子或同 type 原子以查看属性。"

            rows.append(("color", "-", False))

            context = self._selected_atom_property_context()
            if context is not None:
                atom_type, selected_atoms = context
                self._properties_context_atom_type = atom_type
                mass = self.data_model.masses.get(atom_type)
                mass_text = "未定义" if mass is None else f"{mass:g}"
                element_info = infer_element_from_mass(mass)
                element_text = "未定义" if element_info is None else f"{element_info[1]} ({element_info[0]})"
                name_text = self.data_model.mass_names.get(atom_type, "").strip() or "未定义"
                ff_type = self.data_model.mass_labels.get(atom_type, "").strip() or "未定义"

                color = self._color_for_type(atom_type)
                color_text = f"RGB ({color[0]}, {color[1]}, {color[2]})"

                if len(selected_atoms) == 1:
                    charge = selected_atoms[0].charge
                    charge_text = "未定义" if charge is None else f"{charge:.5f}"
                else:
                    charges = [atom.charge for atom in selected_atoms]
                    if charges and all(charge is not None for charge in charges):
                        first_charge = float(charges[0])
                        if all(math.isclose(float(charge), first_charge, rel_tol=1e-9, abs_tol=1e-9) for charge in charges[1:]):
                            charge_text = f"{first_charge:.5f}"
                        else:
                            charge_text = "混合"
                    else:
                        charge_text = "混合"

                rows = [
                    ("charge", charge_text, True),
                    ("element", element_text, False),
                    ("mass", mass_text, True),
                    ("Name", name_text, True),
                    ("ff-type", ff_type, True),
                    ("color", color_text, False),
                ]
                hint_text = f"type={atom_type}，当前选中 {len(selected_atoms)} 个原子。"
            elif self.data_model is not None and self.data_model.selection_kind == "atom":
                hint_text = "当前选择包含多个 type。请选择单个原子或同 type 原子以查看属性。"

            self.properties_hint_label.setText(hint_text)
            self.properties_table.setRowCount(len(rows))
            for row_index, (name, value, editable) in enumerate(rows):
                self._set_property_row(row_index, name, value, editable=editable)
                if name == "color" and context is not None and self._properties_context_atom_type is not None:
                    self._set_property_color_swatch(row_index, self._color_for_type(self._properties_context_atom_type))
            self.properties_table.resizeColumnsToContents()
        finally:
            self._updating_properties_table = False

    def _handle_property_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_properties_table or item.column() != 1 or self.data_model is None:
            return

        atom_type = self._properties_context_atom_type
        if atom_type is None:
            self.update_properties_panel()
            return

        if item.row() == 0:
            text = item.text().strip()
            try:
                new_charge = float(text)
            except ValueError:
                QtWidgets.QMessageBox.warning(self, "电荷无效", "请输入有效的电荷数值。")
                self.update_properties_panel()
                return

            context = self._selected_atom_property_context()
            if context is None:
                self.update_properties_panel()
                return
            _, selected_atoms = context
            selected_ids = {atom.atom_id for atom in selected_atoms}
            changed = self.data_model.set_charge_for_atoms(selected_ids, new_charge)
            self.update_properties_panel()
            self.update_summary_panel()
            self.statusBar().showMessage(f"已将 {changed} 个选中原子的 charge 统一为 {new_charge:g}")
            return

        if item.row() == 2:
            text = item.text().strip()
            try:
                new_mass = float(text)
                if new_mass <= 0:
                    raise ValueError
            except ValueError:
                QtWidgets.QMessageBox.warning(self, "质量无效", "请输入大于 0 的质量数值。")
                self.update_properties_panel()
                return

            current_mass = self.data_model.masses.get(atom_type)
            if current_mass is not None and math.isclose(current_mass, new_mass, rel_tol=1e-9, abs_tol=1e-9):
                return

            self._clear_left_drag_state()
            reply = QtWidgets.QMessageBox.question(
                self,
                "修改质量",
                f"修改质量将同时更新全部 type={atom_type} 的原子，是否继续？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                self.update_properties_panel()
                return

            self.data_model.set_mass_for_type(atom_type, new_mass)
            synced = self._sync_atom_type_property(atom_type, "mass", new_mass)
            self.refresh_scene()
            self.update_selection_panel()
            self.statusBar().showMessage(f"已将 type={atom_type} 的质量更新为 {new_mass:g}，同步 {synced} 个模型")
            return

        if item.row() == 3:
            new_name = item.text().strip()
            if new_name == "未定义":
                new_name = ""
            current_name = self.data_model.mass_names.get(atom_type, "").strip()
            if new_name == current_name:
                return
            self._clear_left_drag_state()
            reply = QtWidgets.QMessageBox.question(
                self,
                "修改 Name",
                f"修改 Name 将同时更新全部 type={atom_type} 的原子注释，是否继续？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                self.update_properties_panel()
                return
            self.data_model.set_name_for_type(atom_type, new_name)
            synced = self._sync_atom_type_property(atom_type, "name", new_name)
            self.update_properties_panel()
            self.update_summary_panel()
            self.statusBar().showMessage(f"已更新 type={atom_type} 的 Name，同步 {synced} 个模型")
            return

        if item.row() == 4:
            new_label = item.text().strip()
            if new_label == "未定义":
                new_label = ""
            current_label = self.data_model.mass_labels.get(atom_type, "").strip()
            if new_label == current_label:
                return

            self._clear_left_drag_state()
            reply = QtWidgets.QMessageBox.question(
                self,
                "修改 ff-type",
                f"修改 ff-type 将同时更新全部 type={atom_type} 的原子，是否继续？",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                self.update_properties_panel()
                return

            self.data_model.set_ff_type_for_type(atom_type, new_label)
            synced = self._sync_atom_type_property(atom_type, "ff", new_label)
            self.update_properties_panel()
            self.update_summary_panel()
            self.statusBar().showMessage(f"已更新 type={atom_type} 的 ff-type，同步 {synced} 个模型")

    def _handle_atom_type_table_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_type_tables or item.column() not in (3, 4) or self.data_model is None:
            return

        type_item = self.atom_type_table.item(item.row(), 0)
        if type_item is None:
            self.update_summary_panel()
            return

        try:
            atom_type = int(type_item.text())
        except ValueError:
            self.update_summary_panel()
            return

        new_label = item.text().strip()
        if new_label == "未定义":
            new_label = ""
        is_name_column = item.column() == 3
        current_label = (self.data_model.mass_names if is_name_column else self.data_model.mass_labels).get(atom_type, "").strip()
        if new_label == current_label:
            return

        self._clear_left_drag_state()
        reply = QtWidgets.QMessageBox.question(
            self,
            "修改 Name" if is_name_column else "修改 ff-type",
            f"修改 {'Name' if is_name_column else 'ff-type'} 将同时更新全部 type={atom_type} 的原子，是否继续？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            self.update_summary_panel()
            return

        if is_name_column:
            self.data_model.set_name_for_type(atom_type, new_label)
        else:
            self.data_model.set_ff_type_for_type(atom_type, new_label)
        self.update_summary_panel()
        self.update_properties_panel()
        self.statusBar().showMessage(f"已更新 type={atom_type} 的 {'Name' if is_name_column else 'ff-type'}")

    def toggle_box_visibility(self) -> None:
        self._show_box = bool(self.toggle_box_action.isChecked())
        self.refresh_scene()

    def toggle_cross_boundary_bonds(self) -> None:
        self._hide_cross_boundary_bonds = bool(self.toggle_cross_boundary_bonds_action.isChecked())
        self.refresh_scene()

    def hide_selected(self) -> None:
        if self.data_model is None:
            return
        self.data_model.hide_selected()
        self.last_picked_atom = None
        self.last_picked_bond = None
        self.refresh_scene()
        self.update_selection_panel()

    def isolate_selected(self) -> None:
        if self.data_model is None:
            return
        self.data_model.isolate_selected()
        self.refresh_scene()
        self.update_selection_panel()

    def show_all(self) -> None:
        if self.data_model is None:
            return
        self.data_model.show_all()
        self.refresh_scene()
        self.update_selection_panel()

    def clear_selection(self) -> None:
        if self.data_model is None:
            return
        self.data_model.clear_selection()
        self.last_picked_atom = None
        self.last_picked_bond = None
        self._update_selection_visuals()
        self.update_summary_panel()
        self.update_selection_panel()

    def select_fragment(self) -> None:
        if self.data_model is None:
            return
        if self.data_model.selection_kind != "atom":
            self.statusBar().showMessage("fragment 选择仅适用于 atom 选择。")
            return
        fragment_atoms = self.data_model.select_fragments_of_selection()
        hidden_hit = bool(fragment_atoms & self.data_model.hidden_atoms)
        self.data_model.reveal_atoms(fragment_atoms)
        self.last_picked_atom = min(fragment_atoms) if len(fragment_atoms) == 1 else None
        self.last_picked_bond = None
        if hidden_hit:
            self.refresh_scene()
        else:
            self._update_selection_visuals()
            self.update_summary_panel()
        self.update_selection_panel()

    def _single_selected_origin_atom(self) -> Optional[AtomRecord]:
        if self.data_model is None:
            return None
        if self.data_model.selection_kind == "atom" and len(self.data_model.selected_atoms) == 1:
            atom_id = next(iter(self.data_model.selected_atoms))
            return self.data_model.atoms.get(atom_id)
        if self.last_picked_atom is not None:
            return self.data_model.atoms.get(self.last_picked_atom)
        return None

    def _position_from_origin_and_spherical(self, origin: AtomRecord, values: Tuple[float, float, float]) -> Tuple[float, float, float]:
        dx, dy, dz = spherical_offset(*values)
        return float(origin.x + dx), float(origin.y + dy), float(origin.z + dz)

    def insert_atom(self) -> None:
        if self.data_model is None:
            return
        origin = self._single_selected_origin_atom()
        if origin is None:
            QtWidgets.QMessageBox.information(self, "插入原子", "请先点选一个原子作为球坐标原点。")
            return
        dialog = InsertAtomDialog(self.data_model, self)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        position = self._position_from_origin_and_spherical(origin, dialog.spherical_values())
        atom_id = self.data_model.insert_atom(dialog.atom_type(), position)
        self.data_model.select_atoms({atom_id})
        self.last_picked_atom = atom_id
        self.last_picked_bond = None
        self.refresh_scene()
        self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(f"已插入 atom id={atom_id}, type={dialog.atom_type()}")

    def insert_fragment(self) -> None:
        if self.data_model is None:
            return
        origin = self._single_selected_origin_atom()
        if origin is None:
            QtWidgets.QMessageBox.information(self, "插入 fragment", "请先点选一个原子作为球坐标原点。")
            return
        if not self.data_model.fragments:
            QtWidgets.QMessageBox.information(self, "插入 fragment", "当前模型没有可复制的 fragment。")
            return
        dialog = InsertFragmentDialog(self.data_model, self)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return
        position = self._position_from_origin_and_spherical(origin, dialog.spherical_values())
        try:
            result = self.data_model.insert_fragment_copy(dialog.fragment_id(), position)
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "插入 fragment 失败", str(exc))
            return
        atom_ids = set(result["atom_ids"])
        self.data_model.select_atoms(atom_ids)
        self.last_picked_atom = min(atom_ids) if len(atom_ids) == 1 else None
        self.last_picked_bond = None
        self.refresh_scene()
        self.update_summary_panel()
        self.update_selection_panel()
        self.statusBar().showMessage(
            f"已复制 fragment {dialog.fragment_id()}，新增 {len(atom_ids)} 个原子、{len(result['bond_ids'])} 个 bond。"
        )

    def _legacy_change_selected_atom_type(self) -> None:
        if self.data_model is None or self.data_model.selection_kind != "atom" or not self.data_model.selected_atoms:
            return
        current_types = sorted({self.data_model.atoms[aid].atom_type for aid in self.data_model.selected_atoms})
        value, ok = QtWidgets.QInputDialog.getInt(
            self,
            "修改 type",
            f"当前选中 {len(self.data_model.selected_atoms)} 个原子。\n请输入新的 type:",
            current_types[0],
            1,
            10_000_000,
            1,
        )
        if not ok:
            return
        insert_existing_type = False
        selected_ids = set(self.data_model.selected_atoms)
        existing_target_atoms = {
            aid
            for aid, atom in self.data_model.atoms.items()
            if atom.atom_type == value and aid not in selected_ids
        }
        if existing_target_atoms:
            self._clear_left_drag_state()
            prompt = QtWidgets.QMessageBox(self)
            prompt.setIcon(QtWidgets.QMessageBox.Question)
            prompt.setWindowTitle("修改 type")
            prompt.setText(f"type={value} 已存在。")
            prompt.setInformativeText("请选择将当前选中原子并入已有 type，还是在该编号处插入一个新的 type。")
            merge_button = prompt.addButton("并入已有 type", QtWidgets.QMessageBox.AcceptRole)
            insert_button = prompt.addButton("插入为新 type", QtWidgets.QMessageBox.DestructiveRole)
            cancel_button = prompt.addButton(QtWidgets.QMessageBox.Cancel)
            prompt.setDefaultButton(merge_button)
            prompt.exec()
            clicked = prompt.clickedButton()
            if clicked is cancel_button:
                return
            insert_existing_type = clicked is insert_button

        self.data_model.change_selected_type(value, insert_existing_type=insert_existing_type)
        self.refresh_scene()
        self.update_selection_panel()
        self.statusBar().showMessage(f"已将选中原子 type 修改为 {value}")

    def add_shell_atoms(self) -> None:
        if self.data_model is None or self.data_model.selection_kind != "atom" or not self.data_model.selected_atoms:
            return

        selected_types = {self.data_model.atoms[aid].atom_type for aid in self.data_model.selected_atoms if aid in self.data_model.atoms}
        if len(selected_types) != 1:
            QtWidgets.QMessageBox.warning(self, "添加 Shell 原子", "一次只能对一个 type 的原子添加 shell 原子。")
            return

        original_type = next(iter(selected_types))
        if self.data_model.masses.get(original_type) is None:
            QtWidgets.QMessageBox.warning(self, "添加 Shell 原子", f"type={original_type} 未定义质量，无法生成 shell 原子。")
            return

        selected_ids = set(self.data_model.selected_atoms)
        all_type_ids = self.data_model.atom_ids_of_type(original_type)
        split_selected = False

        if selected_ids != all_type_ids:
            self._clear_left_drag_state()
            prompt = QtWidgets.QMessageBox(self)
            prompt.setIcon(QtWidgets.QMessageBox.Question)
            prompt.setWindowTitle("添加 Shell 原子")
            prompt.setText("当前只选中了该 type 的部分原子。")
            prompt.setInformativeText(
                "是否先将当前选中的原子列为一个新的 type，再只对这部分原子添加 shell？\n"
                "选择“否”则默认对当前 type 的全部原子添加 shell。"
            )
            split_button = prompt.addButton("是，拆分为新 type", QtWidgets.QMessageBox.YesRole)
            full_button = prompt.addButton("否，对全部同 type 原子添加", QtWidgets.QMessageBox.NoRole)
            cancel_button = prompt.addButton("取消", QtWidgets.QMessageBox.RejectRole)
            prompt.exec()
            clicked = prompt.clickedButton()
            if clicked == cancel_button:
                return
            split_selected = clicked == split_button
            if clicked not in (split_button, full_button):
                return

        result = self.data_model.add_shell_atoms(
            original_type,
            selected_ids,
            split_selected_atoms=split_selected,
            z_offset=0.05,
        )
        synced = self._sync_add_shell_by_type(original_type) if not split_selected else 0
        shell_ids = set(result["shell_atom_ids"])
        self.data_model.reveal_atoms(shell_ids)
        self.data_model.select_atoms(shell_ids)
        self.last_picked_atom = min(shell_ids) if len(shell_ids) == 1 else None
        self.last_picked_bond = None
        self.refresh_scene()
        self.update_selection_panel()
        bond_count = len(result["shell_bond_ids"])
        bond_type = result["shell_bond_type"]

        if split_selected:
            self.statusBar().showMessage(
                f"已将选中原子拆分为 type={result['core_type']}，并新增 {len(shell_ids)} 个 shell 原子(type={result['shell_type']})。"
            )
        else:
            self.statusBar().showMessage(
                f"已为 type={original_type} 新增 {len(shell_ids)} 个 shell 原子(type={result['shell_type']})。"
            )

        self.statusBar().showMessage(
            f"core-shell bond added: {bond_count} bonds, bond type={bond_type}; shell atom type={result['shell_type']}; 同步 {synced} 个模型"
        )

    def change_selected_type(self) -> None:
        if self.data_model is None or not self.data_model.has_selection():
            return

        kind = self.data_model.selection_kind
        if kind == "atom":
            selected_ids = set(self.data_model.selected_atoms)
            current_types = sorted({self.data_model.atoms[aid].atom_type for aid in selected_ids if aid in self.data_model.atoms})
            selected_count = len(selected_ids)
            object_name = "atom"
        elif kind in TOPOLOGY_KIND_TO_SECTION:
            selected_ids = set(self.data_model.topology_id_set(kind))
            selected_records = [record for record in self.data_model.topology_records(kind) if record.record_id in selected_ids]
            current_types = sorted({record.record_type for record in selected_records})
            selected_count = len(selected_records)
            object_name = kind
        else:
            return

        if not current_types or selected_count == 0:
            return
        source_types = set(current_types)
        can_sync_by_type = self._current_selection_covers_full_types(kind or "", source_types)

        value, ok = QtWidgets.QInputDialog.getInt(
            self,
            "修改 type",
            f"当前选中 {selected_count} 个 {object_name}。\n请输入新的 type:",
            current_types[0],
            1,
            10_000_000,
            1,
        )
        if not ok:
            return

        if kind == "atom":
            existing_targets = {
                aid
                for aid, atom in self.data_model.atoms.items()
                if atom.atom_type == value and aid not in selected_ids
            }
        else:
            existing_targets = {
                record.record_id
                for record in self.data_model.topology_records(kind)
                if record.record_type == value and record.record_id not in selected_ids
            }

        insert_existing_type = False
        if existing_targets:
            self._clear_left_drag_state()
            prompt = QtWidgets.QMessageBox(self)
            prompt.setIcon(QtWidgets.QMessageBox.Question)
            prompt.setWindowTitle("修改 type")
            prompt.setText(f"type={value} 已存在。")
            prompt.setInformativeText(
                f"请选择将当前选中的 {object_name} 并入已有 type，还是在该编号处插入一个新的 type。"
            )
            merge_button = prompt.addButton("并入已有 type", QtWidgets.QMessageBox.AcceptRole)
            insert_button = prompt.addButton("插入为新 type", QtWidgets.QMessageBox.DestructiveRole)
            cancel_button = prompt.addButton(QtWidgets.QMessageBox.Cancel)
            prompt.setDefaultButton(merge_button)
            prompt.exec()
            clicked = prompt.clickedButton()
            if clicked is cancel_button:
                return
            insert_existing_type = clicked is insert_button

        self.data_model.change_selected_type(value, insert_existing_type=insert_existing_type)
        synced = (
            self._sync_change_selected_type_by_types(kind or "", source_types, value, insert_existing_type)
            if can_sync_by_type
            else 0
        )
        self.refresh_scene()
        self.update_summary_panel()
        self.update_selection_panel()
        suffix = f"，同步 {synced} 个模型" if can_sync_by_type else "，当前选择不是完整 type，未同步"
        self.statusBar().showMessage(f"已将选中 {object_name} 的 type 修改为 {value}{suffix}")

    def copy_selected_ids(self) -> None:
        if self.data_model is None or not self.data_model.has_selection():
            return
        ids_text = " ".join(map(str, sorted(self.data_model.current_selection_ids())))
        QtWidgets.QApplication.clipboard().setText(ids_text)
        self.statusBar().showMessage(f"已复制选中的 {self.data_model.selection_kind or 'item'} ID。")

    def _handle_property_cell_double_clicked(self, row: int, column: int) -> None:
        if column != 1 or self.data_model is None or self._properties_context_atom_type is None:
            return
        name_item = self.properties_table.item(row, 0)
        if name_item is None or name_item.text() != "color":
            return

        atom_type = int(self._properties_context_atom_type)
        current_color = self._color_for_type(atom_type)
        dialog = AtomColorDialog(current_color, self)
        if dialog.exec() != QtWidgets.QDialog.Accepted:
            return

        new_color = dialog.color()
        if new_color == current_color:
            return
        self._clear_left_drag_state()
        self.data_model.set_color_for_type(atom_type, new_color)
        synced = self._sync_atom_type_property(atom_type, "color", new_color)
        self.update_properties_panel()
        self.refresh_scene()
        self.statusBar().showMessage(f"已更新 type={atom_type} 的显示颜色为 RGB {new_color}，同步 {synced} 个模型")

    def _handle_topology_type_table_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_type_tables or item.column() not in (2, 3) or self.data_model is None:
            return

        table_to_kind = {
            self.bond_type_table: "bond",
            self.angle_type_table: "angle",
            self.dihedral_type_table: "dihedral",
            self.improper_type_table: "improper",
        }
        table = self.sender()
        kind = table_to_kind.get(table)
        if kind is None:
            return
        type_item = table.item(item.row(), 0)
        if type_item is None:
            self.update_summary_panel()
            return
        try:
            type_id = int(type_item.text())
        except ValueError:
            self.update_summary_panel()
            return

        new_label = item.text().strip()
        if new_label == "未定义":
            new_label = ""
        is_name_column = item.column() == 2
        current_label = (self.data_model.coeff_names if is_name_column else self.data_model.coeff_labels).get(kind, {}).get(type_id, "").strip()
        if new_label == current_label:
            return

        self._clear_left_drag_state()
        reply = QtWidgets.QMessageBox.question(
            self,
            "修改 Name" if is_name_column else "修改 ff-type",
            f"修改 {kind} {'Name' if is_name_column else 'ff-type'} 将同时更新全部 type={type_id} 的 {kind}，是否继续？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            self.update_summary_panel()
            return

        if is_name_column:
            self.data_model.set_coeff_name_for_type(kind, type_id, new_label)
            synced = self._sync_topology_type_property(kind, type_id, "name", new_label)
        else:
            self.data_model.set_coeff_label_for_type(kind, type_id, new_label)
            synced = self._sync_topology_type_property(kind, type_id, "ff", new_label)
        self.update_summary_panel()
        self.statusBar().showMessage(f"已更新 {kind} type={type_id} 的 {'Name' if is_name_column else 'ff-type'}，同步 {synced} 个模型")

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:  # noqa: N802
        if self.data_model is None or not self.data_model.dirty:
            event.accept()
            self._unregister_window()
            return

        reply = QtWidgets.QMessageBox.question(
            self,
            "未保存修改",
            "当前修改尚未保存。关闭前是否另存为？",
            QtWidgets.QMessageBox.Save | QtWidgets.QMessageBox.Discard | QtWidgets.QMessageBox.Cancel,
            QtWidgets.QMessageBox.Save,
        )
        if reply == QtWidgets.QMessageBox.Save:
            if self.save_as_dialog():
                event.accept()
                self._unregister_window()
            else:
                event.ignore()
            return
        if reply == QtWidgets.QMessageBox.Discard:
            event.accept()
            self._unregister_window()
            return
        event.ignore()

    def _unregister_window(self) -> None:
        global SYNC_GROUP_WINDOWS
        if self in OPEN_VIEWER_WINDOWS:
            OPEN_VIEWER_WINDOWS.remove(self)
        if self in SYNC_GROUP_WINDOWS:
            SYNC_GROUP_WINDOWS = [window for window in SYNC_GROUP_WINDOWS if window is not self]

    def reset_camera(self) -> None:
        if self.plotter is None:
            return
        self.plotter.reset_camera()
        self.plotter.render()


def infer_element_from_mass(mass: Optional[float]) -> Optional[Tuple[str, str]]:
    if mass is None or mass <= 0:
        return None
    best_symbol = ""
    best_name = ""
    best_mass = 0.0
    best_delta = float("inf")
    for symbol, name, reference_mass in ELEMENT_MASS_TABLE:
        delta = abs(float(mass) - reference_mass)
        if delta < best_delta:
            best_symbol = symbol
            best_name = name
            best_mass = reference_mass
            best_delta = delta
    tolerance = max(0.2, best_mass * 0.015)
    if best_delta <= tolerance:
        return best_symbol, best_name
    return None


def hsv_to_rgb(h: float, s: float, v: float) -> Tuple[float, float, float]:
    if s == 0.0:
        return v, v, v
    i = int(h * 6.0)
    f = (h * 6.0) - i
    p = v * (1.0 - s)
    q = v * (1.0 - s * f)
    t = v * (1.0 - s * (1.0 - f))
    i %= 6
    if i == 0:
        return v, t, p
    if i == 1:
        return q, v, p
    if i == 2:
        return p, v, t
    if i == 3:
        return p, q, v
    if i == 4:
        return t, p, v
    return v, p, q


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="LAMMPS data viewer/editor")
    parser.add_argument("file", nargs="?", help="待打开的 LAMMPS data 文件")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    app = QtWidgets.QApplication(sys.argv)
    window = ViewerMainWindow(Path(args.file) if args.file else None)
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
